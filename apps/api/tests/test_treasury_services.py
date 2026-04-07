"""
Tests for treasury_parser and treasury_template services.
These test the shared logic independently from the API routes.
"""

import json
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.services.treasury_parser import (
    TreasuryMovement,
    TreasuryStatement,
    _detect_bank,
    _extract_counterparty,
    _extract_reference,
    _movement_category,
    _movement_type,
    _normalize,
    _normalize_spaces,
    _parse_bbva,
    _parse_banregio,
    _parse_bajio,
    _parse_santander,
    _parse_date,
    _parse_money,
    analysis_payload,
    statement_from_payload,
    statements_from_analysis_json,
)
from app.services.treasury_template import (
    _movement_amount,
    _movement_signature,
    _sheet_bank_hint,
    _sheet_kind,
    _statement_kind,
    prepare_balance_template,
    prepare_movement_template,
    render_balance_workbook,
    render_movement_workbook,
)


BBVA_SAMPLE = """
BBVA
Fecha de consulta
03/02/2026 10:06:58 AM
No. Contrato
00610828
Nombre del Cliente
CONSTRUCCIONES LOROS SA DE CV
Detalle de Movimientos
Fecha
Concepto/ Referencia
Cargo
Abono
Saldo
31/01/2026
PAGO DE NOMINA / BC 4203836649 CONSTRUCCIONES LOROS SA DE CV
-63,574.84
46,670.52
29/01/2026
DEPOSITO DE TERCERO / REFBNTC00642703 SALDO DE FACTURA BMRCASH 893,582.12 4,833,729.46
"""

BAJIO_SAMPLE = """
CUENTA CONECTA BANBAJIO
Cuenta: 339222200201
Saldo Total: $23,166.52
1
12-Dic-2025
SPEI Recibido:
Institucion contraparte: BBVA MEXICO Ordenante: CARRETERAS Y EDIFICACIONES DE MICHOACAN
Referencia: 121225
$4,766.66
$23,166.52
Hora: 17:58:30
"""


class TestTreasuryParser(unittest.TestCase):

    def test_detect_bank_bbva(self):
        self.assertEqual(_detect_bank("BBVA Net Cash reporte"), "BBVA")

    def test_detect_bank_banregio(self):
        self.assertEqual(_detect_bank("banregio estado de cuenta"), "Banregio")

    def test_detect_bank_banbajio(self):
        self.assertEqual(_detect_bank("cuenta conecta banbajio"), "BanBajio")

    def test_detect_bank_monex(self):
        self.assertEqual(_detect_bank("Sistema Corporativo Monex"), "Monex")

    def test_detect_bank_santander(self):
        self.assertEqual(_detect_bank("Contrato CMC: 12345"), "Santander")

    def test_detect_bank_unknown(self):
        self.assertEqual(_detect_bank("some random text"), "Desconocido")

    def test_parse_money(self):
        self.assertAlmostEqual(_parse_money("$1,234.56"), 1234.56)
        self.assertAlmostEqual(_parse_money("-$500.00"), -500.00)
        self.assertIsNone(_parse_money(None))
        self.assertIsNone(_parse_money(""))

    def test_parse_money_ocr_corrections(self):
        self.assertAlmostEqual(_parse_money("1,O00.00"), 1000.00)

    def test_parse_date_formats(self):
        self.assertEqual(_parse_date("15/03/2026"), "2026-03-15")
        self.assertEqual(_parse_date("2026-03-15"), "2026-03-15")
        self.assertEqual(_parse_date("15-Mar-2026"), "2026-03-15")
        self.assertIsNone(_parse_date(None))
        self.assertIsNone(_parse_date(""))

    def test_parse_date_spanish_months(self):
        self.assertEqual(_parse_date("12-Dic-2025"), "2025-12-12")
        self.assertEqual(_parse_date("01-ENE-2026"), "2026-01-01")

    def test_normalize(self):
        self.assertEqual(_normalize("  Café  Niño  "), "cafe nino")

    def test_normalize_spaces(self):
        self.assertEqual(_normalize_spaces("  hello   world  "), "hello world")
        self.assertIsNone(_normalize_spaces(None))

    def test_movement_type(self):
        self.assertEqual(_movement_type(100.0, None), "cargo")
        self.assertEqual(_movement_type(None, 200.0), "abono")
        self.assertEqual(_movement_type(None, None), "informativo")

    def test_movement_category_iva_comision(self):
        self.assertEqual(_movement_category("IVA por comision", None, 12.0, None), "iva_comision")

    def test_movement_category_nomina(self):
        self.assertEqual(_movement_category("PAGO DE NOMINA", None, 5000.0, None), "nomina")

    def test_movement_category_spei_recibido(self):
        self.assertEqual(_movement_category("SPEI Recibido", None, None, 1000.0), "transferencia_entrada")

    def test_extract_counterparty(self):
        self.assertEqual(
            _extract_counterparty("Ordenante: CARRETERAS Y EDIFICACIONES DE MICHOACAN Cuenta: 123"),
            "CARRETERAS Y EDIFICACIONES DE MICHOACAN",
        )
        self.assertIsNone(_extract_counterparty(None))

    def test_extract_reference(self):
        self.assertEqual(_extract_reference("CH-12345 pago"), "CH-12345")
        self.assertIsNone(_extract_reference(None))

    def test_parse_bbva(self):
        statement = _parse_bbva(BBVA_SAMPLE, "bbva.pdf", ocr_used=False)
        self.assertEqual(statement.bank, "BBVA")
        self.assertEqual(statement.account_holder, "CONSTRUCCIONES LOROS SA DE CV")
        self.assertEqual(len(statement.movements), 2)
        self.assertEqual(statement.movements[0].movement_type, "cargo")
        self.assertAlmostEqual(statement.movements[0].debit, 63574.84)

    def test_parse_bajio(self):
        statement = _parse_bajio(BAJIO_SAMPLE, "bajio.pdf", ocr_used=False)
        self.assertEqual(statement.bank, "BanBajio")
        self.assertEqual(len(statement.movements), 1)
        self.assertEqual(statement.movements[0].counterparty, "CARRETERAS Y EDIFICACIONES DE MICHOACAN")

    def test_analysis_payload_roundtrip(self):
        statement = TreasuryStatement(
            id="test-1", source_file="test.pdf", bank="BBVA", ocr_used=False,
            account_number="12345", closing_balance=1000.0, raw_text="test",
            movements=[
                TreasuryMovement(
                    statement_id="test-1", source_file="test.pdf", bank="BBVA",
                    sequence=1, movement_date="2026-01-15", description="DEPOSITO",
                    credit=500.0, balance=1000.0,
                )
            ],
        )
        payload = analysis_payload([statement])
        self.assertEqual(payload["summary"]["statements"], 1)
        self.assertEqual(payload["summary"]["movements"], 1)

        rebuilt = statements_from_analysis_json(json.dumps(payload))
        self.assertIsNotNone(rebuilt)
        self.assertEqual(len(rebuilt), 1)
        self.assertEqual(rebuilt[0].movements[0].credit, 500.0)

    def test_statements_from_analysis_json_empty(self):
        self.assertIsNone(statements_from_analysis_json(None))
        self.assertIsNone(statements_from_analysis_json(""))
        self.assertIsNone(statements_from_analysis_json("  "))


class TestTreasuryTemplate(unittest.TestCase):

    def test_sheet_bank_hint(self):
        self.assertEqual(_sheet_bank_hint("BBVA 3599"), "BBVA")
        self.assertEqual(_sheet_bank_hint("Banregio MXN"), "Banregio")
        self.assertEqual(_sheet_bank_hint("Monex USD"), "Monex")
        self.assertIsNone(_sheet_bank_hint("Hoja 1"))

    def test_sheet_kind(self):
        self.assertEqual(_sheet_kind("BBVA DLL"), "usd")
        self.assertEqual(_sheet_kind("Inversión Monex"), "investment")
        self.assertEqual(_sheet_kind("BBVA 3599"), "operational")

    def test_movement_amount(self):
        m = TreasuryMovement(statement_id="x", source_file="x", bank="x", sequence=1, credit=100.0)
        self.assertEqual(_movement_amount(m), 100.0)
        m2 = TreasuryMovement(statement_id="x", source_file="x", bank="x", sequence=1, debit=50.0)
        self.assertEqual(_movement_amount(m2), -50.0)

    def test_prepare_balance_template(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "saldos.xlsx"
            wb = Workbook()
            ws = wb.active
            ws["D4"] = "BANCO"
            ws["E4"] = "CUENTA"
            ws["H4"] = "PESOS"
            ws["I4"] = "DOLARES"
            ws["D5"] = "BBVA"
            ws["E5"] = "CHEQUES 0197083599"
            ws["H5"] = 0
            wb.save(path)

            statement = TreasuryStatement(
                id="bbva-3599", source_file="bbva.pdf", bank="BBVA", ocr_used=False,
                account_number="0197083599", closing_balance=50000.0, raw_text="BBVA",
            )
            result = prepare_balance_template(path, [statement])
            self.assertEqual(len(result["updates"]), 1)
            self.assertEqual(result["updates"][0]["new_value"], 50000.0)

    def test_render_balance_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "saldos.xlsx"
            wb = Workbook()
            ws = wb.active
            ws["D5"] = "BBVA"
            ws["E5"] = "CHEQUES 3599"
            ws["H5"] = 0
            wb.save(path)

            updates = [{"row_number": 5, "column_key": "pesos", "new_value": 12345.67, "enabled": True}]
            result_bytes = render_balance_workbook(path, updates)

            from openpyxl import load_workbook
            out_path = Path(tmpdir) / "out.xlsx"
            out_path.write_bytes(result_bytes)
            wb2 = load_workbook(out_path)
            self.assertEqual(wb2.active["H5"].value, 12345.67)

    def test_prepare_movement_template(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "movimientos.xlsx"
            self._build_movements_workbook(path)

            statement = TreasuryStatement(
                id="bbva-3599", source_file="bbva.pdf", bank="BBVA", ocr_used=False,
                account_number="0197083599", account_holder="DEESA",
                closing_balance=935.60, raw_text="BBVA",
                movements=[
                    TreasuryMovement(
                        statement_id="bbva-3599", source_file="bbva.pdf", bank="BBVA",
                        sequence=1, account_number="0197083599",
                        movement_date="2026-03-21", description="IVA COM SERVICIOS BNTC",
                        debit=66.40, balance=935.60,
                    )
                ],
            )
            result = prepare_movement_template(path, [statement])
            self.assertEqual(len(result["drafts"]), 1)
            self.assertEqual(result["drafts"][0]["sheet_name"], "BBVA 3599")

    def _build_movements_workbook(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "BBVA 3599"
        headers = [
            "TIPO DE MOVIMIENTO", "FECHA", "EMPRESA", "CAJA", "N° CHQ",
            "A NOMBRE DE", "GRUPO", "UNIDAD DE NEGOCIO", "OBRA",
            "CONCILIACION", "CONCEPTO DETALLADO", "DEPOSITOS", "RETIROS",
            "DESGLOSE", "SALDO", "OBSERVACIONES",
        ]
        for idx, value in enumerate(headers, start=1):
            ws.cell(4, idx).value = value

        ws["A6"] = "TRANSFERENCIA"
        ws["B6"] = datetime(2026, 3, 20)
        ws["C6"] = "DEESA"
        ws["F6"] = "BBVA"
        ws["G6"] = "OF CENTRAL"
        ws["J6"] = "COMISIONES BANCARIAS"
        ws["K6"] = "IVA COM SERVICIOS BNTC"
        ws["M6"] = 66.40

        table = Table(displayName="Tabla15", ref="A4:P10")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(table)
        wb.save(path)


if __name__ == "__main__":
    unittest.main()
