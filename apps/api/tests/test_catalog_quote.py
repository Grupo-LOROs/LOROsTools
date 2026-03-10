import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from app.catalog_quote.models import ContainerOffer, PriceTier, Product
from app.routes.catalog_quote import (
    PricedLine,
    _choose_tier,
    _compute_totals,
    _price_line,
    _render_quote_pdf,
    _render_quote_xlsx,
)


class CatalogQuoteTests(unittest.TestCase):
    def test_choose_tier_keeps_original_behavior_below_minimum(self):
        product = Product(sku="SKU-1", description="Producto demo", unit="PZA", price_list_id=1)
        product.tiers = [
            PriceTier(price_list_id=1, product_id=1, min_qty=20, label="+20 piezas", unit_price=239.0),
            PriceTier(price_list_id=1, product_id=1, min_qty=100, label="+100 piezas", unit_price=229.0),
        ]

        tier = _choose_tier(product, 1)

        self.assertEqual(tier.min_qty, 20)
        self.assertEqual(tier.label, "+20 piezas")

    def test_price_line_supports_container_products(self):
        product = Product(sku="SKU-CONT", description="Producto contenedor", unit="PZA", price_list_id=1)
        product.tiers = []
        product.container_offer = ContainerOffer(
            price_list_id=1,
            product_id=1,
            container_qty=540,
            container_price=55000.0,
            notes="Lead time 45 days",
        )

        priced = _price_line(product, 2, "CONTENEDOR_POR_CONTENEDOR")

        self.assertEqual(priced.unit, "CONT")
        self.assertEqual(priced.quantity, 2)
        self.assertEqual(priced.unit_price, 55000.0)
        self.assertEqual(priced.line_total, 110000.0)
        self.assertEqual(priced.rule, "Precio por contenedor")

    def test_compute_totals_breaks_out_iva_when_included(self):
        base, iva, total = _compute_totals(239.0, "included", 0.16)

        self.assertEqual(base, 206.03)
        self.assertEqual(iva, 32.97)
        self.assertEqual(total, 239.0)

    def test_render_quote_xlsx_writes_headers_lines_and_terms(self):
        lines = [
            PricedLine(
                sku="SKU-1",
                description="Producto uno",
                unit="PZA",
                quantity=3,
                unit_price=239.0,
                line_total=717.0,
                rule="Tier +20 piezas (min 20)",
            ),
            PricedLine(
                sku="SKU-2",
                description="Producto dos",
                unit="PZA",
                quantity=1,
                unit_price=150.5,
                line_total=150.5,
                rule="Tier +1 pieza (min 1)",
            ),
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            output = Path(tmpdir) / "quote.xlsx"
            _render_quote_xlsx(output, "Morelia", date(2026, 3, 10), lines, ["Cliente: Demo", "Precios con IVA"])

            wb = load_workbook(output)
            ws = wb.active

            self.assertEqual(ws["F15"].value, "Morelia")
            self.assertEqual(ws["F16"].value, "10/03/2026")
            self.assertEqual(ws["A19"].value, "SKU-1")
            self.assertEqual(ws["D19"].value, 3)
            self.assertEqual(ws["E19"].value, 239.0)
            self.assertEqual(ws["F19"].value, "=D19*E19")
            self.assertEqual(ws["A20"].value, "SKU-2")
            term_values = [ws[f"B{row}"].value for row in range(26, ws.max_row + 1)]
            self.assertIn("Cliente: Demo", term_values)
            self.assertIn("Precios con IVA", term_values)

    def test_render_quote_pdf_creates_non_empty_file(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output = Path(tmpdir) / "quote.pdf"
            _render_quote_pdf(
                pdf_path=output,
                folio="ERA-000123",
                city="Morelia",
                quote_date=date(2026, 3, 10),
                lines=[
                    PricedLine(
                        sku="SKU-1",
                        description="Producto uno",
                        unit="PZA",
                        quantity=3,
                        unit_price=239.0,
                        line_total=717.0,
                        rule="Tier +20 piezas (min 20)",
                    )
                ],
                subtotal=617.0,
                iva=100.0,
                total=717.0,
                vendor_name="Vendedor demo",
                customer_name="Cliente demo",
                terms=["Precios con IVA"],
            )

            self.assertTrue(output.exists())
            self.assertGreater(output.stat().st_size, 1000)


if __name__ == "__main__":
    unittest.main()
