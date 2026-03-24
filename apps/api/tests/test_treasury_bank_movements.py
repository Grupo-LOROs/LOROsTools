import tempfile
import unittest
from datetime import datetime
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.routes.treasury_bank_movements import (
    TreasuryMovement,
    TreasuryStatement,
    _analysis_payload,
    _detect_bank,
    _extract_counterparty,
    _movement_category,
    _parse_bajio,
    _parse_banregio,
    _parse_bbva,
    _prepare_balance_template,
    _prepare_movement_template,
    _parse_santander,
    _render_balance_workbook,
    _render_movement_workbook,
    _statements_from_analysis_json,
)


BBVA_SAMPLE = """
BBVA
Fecha de consulta
03/02/2026 10:06:58 AM
No. Contrato
00610828
Nombre del Cliente
CONSTRUCCIONES LOROS SA DE CV
Detalle de Movimientos
Fecha
Concepto/ Referencia
Cargo
Abono
Saldo
31/01/2026
PAGO DE NOMINA / BC 4203836649 CONSTRUCCIONES LOROS SA DE CV
-63,574.84
46,670.52
29/01/2026
DEPOSITO DE TERCERO / REFBNTC00642703 SALDO DE FACTURA BMRCASH 893,582.12 4,833,729.46
"""

BANREGIO_SAMPLE = """
banregio
Detalle de Movimientos Cuenta de Cheques
CONSTRUCCIONES LOROS S.A. DE C.V.
CUENTA: 210994250013
CLABE: 058470000000903792
Fecha Inicio:01/02/2026 - Fecha Fin: 28/02/2026
Saldo Inicial: $27635.23
Fecha
Descripción
Referencia
Cargos
Abonos
Saldo
_Sin registros
Total Cargos: $0.00
Total Abonos: $0.00
Saldo Final: $27635.23
"""

BAJIO_SAMPLE = """
CUENTA CONECTA BANBAJIO
Cuenta: 339222200201
Saldo Total: $23,166.52
#
Fecha Movimiento
Descripción
Cargos
Abonos
Saldo
1
12-Dic-2025
SPEI Recibido:
Institucion contraparte: BBVA MEXICO Ordenante: CARRETERAS Y EDIFICACIONES DE MICHOACAN
Referencia: 121225
$4,766.66
$23,166.52
Hora: 17:58:30
Clave de Rastreo: 002601002512150000887834 Concepto del Pago: PAGO
"""

SANTANDER_SAMPLE = """
Consulta de Movimientos de la Cuenta de Cheques
Contrato CMC: 80152454770 ENERGIA RENOVABLE DE AMERICA SA DE CV
Número de Cuenta: 65511207359
Periodo: 01/10/2025 al 26/12/2025
Saldo Inicial: $1,912.15 MXN
Saldo Final: $15,000.00 MXN
Importe Total Abonos: $3,117,598.44 MXN
Importe Total Cargos: $3,122,798.44 MXN
Cuenta
Fecha
Hora
Sucursal
Descripción
Importe Cargo
Importe Abono
Saldo
Referencia
Concepto
Descripción Larga
65511207359
03112
025
17:28
7465
ABONO
TRANSFERENCIA
SPEI
0.00
40,390.40
40,390.40
005240091
TRASPASO BBVA A SANTANDER
012470001692707448
ABONO TRANSFERENCIA SPEI
65511207359
03112
025
17:28
7465
I V A  POR
COMISION
400.00
0.00
37,490.40
OCT 2025
I V A  POR COMISION
"""


class TreasuryBankMovementsTests(unittest.TestCase):
    def test_detect_bank(self):
        self.assertEqual(_detect_bank(BBVA_SAMPLE), "BBVA")
        self.assertEqual(_detect_bank(BANREGIO_SAMPLE), "Banregio")
        self.assertEqual(_detect_bank(BAJIO_SAMPLE), "BanBajio")
        self.assertEqual(_detect_bank(SANTANDER_SAMPLE), "Santander")

    def test_parse_bbva_credit_and_debit(self):
        statement = _parse_bbva(BBVA_SAMPLE, "bbva.pdf", ocr_used=True)
        self.assertEqual(statement.bank, "BBVA")
        self.assertEqual(len(statement.movements), 2)
        self.assertEqual(statement.movements[0].movement_type, "cargo")
        self.assertEqual(statement.movements[1].movement_type, "abono")
        self.assertAlmostEqual(statement.movements[1].credit or 0, 893582.12)

    def test_parse_banregio_without_rows(self):
        statement = _parse_banregio(BANREGIO_SAMPLE, "banregio.pdf", ocr_used=True)
        self.assertEqual(statement.account_number, "210994250013")
        self.assertEqual(len(statement.movements), 0)
        self.assertTrue(any("no contiene movimientos" in warning.lower() for warning in statement.warnings))

    def test_parse_bajio_transfer_in(self):
        statement = _parse_bajio(BAJIO_SAMPLE, "bajio.pdf", ocr_used=True)
        self.assertEqual(len(statement.movements), 1)
        movement = statement.movements[0]
        self.assertEqual(movement.movement_type, "abono")
        self.assertEqual(movement.category, "transferencia_entrada")
        self.assertEqual(movement.counterparty, "CARRETERAS Y EDIFICACIONES DE MICHOACAN")

    def test_parse_santander_rows(self):
        statement = _parse_santander(SANTANDER_SAMPLE, "santander.pdf", ocr_used=False)
        self.assertEqual(len(statement.movements), 2)
        self.assertEqual(statement.movements[0].credit, 40390.40)
        self.assertEqual(statement.movements[1].category, "iva_comision")

    def test_counterparty_and_category_helpers(self):
        self.assertEqual(
            _extract_counterparty("Ordenante: CONSTRUCCIONES LOROS SA DE CV | Cuenta ordenante: 123"),
            "CONSTRUCCIONES LOROS SA DE CV",
        )
        self.assertEqual(_movement_category("IVA por comision", None, 12.0, None), "iva_comision")

    def test_prepare_movement_template_reuses_history_from_matching_sheet(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "movimientos.xlsx"
            self._build_movements_workbook(workbook_path)

            statement = TreasuryStatement(
                id="bbva-3599",
                source_file="bbva-3599.pdf",
                bank="BBVA",
                ocr_used=False,
                account_number="0197083599",
                account_holder="DEESA",
                closing_balance=935.60,
                raw_text="IVA COM SERVICIOS BNTC",
                movements=[
                    TreasuryMovement(
                        statement_id="bbva-3599",
                        source_file="bbva-3599.pdf",
                        bank="BBVA",
                        sequence=1,
                        account_number="0197083599",
                        movement_date="2026-03-21",
                        description="IVA COM SERVICIOS BNTC / 00474096",
                        debit=66.40,
                        balance=935.60,
                    )
                ],
            )

            prepared = _prepare_movement_template(workbook_path, [statement])

            self.assertEqual(len(prepared["drafts"]), 1)
            draft = prepared["drafts"][0]
            self.assertEqual(draft["sheet_name"], "BBVA 3599")
            self.assertEqual(draft["values"]["reconciliation"], "COMISIONES BANCARIAS")
            self.assertEqual(draft["values"]["group"], "OF CENTRAL")
            self.assertEqual(draft["suggestion_source"], "historial")

    def test_render_movement_workbook_writes_on_next_available_row(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "movimientos.xlsx"
            self._build_movements_workbook(workbook_path)

            statement = TreasuryStatement(
                id="bbva-3599",
                source_file="bbva-3599.pdf",
                bank="BBVA",
                ocr_used=False,
                account_number="0197083599",
                account_holder="DEESA",
                closing_balance=935.60,
                raw_text="IVA COM SERVICIOS BNTC",
                movements=[
                    TreasuryMovement(
                        statement_id="bbva-3599",
                        source_file="bbva-3599.pdf",
                        bank="BBVA",
                        sequence=1,
                        account_number="0197083599",
                        movement_date="2026-03-21",
                        description="IVA COM SERVICIOS BNTC / 00474096",
                        debit=66.40,
                        balance=935.60,
                    )
                ],
            )
            prepared = _prepare_movement_template(workbook_path, [statement])

            rendered = _render_movement_workbook(workbook_path, prepared["drafts"])
            output_path = Path(tmpdir) / "movimientos_actualizados.xlsx"
            output_path.write_bytes(rendered)

            wb = load_workbook(output_path)
            ws = wb["BBVA 3599"]

            self.assertEqual(ws["A7"].value, "TRANSFERENCIA")
            self.assertEqual(ws["F7"].value, "BBVA")
            self.assertEqual(ws["J7"].value, "COMISIONES BANCARIAS")
            self.assertEqual(ws["K7"].value, "IVA COM SERVICIOS BNTC")
            self.assertEqual(ws["M7"].value, 66.40)
            self.assertTrue(str(ws["O7"].value).startswith("="))

    def test_prepare_and_render_balance_template_updates_matching_row(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "saldos.xlsx"
            self._build_balance_workbook(workbook_path)

            statement = TreasuryStatement(
                id="bbva-3599",
                source_file="bbva-3599.pdf",
                bank="BBVA",
                ocr_used=False,
                account_number="0197083599",
                account_holder="DEESA",
                closing_balance=183871.45,
                raw_text="BBVA",
            )

            prepared = _prepare_balance_template(workbook_path, [statement])

            self.assertEqual(len(prepared["updates"]), 1)
            update = prepared["updates"][0]
            self.assertEqual(update["row_number"], 5)
            self.assertEqual(update["column_key"], "pesos")
            self.assertEqual(update["new_value"], 183871.45)

            rendered = _render_balance_workbook(workbook_path, prepared["updates"])
            output_path = Path(tmpdir) / "saldos_actualizados.xlsx"
            output_path.write_bytes(rendered)

            wb = load_workbook(output_path, data_only=False)
            ws = wb.active
            self.assertEqual(ws["H5"].value, 183871.45)

    def test_analysis_payload_can_be_reused_without_reparsing_pdfs(self):
        statement = TreasuryStatement(
            id="bbva-3599",
            source_file="bbva-3599.pdf",
            bank="BBVA",
            ocr_used=False,
            account_number="0197083599",
            account_holder="DEESA",
            closing_balance=183871.45,
            raw_text="BBVA",
            movements=[
                TreasuryMovement(
                    statement_id="bbva-3599",
                    source_file="bbva-3599.pdf",
                    bank="BBVA",
                    sequence=1,
                    account_number="0197083599",
                    movement_date="2026-03-21",
                    description="DEPOSITO",
                    credit=1500.0,
                    balance=183871.45,
                )
            ],
        )

        payload = _analysis_payload([statement])
        rebuilt = _statements_from_analysis_json(json.dumps(payload))

        self.assertIsNotNone(rebuilt)
        self.assertEqual(len(rebuilt or []), 1)
        self.assertEqual((rebuilt or [])[0].source_file, "bbva-3599.pdf")
        self.assertEqual(len((rebuilt or [])[0].movements), 1)
        self.assertEqual((rebuilt or [])[0].movements[0].credit, 1500.0)

    def _build_movements_workbook(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "BBVA 3599"
        headers = [
            "TIPO DE MOVIMIENTO",
            "FECHA",
            "EMPRESA",
            "CAJA",
            "N° CHQ",
            "A NOMBRE DE",
            "GRUPO",
            "UNIDAD DE NEGOCIO",
            "OBRA",
            "CONCILIACION",
            "CONCEPTO DETALLADO",
            "DEPOSITOS",
            "RETIROS",
            "DESGLOSE",
            "SALDO",
            "OBSERVACIONES",
        ]
        for idx, value in enumerate(headers, start=1):
            ws.cell(4, idx).value = value

        ws["B5"] = datetime(2026, 3, 20)
        ws["F5"] = "SALDO INICIAL"
        ws["O5"] = 1000

        ws["A6"] = "TRANSFERENCIA"
        ws["B6"] = datetime(2026, 3, 20)
        ws["C6"] = "DEESA"
        ws["F6"] = "BBVA"
        ws["G6"] = "OF CENTRAL"
        ws["H6"] = "OF CENTRAL"
        ws["I6"] = "CORPORATIVO"
        ws["J6"] = "COMISIONES BANCARIAS"
        ws["K6"] = "IVA COM SERVICIOS BNTC"
        ws["M6"] = 66.40
        ws["O6"] = "=O5+L6-M6"

        table = Table(displayName="Tabla15", ref="A4:P10")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
        wb.save(path)

    def _build_balance_workbook(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "23-03-26"
        ws["D4"] = "BANCO"
        ws["E4"] = "DEESA"
        ws["H4"] = "SALDO B PESOS"
        ws["I4"] = "SALDO B DOLARES"
        ws["D5"] = "BBVA"
        ws["E5"] = "CHEQUES 0197083599"
        ws["H5"] = 0
        wb.save(path)


if __name__ == "__main__":
    unittest.main()
