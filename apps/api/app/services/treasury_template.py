"""
Treasury Template Processing
─────────────────────────────
Matching and rendering logic for movement and balance Excel templates.

This module is framework-agnostic: no FastAPI dependency.
Both the API routes and the worker import from here.
"""

from __future__ import annotations

import copy
import io
import re
import unicodedata
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import range_boundaries

from .treasury_parser import (
    CATEGORY_TO_RECONCILIATION,
    TREASURY_STOPWORDS,
    TreasuryMovement,
    TreasuryStatement,
    _normalize,
    _normalize_spaces,
    _parse_date,
    _parse_money,
)


# ── Constants ─────────────────────────────────────────────

MOVEMENT_FIELD_LABELS = {
    "sheet_name": "Hoja destino",
    "movement_type": "Tipo de movimiento",
    "company": "Empresa",
    "payee": "A nombre de",
    "group": "Grupo",
    "business_unit": "Unidad de negocio",
    "project": "Obra",
    "reconciliation": "Conciliación",
    "specific_concept": "Concepto específico",
    "detailed_concept": "Concepto detallado",
    "observations": "Observaciones",
}

MOVEMENT_FIELD_ALIASES = {
    "movement_type": (
        "tipo de movimiento",
        "tipo de moviento",
    ),
    "date": ("fecha",),
    "company": ("empresa",),
    "cashbox": ("caja",),
    "check_number": ("n chq", "n chq.", "no cheque", "no. cheque"),
    "payee": ("a nombre de",),
    "group": ("grupo",),
    "business_unit": ("unidad de negocio",),
    "project": ("obra",),
    "reconciliation": ("conciliacion",),
    "specific_concept": ("concepto especifico",),
    "detailed_concept": ("concepto detallado",),
    "deposits": ("depositos",),
    "withdrawals": ("retiros",),
    "breakdown": ("desglose", "desglose de retiros"),
    "balance": ("saldo",),
    "observations": ("observaciones", "columna1"),
}

MOVEMENT_EDITABLE_FIELDS = (
    "movement_type",
    "company",
    "payee",
    "group",
    "business_unit",
    "project",
    "reconciliation",
    "specific_concept",
    "detailed_concept",
    "observations",
)

DEFAULT_MOVEMENT_TYPES = ("TRANSFERENCIA", "INVERSIÓN", "DEPÓSITO", "CHEQUE", "CARGO", "ABONO")


# ── Helpers ───────────────────────────────────────────────

def _as_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return _normalize_spaces(str(value))


def _iso_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return _parse_date(_as_text(value))


def _float_value(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    return _parse_money(str(value))


def _header_name(value: Any) -> str:
    text = _normalize_spaces(str(value or "")) or ""
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[°º#./_()-]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def _canonical_field_name(value: Any) -> str | None:
    normalized = _header_name(value)
    for field_name, aliases in MOVEMENT_FIELD_ALIASES.items():
        for alias in aliases:
            if normalized == alias:
                return field_name
    return None


def _extract_digit_tokens(*values: str | None) -> list[str]:
    tokens: list[str] = []
    for value in values:
        raw = re.sub(r"\D", "", value or "")
        if len(raw) >= 4:
            tokens.append(raw)
            tokens.append(raw[-4:])
            tokens.append(raw[-6:])
    seen: set[str] = set()
    ordered: list[str] = []
    for token in tokens:
        if token and token not in seen:
            seen.add(token)
            ordered.append(token)
    return ordered


def _search_tokens(*values: str | None) -> set[str]:
    text = " ".join(part for part in values if part)
    normalized = _normalize(text)
    tokens = set()
    for token in re.split(r"[^a-z0-9]+", normalized):
        if len(token) < 3 or token in TREASURY_STOPWORDS:
            continue
        tokens.add(token)
    return tokens


def _statement_label(statement: TreasuryStatement) -> str:
    return statement.account_number or statement.contract or statement.source_file


def _movement_amount(movement: TreasuryMovement) -> float:
    if movement.credit is not None:
        return round(float(movement.credit), 2)
    if movement.debit is not None:
        return round(float(-movement.debit), 2)
    return 0.0


def _source_movement_signature(
    *,
    movement_date: str | None,
    amount: float | None,
    detail: str | None,
) -> str:
    detail_norm = _normalize_spaces(detail) or ""
    amount_label = f"{round(float(amount or 0.0), 2):.2f}"
    return "|".join([movement_date or "", amount_label, _normalize(detail_norm)])


def _movement_signature(movement: TreasuryMovement) -> str:
    detail = " ".join(part for part in [movement.description, movement.concept, movement.long_description] if part)
    return _source_movement_signature(
        movement_date=movement.movement_date,
        amount=_movement_amount(movement),
        detail=detail,
    )


# ── Sheet analysis ────────────────────────────────────────

def _sheet_bank_hint(title: str) -> str | None:
    normalized = _normalize(title)
    if "bbva" in normalized:
        return "BBVA"
    if "banreg" in normalized:
        return "Banregio"
    if "bajio" in normalized:
        return "BanBajio"
    if "monex" in normalized:
        return "Monex"
    if "santander" in normalized:
        return "Santander"
    if "scotia" in normalized:
        return "Scotia"
    return None


def _sheet_kind(title: str) -> str:
    normalized = _normalize(title)
    if any(token in normalized for token in ("dll", "usd", "dolares", "dolares", "usd")):
        return "usd"
    if any(token in normalized for token in ("inv", "inver", "inversion", "fondo")):
        return "investment"
    return "operational"


def _statement_kind(statement: TreasuryStatement) -> str:
    normalized = _normalize(
        " ".join(
            part
            for part in [
                statement.source_file,
                statement.currency,
                statement.account_number,
                statement.contract,
                statement.raw_text[:1800],
            ]
            if part
        )
    )
    if any(token in normalized for token in ("usd", "dll", "dolar", "dolares", "dolares")):
        return "usd"
    if any(token in normalized for token in ("fondo", "inversion", "inver", "mesa dinero", "bursatil")):
        return "investment"
    return "operational"


def _movement_type_default(movement: TreasuryMovement, sheet_name: str | None = None) -> str:
    if sheet_name and _sheet_kind(sheet_name) == "investment":
        return "INVERSIÓN"
    if movement.category == "deposito":
        return "DEPÓSITO"
    if movement.category == "cheque":
        return "CHEQUE"
    return "TRANSFERENCIA"


def _default_reconciliation(movement: TreasuryMovement) -> str | None:
    if movement.category in CATEGORY_TO_RECONCILIATION:
        return CATEGORY_TO_RECONCILIATION[movement.category]
    if movement.movement_type == "cargo":
        return "EGRESO"
    if movement.movement_type == "abono":
        return "INGRESO"
    return None


def _default_payee(movement: TreasuryMovement) -> str | None:
    if movement.counterparty:
        return movement.counterparty
    if movement.reference and movement.reference.isdigit():
        return None
    if movement.description:
        return movement.description[:120]
    return movement.bank


def _field_value_counts(history_rows: list[dict[str, Any]]) -> dict[str, Counter[str]]:
    counts = {field: Counter() for field in MOVEMENT_EDITABLE_FIELDS}
    counts["movement_type"].update(DEFAULT_MOVEMENT_TYPES)
    for row in history_rows:
        values = row["values"]
        for field in MOVEMENT_EDITABLE_FIELDS:
            value = _normalize_spaces(values.get(field))
            if value:
                counts[field][value] += 1
    return counts


def _find_sheet_header(ws) -> tuple[int, int, int]:
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        row_values = [ws.cell(row_idx, col).value for col in range(1, min(ws.max_column, 20) + 1)]
        headers = [_canonical_field_name(value) for value in row_values]
        if {"date", "deposits", "withdrawals"} <= {item for item in headers if item}:
            return row_idx, 1, min(ws.max_column, 20)
    raise ValueError(f"No se pudo detectar el encabezado de la hoja {ws.title}.")


def _history_label(values: dict[str, str | None]) -> str | None:
    parts = [values.get("payee"), values.get("reconciliation"), values.get("detailed_concept")]
    label = " · ".join(part for part in parts if part)
    return label[:140] if label else None


def _read_sheet_profile(ws) -> dict[str, Any]:
    table_name = next(iter(ws.tables.keys()), None) if ws.tables else None
    if table_name:
        table = ws.tables[table_name]
        start_col, start_row, end_col, end_row = range_boundaries(table.ref)
    else:
        start_row, start_col, end_col = _find_sheet_header(ws)
        end_row = ws.max_row

    field_columns: dict[str, int] = {}
    for col in range(start_col, end_col + 1):
        field_name = _canonical_field_name(ws.cell(start_row, col).value)
        if field_name and field_name not in field_columns:
            field_columns[field_name] = col

    history_rows: list[dict[str, Any]] = []
    existing_signatures: set[str] = set()
    data_start_row = start_row + 1
    scan_end_row = max(end_row, ws.max_row)
    for row_idx in range(data_start_row, scan_end_row + 1):
        row_values: dict[str, Any] = {
            field: ws.cell(row_idx, column).value for field, column in field_columns.items()
        }
        has_content = any(
            row_values.get(field) not in (None, "")
            for field in ("date", "payee", "detailed_concept", "deposits", "withdrawals", "reconciliation", "group")
        )
        if not has_content:
            continue

        payload = {
            field: _normalize_spaces(_as_text(row_values.get(field))) for field in MOVEMENT_EDITABLE_FIELDS
        }
        payload["date"] = _iso_date(row_values.get("date"))
        payload["deposits"] = _float_value(row_values.get("deposits"))
        payload["withdrawals"] = _float_value(row_values.get("withdrawals"))
        payload["row_number"] = row_idx
        payload["search_tokens"] = _search_tokens(
            payload.get("payee"),
            payload.get("reconciliation"),
            payload.get("specific_concept"),
            payload.get("detailed_concept"),
            payload.get("observations"),
        )
        history_rows.append(
            {
                "row_number": row_idx,
                "values": payload,
                "label": _history_label(payload),
            }
        )

        amount = (payload.get("deposits") or 0.0) - (payload.get("withdrawals") or 0.0)
        existing_signatures.add(
            _source_movement_signature(
                movement_date=payload.get("date"),
                amount=amount,
                detail=payload.get("detailed_concept"),
            )
        )

    value_counts = _field_value_counts(history_rows)
    field_options = {
        field: [value for value, _ in value_counts[field].most_common(24)] for field in MOVEMENT_EDITABLE_FIELDS
    }
    defaults = {
        field: (field_options[field][0] if field_options[field] else None) for field in MOVEMENT_EDITABLE_FIELDS
    }
    defaults["movement_type"] = defaults["movement_type"] or _movement_type_default(
        TreasuryMovement(statement_id="", source_file="", bank="", sequence=0),
        ws.title,
    )

    return {
        "name": ws.title,
        "bank_hint": _sheet_bank_hint(ws.title),
        "sheet_kind": _sheet_kind(ws.title),
        "table_name": table_name,
        "table_start_col": start_col,
        "table_end_col": end_col,
        "table_end_row": end_row,
        "header_row": start_row,
        "data_start_row": data_start_row,
        "field_columns": field_columns,
        "field_options": field_options,
        "defaults": defaults,
        "history_rows": history_rows,
        "existing_signatures": existing_signatures,
    }


# ── Statement / sheet scoring ─────────────────────────────

def _score_statement_sheet(statement: TreasuryStatement, profile: dict[str, Any]) -> int:
    score = 0
    if profile["bank_hint"] == statement.bank:
        score += 20
    elif profile["bank_hint"]:
        return -100

    statement_digits = _extract_digit_tokens(statement.account_number, statement.clabe, statement.contract, statement.source_file)
    title_normalized = _normalize(profile["name"])
    if any(token and token in re.sub(r"\D", "", title_normalized) for token in statement_digits if len(token) >= 4):
        score += 12
    elif any(token and token in title_normalized for token in statement_digits if len(token) == 4):
        score += 10

    statement_kind_val = _statement_kind(statement)
    if profile["sheet_kind"] == statement_kind_val:
        score += 6
    elif profile["sheet_kind"] != "operational" and statement_kind_val != "operational":
        score -= 4

    raw_hint = _normalize(" ".join(part for part in [statement.source_file, statement.raw_text[:600]] if part))
    if "obra publica" in raw_hint and "op" in title_normalized:
        score += 4
    return score


def _rank_statement_sheets(statement: TreasuryStatement, profiles: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ranked = [
        {
            "name": profile["name"],
            "score": _score_statement_sheet(statement, profile),
            "bank_hint": profile["bank_hint"],
        }
        for profile in profiles
    ]
    ranked.sort(key=lambda item: (-item["score"], item["name"]))
    bank_matches = [item for item in ranked if item["bank_hint"] == statement.bank]
    if bank_matches:
        return bank_matches
    useful = [item for item in ranked if item["score"] > -100]
    return useful or ranked


# ── History matching ──────────────────────────────────────

def _match_history_row(profile: dict[str, Any], movement: TreasuryMovement) -> tuple[dict[str, Any] | None, int]:
    movement_tokens = _search_tokens(
        movement.counterparty,
        movement.description,
        movement.concept,
        movement.long_description,
        movement.reference,
    )
    if not movement_tokens:
        return None, 0

    best_row: dict[str, Any] | None = None
    best_score = 0
    for row in profile["history_rows"]:
        row_tokens = row["values"]["search_tokens"]
        if not row_tokens:
            continue
        overlap = len(movement_tokens & row_tokens)
        if overlap == 0:
            continue
        score = overlap * 2
        row_payee = _normalize(row["values"].get("payee") or "")
        if movement.counterparty and row_payee and _normalize(movement.counterparty) == row_payee:
            score += 8
        if movement.reference and movement.reference in (row["values"].get("detailed_concept") or ""):
            score += 4
        amount_delta = abs(
            ((row["values"].get("deposits") or 0.0) - (row["values"].get("withdrawals") or 0.0))
            - _movement_amount(movement)
        )
        if amount_delta <= 0.01:
            score += 3
        if score > best_score:
            best_row = row
            best_score = score
    return best_row, best_score


# ── Draft building ────────────────────────────────────────

def _initial_draft_values(
    statement: TreasuryStatement,
    movement: TreasuryMovement,
    sheet_name: str | None,
    profile: dict[str, Any] | None,
    history_row: dict[str, Any] | None,
) -> dict[str, Any]:
    values = {
        "movement_type": None,
        "date": movement.movement_date,
        "company": None,
        "payee": None,
        "group": None,
        "business_unit": None,
        "project": None,
        "reconciliation": None,
        "specific_concept": None,
        "detailed_concept": None,
        "deposits": movement.credit,
        "withdrawals": movement.debit,
        "breakdown": movement.debit if movement.debit else None,
        "observations": None,
    }

    if history_row:
        for field in MOVEMENT_EDITABLE_FIELDS:
            values[field] = history_row["values"].get(field)

    if profile:
        for field in MOVEMENT_EDITABLE_FIELDS:
            if not values.get(field):
                values[field] = profile["defaults"].get(field)

    values["movement_type"] = values["movement_type"] or _movement_type_default(movement, sheet_name)
    values["company"] = values["company"] or statement.account_holder or (profile["defaults"].get("company") if profile else None) or "DEESA"
    values["payee"] = values["payee"] or _default_payee(movement)
    values["reconciliation"] = values["reconciliation"] or _default_reconciliation(movement)
    values["specific_concept"] = values["specific_concept"] or values["reconciliation"]
    values["detailed_concept"] = values["detailed_concept"] or _normalize_spaces(
        " ".join(part for part in [movement.description, movement.concept, movement.long_description] if part)
    )
    return values


def _missing_draft_fields(values: dict[str, Any], profile: dict[str, Any] | None, sheet_name: str | None) -> list[str]:
    missing: list[str] = []
    if not sheet_name:
        missing.append("sheet_name")

    review_fields = ("movement_type", "company", "payee", "reconciliation", "detailed_concept")
    if profile:
        review_fields += tuple(
            field
            for field in ("group", "business_unit", "project")
            if field in profile["field_columns"]
        )

    seen: set[str] = set()
    for field in review_fields:
        if field in seen:
            continue
        seen.add(field)
        if not _normalize_spaces(_as_text(values.get(field))):
            missing.append(field)
    return missing


def _build_movement_draft(
    statement: TreasuryStatement,
    movement: TreasuryMovement,
    sheet_name: str | None,
    sheet_options: list[str],
    profile: dict[str, Any] | None,
) -> dict[str, Any]:
    history_row, history_score = _match_history_row(profile, movement) if profile else (None, 0)
    values = _initial_draft_values(statement, movement, sheet_name, profile, history_row)
    missing_fields = _missing_draft_fields(values, profile, sheet_name)
    suggestion_source = "historial" if history_row and history_score >= 4 else "automático"

    return {
        "draft_id": f"{statement.id}-{movement.sequence}",
        "statement_id": statement.id,
        "statement_label": _statement_label(statement),
        "sheet_name": sheet_name,
        "sheet_options": sheet_options,
        "suggestion_source": suggestion_source,
        "suggestion_score": history_score,
        "matched_history_label": history_row.get("label") if history_row else None,
        "missing_fields": missing_fields,
        "needs_review": bool(missing_fields) or suggestion_source != "historial",
        "movement": {
            "sequence": movement.sequence,
            "bank": movement.bank,
            "account_number": movement.account_number,
            "movement_date": movement.movement_date,
            "description": movement.description,
            "concept": movement.concept,
            "reference": movement.reference,
            "counterparty": movement.counterparty,
            "debit": movement.debit,
            "credit": movement.credit,
            "balance": movement.balance,
        },
        "values": values,
    }


# ── Public: prepare templates ─────────────────────────────

def prepare_movement_template(template_path: Path, statements: list[TreasuryStatement]) -> dict[str, Any]:
    """Analyze a movement template and match statements to sheets."""
    workbook = load_workbook(template_path)
    profiles = [_read_sheet_profile(ws) for ws in workbook.worksheets]
    profile_map = {profile["name"]: profile for profile in profiles}

    drafts: list[dict[str, Any]] = []
    unmatched_statements: list[str] = []
    skipped_duplicates = 0
    review_count = 0

    for statement in statements:
        ranked = _rank_statement_sheets(statement, profiles)
        sheet_options = [item["name"] for item in ranked if item["score"] > -100]
        if not sheet_options:
            sheet_options = [profile["name"] for profile in profiles]
        best_score = ranked[0]["score"] if ranked else -100
        selected_sheet = sheet_options[0] if sheet_options and best_score > -100 else None
        if not selected_sheet:
            unmatched_statements.append(_statement_label(statement))
        profile = profile_map.get(selected_sheet) if selected_sheet else None

        for movement in statement.movements:
            source_signature = _movement_signature(movement)
            if profile and source_signature in profile["existing_signatures"]:
                skipped_duplicates += 1
                continue

            draft = _build_movement_draft(statement, movement, selected_sheet, sheet_options, profile)
            drafts.append(draft)
            if draft["needs_review"]:
                review_count += 1
            if profile:
                profile["existing_signatures"].add(source_signature)

    return {
        "filename": template_path.name,
        "sheets": [
            {
                "name": profile["name"],
                "bank_hint": profile["bank_hint"],
                "sheet_kind": profile["sheet_kind"],
                "field_options": profile["field_options"],
                "defaults": profile["defaults"],
            }
            for profile in profiles
        ],
        "drafts": drafts,
        "review_count": review_count,
        "skipped_duplicates": skipped_duplicates,
        "unmatched_statements": unmatched_statements,
    }


def prepare_balance_template(template_path: Path, statements: list[TreasuryStatement]) -> dict[str, Any]:
    """Analyze a balance template and match statements to rows."""
    workbook = load_workbook(template_path)
    ws = workbook[workbook.sheetnames[0]]

    updates: list[dict[str, Any]] = []
    matched_statement_ids: set[str] = set()
    for row_idx in range(1, ws.max_row + 1):
        bank_value = _normalize_spaces(_as_text(ws.cell(row_idx, 4).value))
        account_label = _normalize_spaces(_as_text(ws.cell(row_idx, 5).value))
        clabe_value = _normalize_spaces(_as_text(ws.cell(row_idx, 6).value))
        if not bank_value and not account_label:
            continue

        candidates = []
        for statement in statements:
            if statement.closing_balance is None:
                continue
            score, reason = _score_balance_row(statement, bank_value, account_label, clabe_value)
            if score <= 0:
                continue
            candidates.append((score, reason, statement))
        if not candidates:
            continue

        score, reason, statement = sorted(candidates, key=lambda item: (-item[0], item[2].source_file))[0]
        if score < 18:
            continue

        is_dollar_row = _sheet_kind(" ".join(part for part in [bank_value, account_label] if part)) == "usd"
        column_key = "dolares" if is_dollar_row else "pesos"
        current_value = _float_value(ws.cell(row_idx, 9 if is_dollar_row else 8).value)
        updates.append(
            {
                "id": f"balance-{row_idx}",
                "row_number": row_idx,
                "bank": bank_value,
                "account_label": account_label,
                "column_key": column_key,
                "current_value": current_value,
                "new_value": round(float(statement.closing_balance), 2),
                "statement_id": statement.id,
                "statement_label": _statement_label(statement),
                "confidence": round(min(score / 34, 0.99), 2),
                "reason": reason,
                "enabled": True,
            }
        )
        matched_statement_ids.add(statement.id)

    return {
        "filename": template_path.name,
        "sheet_name": ws.title,
        "updates": updates,
        "unmatched_statements": [
            _statement_label(statement) for statement in statements if statement.id not in matched_statement_ids
        ],
    }


def _score_balance_row(statement: TreasuryStatement, bank_value: str | None, account_label: str | None, clabe_value: str | None) -> tuple[int, str]:
    row_text = " ".join(part for part in [bank_value, account_label, clabe_value] if part)
    normalized_row = _normalize(row_text)
    score = 0
    reason_parts: list[str] = []

    bank_hint = _sheet_bank_hint(bank_value or "")
    if bank_hint == statement.bank:
        score += 18
        reason_parts.append("banco")

    statement_digits = _extract_digit_tokens(statement.account_number, statement.clabe, statement.contract, statement.source_file)
    row_digits = _extract_digit_tokens(account_label, clabe_value, bank_value)
    if any(token in row_digits for token in statement_digits if len(token) >= 4):
        score += 12
        reason_parts.append("cuenta")
    elif any(token in normalized_row for token in statement_digits if len(token) == 4):
        score += 8
        reason_parts.append("últimos 4 dígitos")

    row_kind = _sheet_kind(row_text)
    statement_kind_val = _statement_kind(statement)
    if row_kind == statement_kind_val:
        score += 4
        reason_parts.append("tipo de cuenta")
    elif row_kind != "operational" and statement_kind_val != "operational":
        score -= 3

    return score, ", ".join(reason_parts) or "coincidencia general"


# ── Public: render workbooks ──────────────────────────────

def _row_is_blank_for_insert(ws, profile: dict[str, Any], row_idx: int) -> bool:
    for field in ("date", "payee", "detailed_concept", "deposits", "withdrawals"):
        column = profile["field_columns"].get(field)
        if column is None:
            continue
        value = ws.cell(row_idx, column).value
        if value not in (None, ""):
            return False
    return True


def _clone_template_row(ws, profile: dict[str, Any], source_row: int, target_row: int) -> None:
    for col_idx in range(profile["table_start_col"], profile["table_end_col"] + 1):
        source = ws.cell(source_row, col_idx)
        target = ws.cell(target_row, col_idx)
        target._style = copy.copy(source._style)
        target.number_format = source.number_format
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.protection = copy.copy(source.protection)
        if isinstance(source.value, str) and source.value.startswith("="):
            target.value = Translator(source.value, origin=source.coordinate).translate_formula(target.coordinate)
        else:
            target.value = None


def _ensure_target_row(ws, profile: dict[str, Any]) -> int:
    for row_idx in range(profile["data_start_row"], profile["table_end_row"] + 1):
        if _row_is_blank_for_insert(ws, profile, row_idx):
            if row_idx > ws.max_row:
                source_row = max(profile["data_start_row"], min(ws.max_row, row_idx - 1))
                _clone_template_row(ws, profile, source_row, row_idx)
            return row_idx

    source_row = max(profile["data_start_row"], min(ws.max_row, profile["table_end_row"]))
    target_row = profile["table_end_row"] + 1
    _clone_template_row(ws, profile, source_row, target_row)
    if profile["table_name"]:
        table = ws.tables[profile["table_name"]]
        start_col, start_row, end_col, _ = range_boundaries(table.ref)
        table.ref = f"{ws.cell(start_row, start_col).coordinate}:{ws.cell(target_row, end_col).coordinate}"
    profile["table_end_row"] = target_row
    return target_row


def _write_draft_row(ws, profile: dict[str, Any], row_idx: int, draft: dict[str, Any]) -> None:
    values = draft["values"]
    for field, column in profile["field_columns"].items():
        cell = ws.cell(row_idx, column)
        if field == "date":
            if values.get("date"):
                cell.value = datetime.strptime(values["date"], "%Y-%m-%d")
        elif field == "deposits":
            if values.get("deposits") is not None:
                cell.value = round(float(values["deposits"]), 2)
        elif field == "withdrawals":
            if values.get("withdrawals") is not None:
                cell.value = round(float(values["withdrawals"]), 2)
        elif field == "breakdown":
            if values.get("breakdown") is not None:
                cell.value = round(float(values["breakdown"]), 2)
        elif field == "balance":
            if not (isinstance(cell.value, str) and cell.value.startswith("=")) and draft["movement"].get("balance") is not None:
                cell.value = round(float(draft["movement"]["balance"]), 2)
        elif field in values and values.get(field):
            cell.value = values[field]


def render_movement_workbook(template_path: Path, drafts: list[dict[str, Any]]) -> bytes:
    """Apply movement drafts to the template.

    Creates a new sheet for each day that has movements, copying structure
    from the matching template sheet.
    """
    workbook = load_workbook(template_path, keep_links=True)
    # Remove table definitions to prevent openpyxl corruption
    for sheet in workbook.worksheets:
        if hasattr(sheet, "_tables"):
            sheet._tables = []
    profiles = {ws.title: _read_sheet_profile(ws) for ws in workbook.worksheets}

    # Group drafts by sheet_name, then by date
    by_sheet: dict[str, list[dict[str, Any]]] = {}
    for draft in drafts:
        if not draft.get("sheet_name"):
            continue
        by_sheet.setdefault(draft["sheet_name"], []).append(draft)

    for sheet_name, sheet_drafts in by_sheet.items():
        if sheet_name not in profiles:
            continue
        source_profile = profiles[sheet_name]

        # Group this sheet's drafts by date
        by_date: dict[str, list[dict[str, Any]]] = {}
        for draft in sorted(sheet_drafts, key=lambda d: (d["values"].get("date") or "", d["draft_id"])):
            date_key = draft["values"].get("date") or "sin-fecha"
            by_date.setdefault(date_key, []).append(draft)

        for date_key, day_drafts in by_date.items():
            # Build a sheet name like "BBVA Pesos 2026-01-31" or use the original if only one day
            if len(by_date) == 1:
                target_sheet_name = sheet_name
            else:
                # Format date for sheet name (max 31 chars for Excel)
                date_suffix = date_key[5:] if date_key.startswith("20") else date_key  # "01-31"
                target_sheet_name = f"{sheet_name} {date_suffix}"[:31]

            if target_sheet_name == sheet_name:
                # Write directly to the existing sheet
                ws = workbook[sheet_name]
                profile = source_profile
            else:
                # Clone the template sheet for this day
                source_ws = workbook[sheet_name]
                ws = workbook.copy_worksheet(source_ws)
                ws.title = target_sheet_name
                profile = _read_sheet_profile(ws)
                # Clear existing data rows (keep only header)
                for row_idx in range(profile["data_start_row"], profile["table_end_row"] + 1):
                    for col_idx in range(profile["table_start_col"], profile["table_end_col"] + 1):
                        cell = ws.cell(row_idx, col_idx)
                        if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                            cell.value = None

            for draft in day_drafts:
                row_idx = _ensure_target_row(ws, profile)
                _write_draft_row(ws, profile, row_idx, draft)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def render_balance_workbook(template_path: Path, updates: list[dict[str, Any]]) -> bytes:
    """Apply balance updates to the template and return the workbook bytes."""
    workbook = load_workbook(template_path, keep_links=True)
    # Remove any table definitions to prevent openpyxl from corrupting them on save
    for sheet in workbook.worksheets:
        if hasattr(sheet, "_tables"):
            sheet._tables = []
    ws = workbook[workbook.sheetnames[0]]
    for update in updates:
        if not update.get("enabled", True):
            continue
        row_idx = int(update["row_number"])
        column_idx = 9 if update.get("column_key") == "dolares" else 8
        ws.cell(row_idx, column_idx).value = round(float(update["new_value"]), 2)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
