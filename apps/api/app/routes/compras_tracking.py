from __future__ import annotations

import csv
import re
import tempfile
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy.orm import Session

from app.core.config import settings
from app.db.models import AppDefinition, Job, JobFile, User
from app.db.session import get_db
from app.deps import ensure_app_access, require_user

try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None


router = APIRouter(prefix="/tools/compras/importaciones-tracking", tags=["tools-compras-tracking"])

TRACKING_APP_KEY = "era_compras_seguimiento_importaciones"
IMPORTACIONES_SOURCE_APP_KEY = "era_importaciones_generador_oc"
UPLOAD_PREFIX_RX = re.compile(r"^\d{2,3}-(?P<name>.+)$")
STOP_BLOCK_MARKERS = (
    "CONSIGNEE",
    "IMPORTER",
    "IMPORTADOR",
    "TO INVOICE",
    "PACKING LIST",
    "PORT ",
    "VESSEL",
    "COUNTRY OF ORIGIN",
    "PAIS DE ORIGEN",
    "INCOTERM",
    "TERMS OF DELIVERY",
    "MARKS",
    "DESCRIPTION OF GOODS",
    "FLETE",
    "TYPE OF CONTAINER",
    "TIPE OF CONTAINER",
)

FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "order_number": ("no po", "po", "orden de compra", "pedido", "folio", "no. po"),
    "general_po": ("general po", "no general po", "no. general po"),
    "invoice_number": ("factura", "invoice", "invoice no", "no factura"),
    "supplier_display": ("proveedor", "vendor", "seller", "vendedor", "nombre proveedor"),
    "container": ("contenedor", "container"),
    "start_production": ("start of production", "inicio produccion", "inicio producción"),
    "end_production": ("end of production", "fin produccion", "fin producción"),
    "inspection_day": ("inspection day", "inspeccion", "inspección"),
    "etd": ("etd", "fecha salida", "salida puerto"),
    "eta": ("eta", "fecha llegada", "llegada puerto", "arribo puerto"),
    "port_arrival": ("port arrival", "arribo puerto real", "llegada puerto real"),
    "customs_release": ("customs release", "liberacion aduana", "liberación aduana", "despacho"),
    "warehouse_arrival": ("warehouse arrival", "llegada almacen", "llegada almacén", "recepcion almacen", "recepción almacén"),
    "current_stage": ("stage", "etapa", "fase", "milestone", "seguimiento actual"),
    "status": ("status", "estatus", "estado"),
    "terminal": ("terminal",),
    "comments": ("comments", "comentarios", "observaciones", "notas"),
}

STAGE_KEYWORDS: dict[str, tuple[str, ...]] = {
    "order_created": ("orden", "order", "po"),
    "production_start": ("inicio produccion", "inicio producción", "start of production", "produccion", "producción"),
    "production_end": ("fin produccion", "fin producción", "end of production"),
    "inspection": ("inspeccion", "inspección", "inspection"),
    "departure": ("etd", "salida", "embarque", "departure"),
    "arrival_port": ("eta", "arribo puerto", "llegada puerto", "port arrival"),
    "customs": ("aduana", "despacho", "customs", "liberacion", "liberación"),
    "warehouse": ("almacen", "almacén", "recepcion", "recepción", "warehouse"),
}


@dataclass
class TrackingMilestone:
    key: str
    label: str
    date: str | None
    status: str


@dataclass
class ShipmentTracking:
    id: str
    source: str
    order_number: str | None = None
    general_po: str | None = None
    invoice_number: str | None = None
    supplier_display: str | None = None
    supplier_name: str | None = None
    container: str | None = None
    origin_port: str | None = None
    destination_port: str | None = None
    terminal: str | None = None
    incoterm: str | None = None
    total_usd: float | None = None
    order_date: str | None = None
    start_production: str | None = None
    end_production: str | None = None
    inspection_day: str | None = None
    etd: str | None = None
    eta: str | None = None
    port_arrival: str | None = None
    customs_release: str | None = None
    warehouse_arrival: str | None = None
    current_stage: str | None = None
    status: str | None = None
    comments: str | None = None
    source_updated_at: str | None = None
    has_operational_data: bool = False
    progress_pct: int = 0
    stage_label: str = "Pendiente"
    milestones: list[TrackingMilestone] = field(default_factory=list)


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    for source, target in (
        ("á", "a"),
        ("é", "e"),
        ("í", "i"),
        ("ó", "o"),
        ("ú", "u"),
        ("ñ", "n"),
        (".", " "),
        ("_", " "),
        ("-", " "),
        ("/", " "),
    ):
        text = text.replace(source, target)
    return " ".join(text.split())


def _to_iso_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    raw = str(value).strip().replace("/", "-").replace(".", "-")
    raw = re.sub(r"\s*-\s*", "-", raw)
    raw = " ".join(raw.split())
    for fmt in ("%d-%m-%Y", "%m-%d-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            pass

    month_match = re.search(r"\b(\d{1,2})-([A-Za-z]{3})-?(20\d{2})\b", raw)
    if month_match:
        day, month, year = month_match.groups()
        months = {
            "JAN": 1,
            "FEB": 2,
            "MAR": 3,
            "APR": 4,
            "MAY": 5,
            "JUN": 6,
            "JUL": 7,
            "AUG": 8,
            "SEP": 9,
            "OCT": 10,
            "NOV": 11,
            "DEC": 12,
        }
        month_num = months.get(month.upper())
        if month_num:
            try:
                return datetime(int(year), month_num, int(day)).date().isoformat()
            except ValueError:
                return None
    return None


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "").replace("$", "").strip())
    except Exception:
        return None


def _extract_pdf_text(path: Path) -> str:
    if fitz is None:
        raise RuntimeError("PyMuPDF no está disponible en la API.")
    document = fitz.open(str(path))
    try:
        pages = [page.get_text("text") for page in document]
    finally:
        document.close()
    return "\n".join(pages).replace("\xa0", " ")


def _clean_lines(text: str) -> list[str]:
    return [line.strip() for line in text.splitlines() if line and line.strip()]


def _line_value_after_label(line: str, labels: tuple[str, ...]) -> str | None:
    normalized = line.strip()
    for label in labels:
        pattern = rf"^{label}\s*[:.]?\s*(.+)$"
        match = re.match(pattern, normalized, re.IGNORECASE)
        if match:
            value = match.group(1).strip(" :.-")
            if value:
                return value
            return None
    return None


def _extract_labeled_value(lines: list[str], labels: tuple[str, ...], max_lookahead: int = 3) -> str | None:
    for index, line in enumerate(lines):
        inline_value = _line_value_after_label(line, labels)
        if inline_value:
            return inline_value

        if not any(re.match(rf"^{label}\s*[:.]?$", line.strip(), re.IGNORECASE) for label in labels):
            continue

        for candidate in lines[index + 1 : index + 1 + max_lookahead]:
            value = candidate.strip(" :.-")
            if value:
                return value
    return None


def _extract_labeled_block(lines: list[str], labels: tuple[str, ...], max_lines: int = 5) -> list[str]:
    for index, line in enumerate(lines):
        if not any(label in line.upper() for label in labels):
            continue

        tail = re.sub(r"^EXPORTER\s*(\([^)]*\))?\s*", "", line, flags=re.IGNORECASE)
        tail = re.sub(r"^PROVEEDOR\s*:?\s*", "", tail, flags=re.IGNORECASE)
        tail = tail.strip(" :.-")
        block = [tail] if tail else []

        for candidate in lines[index + 1 : index + 1 + max_lines]:
            value = candidate.strip()
            if not value:
                break
            if any(value.upper().startswith(marker) for marker in STOP_BLOCK_MARKERS):
                break
            block.append(value)

        return block
    return []


def _extract_supplier_name(lines: list[str], labels: tuple[str, ...]) -> str | None:
    for index, line in enumerate(lines):
        upper = line.upper()
        if not any(label in upper for label in labels):
            continue

        tail = re.sub(r"^EXPORTER\s*(\([^)]*\))?\s*", "", line, flags=re.IGNORECASE).strip(" :.-")
        if tail and tail.upper() not in labels:
            return tail

        for candidate in lines[index + 1 : index + 4]:
            value = candidate.strip()
            if value and value.upper() not in labels:
                return value
    return None


def _extract_invoice_from_lines(lines: list[str]) -> tuple[str | None, str | None]:
    invoice_labels = ("INVOICE\\s+NO", "INVOICE\\s+NUMBER", "FACTURA\\s+NO")

    for index, line in enumerate(lines):
        if "TO INVOICE NO. DATE" not in line.upper():
            continue
        candidate = lines[index + 1] if index + 1 < len(lines) else ""
        match = re.search(
            r"([A-Z0-9-]{4,})\s+(\d{1,2}-[A-Z]{3}-?20\d{2}|\d{1,2}[/-]\d{1,2}[/-]20\d{2})$",
            candidate,
            re.IGNORECASE,
        )
        if match:
            return match.group(1).strip(), _to_iso_date(match.group(2))

    for index, line in enumerate(lines):
        inline_value = _line_value_after_label(line, invoice_labels)
        if inline_value:
            date_value = None
            for candidate in lines[index : index + 4]:
                extracted = _line_value_after_label(candidate, ("DATE",))
                if extracted:
                    date_value = extracted
                    break
            if date_value is None:
                date_value = _extract_labeled_value(lines[index : index + 4], ("DATE",), max_lookahead=2)
            return inline_value, _to_iso_date(date_value)

        if re.match(r"^INVOICE\s+NO\s*[:.]?$", line.strip(), re.IGNORECASE):
            invoice_value = _extract_labeled_value(lines[index : index + 4], ("INVOICE\\s+NO",), max_lookahead=2)
            date_value = _extract_labeled_value(lines[index : index + 6], ("DATE",), max_lookahead=2)
            return invoice_value, _to_iso_date(date_value)

    filename_match = re.search(r"\b(\d{2}[A-Z]\d{2}(?:[+/_-]\d{2})?)\b", " ".join(lines[:10]).upper())
    if filename_match:
        return filename_match.group(1).replace("_", "+"), None

    return None, None


def _extract_supplier_from_lines(lines: list[str]) -> str | None:
    supplier_block = _extract_labeled_block(lines, ("EXPORTER", "PROVEEDOR"))
    if supplier_block:
        return supplier_block[0]

    supplier_name = _extract_supplier_name(lines, ("EXPORTER",))
    if supplier_name:
        return supplier_name

    for index, line in enumerate(lines):
        if "PACKING LIST" not in line.upper():
            continue
        if index > 0:
            candidate = lines[index - 1].strip()
            if candidate and "PACKING LIST" not in candidate.upper():
                return candidate
        break
    return None


def _extract_ports_from_lines(lines: list[str]) -> tuple[str | None, str | None]:
    origin_port = None
    destination_port = None

    for index, line in enumerate(lines):
        if line.upper() == "FROM" and index + 1 < len(lines):
            candidate = lines[index + 1].strip()
            if "," in candidate:
                origin_port = candidate
        if "TERMS OF DELIVERY" in line.upper():
            for follow_index in range(index + 1, min(index + 5, len(lines))):
                if lines[follow_index].upper() == "TO" and follow_index + 1 < len(lines):
                    candidate = lines[follow_index + 1].strip()
                    if "," in candidate:
                        destination_port = candidate
                        break

    if not origin_port:
        from_value = _extract_labeled_value(lines, ("FROM",), max_lookahead=2)
        if from_value and "," in from_value:
            origin_port = from_value
    if not destination_port:
        to_value = _extract_labeled_value(lines, ("PORT OF DISCHARGE", "TO"), max_lookahead=2)
        if to_value and "," in to_value:
            destination_port = to_value

    return origin_port, destination_port


def _source_name(path: Path) -> str:
    match = UPLOAD_PREFIX_RX.match(path.stem)
    return match.group("name") if match else path.stem


def _fallback_identifier(path: Path) -> str | None:
    match = re.search(r"\b(\d{2}[A-Z]\d{2}(?:[+/_-]\d{2})?)\b", _source_name(path).upper())
    if not match:
        return None
    return match.group(1).replace("_", "+").replace("-", "+", 1) if "_" in match.group(1) else match.group(1)


def _parse_order_pdf(path: Path) -> ShipmentTracking:
    text = _extract_pdf_text(path)
    lines = _clean_lines(text)

    shipment = ShipmentTracking(
        id=path.stem,
        source="pdf",
    )

    order_match = re.search(r"NO\.\s*PO:\s*([A-Z0-9+/\-]+)(?:\s+DATE:|\b)", text, re.IGNORECASE)
    if order_match:
        shipment.order_number = order_match.group(1).strip()

    general_match = re.search(r"NO\.\s*GENERAL\s*PO:\s*([A-Z0-9\-]+)", text, re.IGNORECASE)
    if general_match:
        shipment.general_po = general_match.group(1).strip()

    invoice_number, invoice_date = _extract_invoice_from_lines(lines)
    if invoice_number:
        shipment.invoice_number = invoice_number
        if shipment.order_number is None:
            shipment.order_number = invoice_number
        if shipment.order_date is None:
            shipment.order_date = invoice_date

    if shipment.order_date is None:
        date_value = _extract_labeled_value(lines, ("DATE",), max_lookahead=2)
        shipment.order_date = _to_iso_date(date_value)

    provider_match = re.search(r"PROVE?DOR\s+([A-Z/& ]+)", text, re.IGNORECASE)
    if provider_match:
        shipment.supplier_display = provider_match.group(1).strip()

    shipment.supplier_name = _extract_supplier_from_lines(lines)
    supplier_match = re.search(r"TOTAL,\s*USD\s*\$?[\d,]+\.\d{2}\s+([A-Z][A-Z0-9 .,&/-]{3,})\s+PAIS DE ORIGEN", text, re.IGNORECASE | re.DOTALL)
    if supplier_match and not shipment.supplier_name:
        shipment.supplier_name = " ".join(supplier_match.group(1).split())

    incoterm_match = re.search(r"INCOTERM:\s*([A-Z0-9]+)", text, re.IGNORECASE)
    if incoterm_match:
        shipment.incoterm = incoterm_match.group(1).strip().upper()

    total_match = re.search(r"TOTAL,\s*USD\s*\$?([\d,]+\.\d{2})", text, re.IGNORECASE)
    shipment.total_usd = _to_float(total_match.group(1)) if total_match else None

    line_origin = _extract_labeled_value(lines, ("PORT\\s+OF\\s+LOADING", "PORT\\s+OR\\s+LOADING"), max_lookahead=1)
    if line_origin:
        shipment.origin_port = line_origin
    else:
        origin_match = re.search(
            r"PORT\s+OR\s+LOADING:\s*(.+?)(?:\s+PORT\s+OF\s+DISCHARGE:|$)",
            text,
            re.IGNORECASE | re.DOTALL,
        )
        if origin_match:
            shipment.origin_port = " ".join(origin_match.group(1).split())

    line_destination = _extract_labeled_value(lines, ("PORT\\s+OF\\s+DISCHARGE",), max_lookahead=1)
    if line_destination:
        shipment.destination_port = line_destination
    else:
        destination_match = re.search(
            r"PORT\s+OF\s+DISCHARGE:\s*(.+?)(?:\s+TIPE\s+OF\s+CONTAINER:|\s+TYPE\s+OF\s+CONTAINER:|$)",
            text,
            re.IGNORECASE | re.DOTALL,
        )
        if destination_match:
            shipment.destination_port = " ".join(destination_match.group(1).split())

    if not shipment.origin_port or not shipment.destination_port:
        fallback_origin, fallback_destination = _extract_ports_from_lines(lines)
        shipment.origin_port = shipment.origin_port or fallback_origin
        shipment.destination_port = shipment.destination_port or fallback_destination

    if not shipment.origin_port or not shipment.destination_port:
        vessel_match = re.search(r"VESSEL\s+([A-ZÁÉÍÓÚ ,.-]+)\s+TO\s+([A-ZÁÉÍÓÚ ,.-]+)", text.upper(), re.IGNORECASE)
        if vessel_match:
            shipment.origin_port = shipment.origin_port or " ".join(vessel_match.group(1).split())
            shipment.destination_port = shipment.destination_port or " ".join(vessel_match.group(2).split())

    for attr, pattern in (
        ("start_production", r"Start\s+of\s+the\s+production:\s*([0-9./-]{8,10})"),
        ("end_production", r"End\s+of\s+production:\s*([0-9./-]{8,10})"),
        ("inspection_day", r"Inspection\s+day:\s*([0-9./-]{8,10})"),
        ("etd", r"ETD:\s*([0-9./-]{8,10})"),
        ("eta", r"ETA:\s*([0-9./-]{8,10})"),
    ):
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            setattr(shipment, attr, _to_iso_date(match.group(1)))

    container_match = re.search(r"\b([A-Z]{4}\d{7})\b", text)
    if container_match:
        shipment.container = container_match.group(1)

    normalized_destination = _normalize_header(shipment.destination_port)
    if "lazaro" in normalized_destination:
        shipment.terminal = "APM"
    elif "manzanillo" in normalized_destination:
        shipment.terminal = "MANZANILLO"

    if not shipment.supplier_name and lines:
        try:
            total_index = next(idx for idx, line in enumerate(lines) if "TOTAL, USD" in line.upper())
            for next_line in lines[total_index + 1 : total_index + 4]:
                upper = next_line.upper()
                if upper.startswith("PAIS DE ORIGEN") or upper.startswith("INCOTERM"):
                    break
                shipment.supplier_name = next_line
                break
        except StopIteration:
            pass

    if not shipment.order_number and shipment.invoice_number:
        shipment.order_number = shipment.invoice_number
    if not shipment.order_number:
        shipment.order_number = _fallback_identifier(path)

    shipment.id = shipment.order_number or shipment.general_po or _source_name(path)
    return shipment


def _pick_column(headers: list[str], aliases: tuple[str, ...]) -> int | None:
    normalized_headers = [_normalize_header(item) for item in headers]
    normalized_aliases = {_normalize_header(item) for item in aliases}
    for idx, header in enumerate(normalized_headers):
        if header in normalized_aliases:
            return idx
    for idx, header in enumerate(normalized_headers):
        if any(alias in header for alias in normalized_aliases):
            return idx
    return None


def _parse_operations_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle))
    elif suffix in {".xlsx", ".xlsm"}:
        workbook = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    else:
        raise HTTPException(status_code=400, detail="El archivo operativo debe ser CSV, XLSX o XLSM.")

    header_row = next((row for row in rows if any(cell not in (None, "") for cell in row)), None)
    if not header_row:
        return []

    header_index = rows.index(header_row)
    headers = [str(cell or "").strip() for cell in header_row]
    column_map = {field: _pick_column(headers, aliases) for field, aliases in FIELD_ALIASES.items()}

    parsed: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if not any(cell not in (None, "") for cell in row):
            continue

        item: dict[str, Any] = {}
        for field_name, idx in column_map.items():
            if idx is None or idx >= len(row):
                continue
            value = row[idx]
            if value in (None, ""):
                continue
            item[field_name] = value

        if not any(item.get(key) for key in ("order_number", "general_po", "invoice_number", "container")):
            continue

        for date_field in ("start_production", "end_production", "inspection_day", "etd", "eta", "port_arrival", "customs_release", "warehouse_arrival"):
            if date_field in item:
                item[date_field] = _to_iso_date(item[date_field])

        parsed.append(item)

    return parsed


def _match_stage_key(value: str | None) -> str | None:
    normalized = _normalize_header(value)
    if not normalized:
        return None
    for key, aliases in STAGE_KEYWORDS.items():
        if any(alias in normalized for alias in aliases):
            return key
    return None


def _milestone_blueprint() -> list[tuple[str, str, str]]:
    return [
        ("order_created", "Orden emitida", "order_date"),
        ("production_start", "Inicio de producción", "start_production"),
        ("production_end", "Fin de producción", "end_production"),
        ("inspection", "Inspección", "inspection_day"),
        ("departure", "Salida de puerto (ETD)", "etd"),
        ("arrival_port", "Arribo a puerto", "port_arrival"),
        ("customs", "Despacho aduanal", "customs_release"),
        ("warehouse", "Llegada a almacén", "warehouse_arrival"),
    ]


def _coalesce_port_arrival(shipment: ShipmentTracking) -> None:
    if shipment.port_arrival is None:
        shipment.port_arrival = shipment.eta


def _enrich_status(shipment: ShipmentTracking) -> None:
    _coalesce_port_arrival(shipment)
    today = date.today().isoformat()
    stage_key = _match_stage_key(shipment.current_stage or shipment.status)

    milestones: list[TrackingMilestone] = []
    current_index: int | None = None
    latest_completed_index = -1

    for idx, (key, label, attr) in enumerate(_milestone_blueprint()):
        milestone_date = getattr(shipment, attr)
        status = "upcoming"
        if milestone_date:
            status = "completed" if milestone_date <= today else "scheduled"
        milestones.append(TrackingMilestone(key=key, label=label, date=milestone_date, status=status))
        if status == "completed":
            latest_completed_index = idx

    if stage_key:
        for idx, milestone in enumerate(milestones):
            if milestone.key == stage_key:
                current_index = idx
                break

    if latest_completed_index >= 0 and (current_index is None or latest_completed_index > current_index):
        current_index = latest_completed_index

    if current_index is None:
        for idx, milestone in enumerate(milestones):
            if milestone.status in {"scheduled", "upcoming"}:
                current_index = idx
                break
        if current_index is None:
            current_index = len(milestones) - 1

    for idx, milestone in enumerate(milestones):
        if idx < current_index and milestone.status == "upcoming":
            milestone.status = "completed"
        elif idx == current_index:
            milestone.status = "completed" if milestone.status == "completed" and idx == len(milestones) - 1 else "current"
        elif idx > current_index and milestone.status == "scheduled":
            milestone.status = "upcoming"

    shipment.milestones = milestones
    shipment.progress_pct = int(round((current_index / max(len(milestones) - 1, 1)) * 100))
    shipment.stage_label = milestones[current_index].label if milestones else "Pendiente"


def _merge_operations(shipment: ShipmentTracking, ops: dict[str, Any]) -> ShipmentTracking:
    shipment.has_operational_data = True
    if ops.get("supplier_display"):
        shipment.supplier_display = str(ops["supplier_display"]).strip()
    if ops.get("status"):
        shipment.status = str(ops["status"]).strip()
    if ops.get("current_stage"):
        shipment.current_stage = str(ops["current_stage"]).strip()
    if ops.get("terminal"):
        shipment.terminal = str(ops["terminal"]).strip()
    if ops.get("comments"):
        shipment.comments = str(ops["comments"]).strip()

    for attr in ("container", "start_production", "end_production", "inspection_day", "etd", "eta", "port_arrival", "customs_release", "warehouse_arrival"):
        if ops.get(attr):
            setattr(shipment, attr, str(ops[attr]))

    return shipment


def _shipment_identifiers(item: ShipmentTracking | dict[str, Any]) -> list[str]:
    values = []
    for key in ("order_number", "general_po", "invoice_number", "container"):
        value = getattr(item, key, None) if not isinstance(item, dict) else item.get(key)
        if value:
            values.append(str(value).strip().upper())
    return list(dict.fromkeys(values))


def _latest_tracking_date(shipment: ShipmentTracking) -> str:
    actual_dates = [
        shipment.warehouse_arrival,
        shipment.customs_release,
        shipment.port_arrival,
        shipment.eta,
        shipment.etd,
        shipment.order_date,
    ]
    resolved_dates = [value for value in actual_dates if value]
    if resolved_dates:
        return max(resolved_dates)
    return shipment.source_updated_at or "0000-00-00"


def _merge_tracking_data(pdf_shipments: list[ShipmentTracking], operations_rows: list[dict[str, Any]]) -> tuple[list[ShipmentTracking], int]:
    shipments = list(pdf_shipments)
    index: dict[str, ShipmentTracking] = {}
    for shipment in shipments:
        for identifier in _shipment_identifiers(shipment):
            index[identifier] = shipment

    unmatched = 0
    for row in operations_rows:
        target = None
        for identifier in _shipment_identifiers(row):
            target = index.get(identifier)
            if target:
                break

        if target is None:
            target = ShipmentTracking(
                id=str(row.get("order_number") or row.get("general_po") or row.get("container") or f"operativo-{len(shipments)+1}"),
                source="operational",
                order_number=str(row.get("order_number")).strip() if row.get("order_number") else None,
                general_po=str(row.get("general_po")).strip() if row.get("general_po") else None,
                invoice_number=str(row.get("invoice_number")).strip() if row.get("invoice_number") else None,
                supplier_display=str(row.get("supplier_display")).strip() if row.get("supplier_display") else None,
                container=str(row.get("container")).strip() if row.get("container") else None,
            )
            shipments.append(target)
            unmatched += 1

        _merge_operations(target, row)
        for identifier in _shipment_identifiers(target):
            index[identifier] = target

    for shipment in shipments:
        _enrich_status(shipment)

    shipments.sort(
        key=lambda item: (
            _latest_tracking_date(item),
            item.order_number or item.general_po or item.id,
        ),
        reverse=True,
    )
    return shipments, unmatched


def _load_importaciones_history(db: Session, user: User) -> list[ShipmentTracking]:
    files_root = Path(settings.files_root).resolve()

    query = (
        db.query(JobFile, Job)
        .join(Job, Job.id == JobFile.job_id)
        .filter(
            Job.app_key == IMPORTACIONES_SOURCE_APP_KEY,
            Job.status == "succeeded",
            JobFile.role == "input",
        )
        .order_by(Job.created_at.desc(), JobFile.created_at.desc())
    )
    if not user.is_admin:
        query = query.filter(Job.created_by == user.username)

    latest_by_identifier: dict[str, ShipmentTracking] = {}

    for job_file, job in query.all():
        if not job_file.filename.lower().endswith(".pdf"):
            continue
        abs_path = (files_root / job_file.path).resolve()
        if not abs_path.exists():
            continue
        shipment = _parse_order_pdf(abs_path)
        shipment.source = "importaciones-history"
        shipment.source_updated_at = job.created_at.isoformat() if job.created_at else None
        identifiers = _shipment_identifiers(shipment)
        if not identifiers:
            identifiers = [shipment.id]
        if any(identifier in latest_by_identifier for identifier in identifiers):
            continue
        for identifier in identifiers:
            latest_by_identifier[identifier] = shipment

    unique_shipments: dict[str, ShipmentTracking] = {}
    for shipment in latest_by_identifier.values():
        unique_shipments[shipment.id] = shipment

    return list(unique_shipments.values())


def _save_upload(path: Path, upload: UploadFile) -> Path:
    with path.open("wb") as handle:
        handle.write(upload.file.read())
    return path


@router.post("/analyze")
async def analyze_import_tracking(
    order_pdfs: list[UploadFile] | None = File(default=None),
    operations_file: UploadFile | None = File(default=None),
    stage_hint: str | None = Form(default=None),
    use_importaciones_history: bool = Form(default=False),
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    app = db.get(AppDefinition, TRACKING_APP_KEY)
    if not app or not app.enabled:
        raise HTTPException(status_code=404, detail="App de seguimiento no disponible.")

    ensure_app_access(user, TRACKING_APP_KEY, db)

    manual_pdfs = [item for item in (order_pdfs or []) if item.filename]
    if not manual_pdfs and not use_importaciones_history:
        raise HTTPException(status_code=400, detail="Debes subir al menos un PDF de orden de compra.")

    with tempfile.TemporaryDirectory(prefix="compras-tracking-") as temp_dir:
        temp_root = Path(temp_dir)
        pdf_paths: list[Path] = []
        for idx, uploaded in enumerate(manual_pdfs):
            filename = Path(uploaded.filename or f"order_{idx}.pdf").name
            if filename.lower().endswith(".pdf") is False:
                raise HTTPException(status_code=400, detail=f"Archivo inválido: {filename}. Solo se aceptan PDFs.")
            pdf_paths.append(_save_upload(temp_root / f"{idx:03d}-{filename}", uploaded))

        shipments = [_parse_order_pdf(path) for path in pdf_paths]
        if use_importaciones_history:
            shipments.extend(_load_importaciones_history(db, user))

        operations_rows: list[dict[str, Any]] = []
        if operations_file is not None:
            op_filename = Path(operations_file.filename or "operativo.xlsx").name
            op_path = _save_upload(temp_root / op_filename, operations_file)
            operations_rows = _parse_operations_rows(op_path)

        merged, unmatched = _merge_tracking_data(shipments, operations_rows)

    if stage_hint:
        for shipment in merged:
            if not shipment.current_stage:
                shipment.current_stage = stage_hint
                _enrich_status(shipment)

    summary = {
        "shipments": len(merged),
        "with_operations": len([item for item in merged if item.has_operational_data]),
        "arrived_port": len([item for item in merged if item.port_arrival]),
        "customs_released": len([item for item in merged if item.customs_release]),
        "warehouse_arrived": len([item for item in merged if item.warehouse_arrival]),
        "unmatched_operations": unmatched,
    }

    return {
        "summary": summary,
        "shipments": [asdict(item) for item in merged],
    }
