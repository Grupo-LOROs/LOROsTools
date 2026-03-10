from __future__ import annotations

import mimetypes
import uuid
from copy import copy
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.catalog_quote.db import make_catalog_session_factory
from app.catalog_quote.models import PriceList, Product, Supplier, Vendor
from app.core.config import settings
from app.db.models import AppDefinition, CatalogQuoteFolioCounter, User
from app.db.session import get_db
from app.deps import ensure_app_access, require_user


router = APIRouter(prefix="/tools/era/ventas/catalog-quote", tags=["tools-era-ventas-catalog-quote"])

CATALOG_APP_KEY = "era_ventas_cotizador_catalogo"
CATALOG_ASSETS_DIR = Path(__file__).resolve().parents[1] / "assets" / "catalog_quote"
CATALOG_DB_PATH = CATALOG_ASSETS_DIR / "app.db"
CATALOG_TEMPLATE_PATH = CATALOG_ASSETS_DIR / "template.xlsx"
CATALOG_SESSION = make_catalog_session_factory(CATALOG_DB_PATH)

ITEMS_START_ROW = 19
TERMS_START_ROW = 26


class VendorOut(BaseModel):
    id: int
    name: str


class ProductOut(BaseModel):
    sku: str
    description: str
    unit: str
    category: str | None
    supplier: str
    price_list_id: int
    has_container_offer: bool


class TierOut(BaseModel):
    min_qty: int
    label: str
    unit_price: float


class ProductDetailOut(ProductOut):
    tiers: list[TierOut]
    container_qty: int | None
    container_price: float | None
    container_notes: str | None


QuoteMode = Literal["MAYOREO", "CONTENEDOR_POR_CONTENEDOR"]


class QuoteItemIn(BaseModel):
    sku: str = Field(..., min_length=1)
    quantity: int = Field(..., gt=0)
    mode: QuoteMode = "MAYOREO"


class QuoteIn(BaseModel):
    serie: str = "A"
    city: str = "Morelia"
    vendor_name: str | None = None
    customer_name: str | None = None
    items: list[QuoteItemIn]
    iva_mode: Literal["included", "excluded"] = "included"
    iva_rate: float = 0.16


class QuoteLineOut(BaseModel):
    sku: str
    description: str
    unit: str
    quantity: int
    unit_price: float
    line_total: float
    price_rule: str


class QuoteTotals(BaseModel):
    subtotal: float
    iva: float
    total: float


class QuoteOut(BaseModel):
    quote_id: str
    folio: str
    date: date
    city: str
    vendor_name: str | None
    customer_name: str | None
    lines: list[QuoteLineOut]
    totals: QuoteTotals
    download_xlsx_url: str
    download_pdf_url: str


@dataclass(frozen=True)
class PricedLine:
    sku: str
    description: str
    unit: str
    quantity: int
    unit_price: float
    line_total: float
    rule: str


def _ensure_catalog_assets() -> None:
    if not CATALOG_DB_PATH.exists():
        raise HTTPException(status_code=500, detail="No está disponible la base del catálogo.")
    if not CATALOG_TEMPLATE_PATH.exists():
        raise HTTPException(status_code=500, detail="No está disponible la plantilla del cotizador.")


def _round2(value: float) -> float:
    return float(f"{value:.2f}")


def _catalog_session():
    db = CATALOG_SESSION()
    try:
        yield db
    finally:
        db.close()


def _next_folio(db: Session, serie: str) -> str:
    normalized = (serie or "A").strip().upper()[:10] or "A"
    counter = db.get(CatalogQuoteFolioCounter, normalized)
    if counter is None:
        counter = CatalogQuoteFolioCounter(serie=normalized, last_folio=0)
        db.add(counter)
        db.flush()

    counter.last_folio += 1
    db.add(counter)
    db.commit()
    db.refresh(counter)
    return f"{normalized}-{counter.last_folio:06d}"


def _latest_product_query(catalog_db: Session, sku: str):
    return (
        catalog_db.query(Product)
        .filter(Product.sku == sku.strip())
        .join(PriceList, Product.price_list_id == PriceList.id)
        .order_by(PriceList.end_date.desc(), PriceList.created_at.desc())
    )


def _choose_tier(product: Product, qty: int):
    tiers = sorted(product.tiers, key=lambda item: item.min_qty)
    if not tiers:
        raise ValueError(f"El producto {product.sku} no tiene tiers de precio.")
    eligible = [item for item in tiers if item.min_qty <= qty]
    return max(eligible, key=lambda item: item.min_qty) if eligible else tiers[0]


def _price_line(product: Product, qty: int, mode: QuoteMode) -> PricedLine:
    has_offer = product.container_offer is not None
    if has_offer and mode == "MAYOREO":
        raise ValueError(f"El producto {product.sku} solo se vende por contenedor.")
    if not has_offer and mode != "MAYOREO":
        raise ValueError(f"El producto {product.sku} no tiene precio por contenedor.")

    if mode == "MAYOREO":
        tier = _choose_tier(product, qty)
        unit_price = _round2(tier.unit_price)
        return PricedLine(
            sku=product.sku,
            description=product.description,
            unit=product.unit,
            quantity=qty,
            unit_price=unit_price,
            line_total=_round2(unit_price * qty),
            rule=f"Tier {tier.label} (min {tier.min_qty})",
        )

    offer = product.container_offer
    if offer is None:
        raise ValueError(f"El producto {product.sku} no tiene precio por contenedor.")

    unit_price = _round2(offer.container_price)
    return PricedLine(
        sku=product.sku,
        description=f"{product.description} (CONTENEDOR x {offer.container_qty} pzas)",
        unit="CONT",
        quantity=qty,
        unit_price=unit_price,
        line_total=_round2(unit_price * qty),
        rule="Precio por contenedor",
    )


def _compute_totals(subtotal: float, iva_mode: Literal["included", "excluded"], iva_rate: float) -> tuple[float, float, float]:
    if iva_rate < 0:
        raise ValueError("La tasa de IVA es inválida.")

    if iva_mode == "included":
        total = subtotal
        base = total / (1 + iva_rate) if iva_rate != -1 else total
        iva = total - base
        return _round2(base), _round2(iva), _round2(total)

    if iva_mode == "excluded":
        base = subtotal
        iva = base * iva_rate
        return _round2(base), _round2(iva), _round2(base + iva)

    raise ValueError("Modo de IVA inválido.")


def _copy_cell_style(src: Cell, dst: Cell) -> None:
    dst._style = copy(src._style)
    dst.number_format = src.number_format
    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)


def _copy_row(ws: Worksheet, src_row: int, dst_row: int, max_col: int) -> None:
    for column in range(1, max_col + 1):
        src = ws.cell(src_row, column)
        dst = ws.cell(dst_row, column)
        _copy_cell_style(src, dst)
        dst.value = None


def _render_quote_xlsx(output_path: Path, city: str, quote_date: date, lines: list[PricedLine], terms: list[str]) -> None:
    wb = load_workbook(CATALOG_TEMPLATE_PATH)
    ws = wb.active

    ws["F15"].value = city
    ws["F16"].value = quote_date.strftime("%d/%m/%Y")

    max_lines_in_template = TERMS_START_ROW - ITEMS_START_ROW
    if len(lines) > max_lines_in_template:
        extra_rows = len(lines) - max_lines_in_template
        ws.insert_rows(TERMS_START_ROW, amount=extra_rows)
        for offset in range(extra_rows):
            _copy_row(ws, ITEMS_START_ROW, TERMS_START_ROW + offset, max_col=6)

    new_terms_start = TERMS_START_ROW + max(0, len(lines) - max_lines_in_template)
    for row in range(ITEMS_START_ROW, new_terms_start):
        for col in ("A", "B", "C", "D", "E", "F"):
            ws[f"{col}{row}"].value = None

    for index, line in enumerate(lines):
        row = ITEMS_START_ROW + index
        ws[f"A{row}"].value = line.sku
        ws[f"B{row}"].value = line.description
        ws[f"C{row}"].value = line.unit
        ws[f"D{row}"].value = line.quantity
        ws[f"E{row}"].value = line.unit_price
        ws[f"F{row}"].value = f"=D{row}*E{row}"

    if terms:
        last_term_row = new_terms_start
        for row in range(new_terms_start, ws.max_row + 1):
            if ws[f"B{row}"].value:
                last_term_row = row

        insert_row = last_term_row + 1
        for term in terms:
            ws.insert_rows(insert_row, amount=1)
            _copy_row(ws, last_term_row, insert_row, max_col=6)
            ws[f"B{insert_row}"].value = term
            insert_row += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _render_quote_pdf(
    pdf_path: Path,
    folio: str,
    city: str,
    quote_date: date,
    lines: list[PricedLine],
    subtotal: float,
    iva: float,
    total: float,
    vendor_name: str | None,
    customer_name: str | None,
    terms: list[str],
) -> None:
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    doc = canvas.Canvas(str(pdf_path), pagesize=letter)
    width, height = letter

    doc.setFont("Helvetica-Bold", 16)
    doc.drawString(0.9 * inch, height - 0.9 * inch, "COTIZACIÓN DE CATÁLOGO")
    doc.setFont("Helvetica", 10)
    header_y = height - 1.25 * inch
    doc.drawString(0.9 * inch, header_y, f"Folio: {folio}")
    doc.drawString(3.0 * inch, header_y, f"Fecha: {quote_date.strftime('%d/%m/%Y')}")
    header_y -= 0.2 * inch
    doc.drawString(0.9 * inch, header_y, f"Ciudad: {city}")
    if vendor_name:
        header_y -= 0.2 * inch
        doc.drawString(0.9 * inch, header_y, f"Vendedor: {vendor_name}")
    if customer_name:
        header_y -= 0.2 * inch
        doc.drawString(0.9 * inch, header_y, f"Cliente: {customer_name}")

    data = [["SKU", "Descripción", "Unidad", "Cantidad", "P.U.", "Importe"]]
    for line in lines:
        data.append([
            line.sku,
            line.description[:58] + ("..." if len(line.description) > 58 else ""),
            line.unit,
            str(line.quantity),
            f"{line.unit_price:,.2f}",
            f"{line.line_total:,.2f}",
        ])

    table = Table(data, colWidths=[1.1 * inch, 3.0 * inch, 0.7 * inch, 0.8 * inch, 0.9 * inch, 1.0 * inch])
    table.setStyle(
        TableStyle(
            [
                ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 9),
                ("FONT", (0, 1), (-1, -1), "Helvetica", 8),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#94a3b8")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    table.wrapOn(doc, width - 1.8 * inch, height)
    table.drawOn(doc, 0.9 * inch, header_y - table._height - 0.35 * inch)

    totals_y = header_y - table._height - 0.65 * inch
    doc.setFont("Helvetica-Bold", 10)
    doc.drawRightString(width - 0.9 * inch, totals_y, f"Subtotal: {subtotal:,.2f}")
    doc.drawRightString(width - 0.9 * inch, totals_y - 0.2 * inch, f"IVA: {iva:,.2f}")
    doc.drawRightString(width - 0.9 * inch, totals_y - 0.4 * inch, f"Total: {total:,.2f}")

    if terms:
        terms_y = totals_y - 0.8 * inch
        doc.setFont("Helvetica-Bold", 10)
        doc.drawString(0.9 * inch, terms_y, "Términos y consideraciones")
        doc.setFont("Helvetica", 8)
        terms_y -= 0.22 * inch
        for term in terms:
            doc.drawString(1.0 * inch, terms_y, f"- {term}")
            terms_y -= 0.18 * inch
            if terms_y < 0.8 * inch:
                doc.showPage()
                terms_y = height - 0.9 * inch
                doc.setFont("Helvetica", 8)

    doc.save()


def _quote_base_dir(user: User, quote_id: str) -> Path:
    root = Path(settings.files_root).resolve()
    out_dir = (root / "catalog_quotes" / user.username / quote_id).resolve()
    if not str(out_dir).startswith(str(root)):
        raise HTTPException(status_code=400, detail="Ruta de salida inválida.")
    return out_dir


def _quote_output_dir(user: User, quote_id: str) -> Path:
    out_dir = _quote_base_dir(user, quote_id)
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def _resolve_quote_file(user: User, quote_id: str, kind: str) -> Path:
    if kind not in {"xlsx", "pdf"}:
        raise HTTPException(status_code=404, detail="Tipo de archivo no soportado.")
    out_dir = _quote_base_dir(user, quote_id)
    matches = sorted(out_dir.glob(f"*.{kind}"))
    if not matches:
        raise HTTPException(status_code=404, detail="Archivo no encontrado.")
    return matches[0]


@router.get("/vendors", response_model=list[VendorOut])
def list_vendors(
    db: Session = Depends(get_db),
    catalog_db: Session = Depends(_catalog_session),
    user: User = Depends(require_user),
):
    _ensure_catalog_assets()
    app = db.get(AppDefinition, CATALOG_APP_KEY)
    if not app or not app.enabled:
        raise HTTPException(status_code=404, detail="Cotizador no disponible.")
    ensure_app_access(user, CATALOG_APP_KEY, db)

    vendors = catalog_db.query(Vendor).order_by(Vendor.name.asc()).all()
    return [VendorOut(id=item.id, name=item.name) for item in vendors]


@router.get("/products", response_model=list[ProductOut])
def search_products(
    q: str = "",
    db: Session = Depends(get_db),
    catalog_db: Session = Depends(_catalog_session),
    user: User = Depends(require_user),
):
    _ensure_catalog_assets()
    app = db.get(AppDefinition, CATALOG_APP_KEY)
    if not app or not app.enabled:
        raise HTTPException(status_code=404, detail="Cotizador no disponible.")
    ensure_app_access(user, CATALOG_APP_KEY, db)

    query = (
        catalog_db.query(Product, Supplier.name.label("supplier_name"), PriceList.id.label("price_list_id"))
        .join(PriceList, Product.price_list_id == PriceList.id)
        .join(Supplier, PriceList.supplier_id == Supplier.id)
    )

    normalized = (q or "").strip()
    if normalized:
        for token in [item for item in normalized.split() if item]:
            like = f"%{token.lower()}%"
            query = query.filter(
                or_(func.lower(Product.sku).like(like), func.lower(Product.description).like(like))
            )

    rows = query.order_by(Product.sku.asc()).limit(80).all()
    return [
        ProductOut(
            sku=product.sku,
            description=product.description,
            unit=product.unit,
            category=product.category,
            supplier=supplier_name,
            price_list_id=price_list_id,
            has_container_offer=product.container_offer is not None,
        )
        for product, supplier_name, price_list_id in rows
    ]


@router.get("/products/{sku}", response_model=ProductDetailOut)
def product_detail(
    sku: str,
    db: Session = Depends(get_db),
    catalog_db: Session = Depends(_catalog_session),
    user: User = Depends(require_user),
):
    _ensure_catalog_assets()
    app = db.get(AppDefinition, CATALOG_APP_KEY)
    if not app or not app.enabled:
        raise HTTPException(status_code=404, detail="Cotizador no disponible.")
    ensure_app_access(user, CATALOG_APP_KEY, db)

    product = _latest_product_query(catalog_db, sku).first()
    if not product:
        raise HTTPException(status_code=404, detail="SKU no encontrado.")

    supplier_name = product.price_list.supplier.name if product.price_list and product.price_list.supplier else ""
    tiers = sorted(product.tiers, key=lambda item: item.min_qty)
    offer = product.container_offer

    return ProductDetailOut(
        sku=product.sku,
        description=product.description,
        unit=product.unit,
        category=product.category,
        supplier=supplier_name,
        price_list_id=product.price_list_id,
        has_container_offer=offer is not None,
        tiers=[TierOut(min_qty=item.min_qty, label=item.label, unit_price=item.unit_price) for item in tiers],
        container_qty=offer.container_qty if offer else None,
        container_price=offer.container_price if offer else None,
        container_notes=offer.notes if offer else None,
    )


@router.post("/quotes", response_model=QuoteOut)
def create_quote(
    payload: QuoteIn,
    db: Session = Depends(get_db),
    catalog_db: Session = Depends(_catalog_session),
    user: User = Depends(require_user),
):
    _ensure_catalog_assets()
    app = db.get(AppDefinition, CATALOG_APP_KEY)
    if not app or not app.enabled:
        raise HTTPException(status_code=404, detail="Cotizador no disponible.")
    ensure_app_access(user, CATALOG_APP_KEY, db)

    if not payload.items:
        raise HTTPException(status_code=400, detail="Agrega al menos una partida.")

    folio = _next_folio(db, payload.serie)
    quote_date = date.today()
    priced_lines: list[PricedLine] = []
    extra_terms: list[str] = []

    if payload.vendor_name:
        extra_terms.append(f"Vendedor: {payload.vendor_name}")
    if payload.customer_name:
        extra_terms.append(f"Cliente: {payload.customer_name}")

    try:
        for item in payload.items:
            product = _latest_product_query(catalog_db, item.sku).first()
            if not product:
                raise ValueError(f"SKU no encontrado: {item.sku}")

            priced = _price_line(product, item.quantity, item.mode)
            priced_lines.append(priced)

            if item.mode != "MAYOREO" and product.container_offer:
                offer = product.container_offer
                extra_terms.append(f"{product.sku}: Unidades por contenedor: {offer.container_qty} pzas.")
                extra_terms.append(f"{product.sku}: Total unidades informativas: {item.quantity * offer.container_qty} pzas.")
                if offer.notes:
                    extra_terms.append(f"{product.sku}: {offer.notes}")

        if payload.iva_mode == "excluded":
            extra_terms.append("Precios más IVA.")
        else:
            extra_terms.append("Los precios ya incluyen IVA.")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    subtotal = sum(item.line_total for item in priced_lines)
    base, iva, total = _compute_totals(subtotal, payload.iva_mode, payload.iva_rate)
    quote_id = uuid.uuid4().hex
    out_dir = _quote_output_dir(user, quote_id)
    xlsx_path = out_dir / f"{folio}.xlsx"
    pdf_path = out_dir / f"{folio}.pdf"

    unique_terms = list(dict.fromkeys(extra_terms))
    _render_quote_xlsx(xlsx_path, payload.city, quote_date, priced_lines, unique_terms)
    _render_quote_pdf(
        pdf_path=pdf_path,
        folio=folio,
        city=payload.city,
        quote_date=quote_date,
        lines=priced_lines,
        subtotal=base,
        iva=iva,
        total=total,
        vendor_name=payload.vendor_name,
        customer_name=payload.customer_name,
        terms=unique_terms,
    )

    return QuoteOut(
        quote_id=quote_id,
        folio=folio,
        date=quote_date,
        city=payload.city,
        vendor_name=payload.vendor_name,
        customer_name=payload.customer_name,
        lines=[
            QuoteLineOut(
                sku=item.sku,
                description=item.description,
                unit=item.unit,
                quantity=item.quantity,
                unit_price=item.unit_price,
                line_total=item.line_total,
                price_rule=item.rule,
            )
            for item in priced_lines
        ],
        totals=QuoteTotals(subtotal=base, iva=iva, total=total),
        download_xlsx_url=f"/tools/era/ventas/catalog-quote/quotes/{quote_id}/download/xlsx",
        download_pdf_url=f"/tools/era/ventas/catalog-quote/quotes/{quote_id}/download/pdf",
    )


@router.get("/quotes/{quote_id}/download/{kind}")
def download_quote(
    quote_id: str,
    kind: str,
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    app = db.get(AppDefinition, CATALOG_APP_KEY)
    if not app or not app.enabled:
        raise HTTPException(status_code=404, detail="Cotizador no disponible.")
    ensure_app_access(user, CATALOG_APP_KEY, db)

    file_path = _resolve_quote_file(user, quote_id, kind)
    media_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
    return FileResponse(path=str(file_path), media_type=media_type, filename=file_path.name)
