from __future__ import annotations

import copy
import io
import json
import re
import tempfile
import unicodedata
import zipfile
from collections import Counter
from dataclasses import asdict, dataclass, field, fields
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from sqlalchemy.orm import Session
from openpyxl.utils.cell import range_boundaries

from app.db.models import AppDefinition, User, UserAppPermission
from app.db.session import get_db
from app.deps import require_user

try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None


router = APIRouter(prefix="/tools/tesoreria/bank-movements", tags=["tools-tesoreria-bank-movements"])

TREASURY_APP_KEYS = (
    "tesoreria_automatizacion_saldos",
    "tesoreria_generacion_conciliacion",
)
DATE_SLASH_RX = re.compile(r"^\d{2}/\d{2}/\d{4}$")
DATE_MON_RX = re.compile(r"^\d{2}-[A-Za-z]{3}-\d{4}$")
SANTANDER_ACCOUNT_RX = re.compile(r"^\d{11}$")
SANTANDER_DATE_PART_RX = re.compile(r"^\d{5}$")
TIME_RX = re.compile(r"^\d{2}:\d{2}$")
BRANCH_RX = re.compile(r"^\d{3,4}$")
MONEY_ONLY_RX = re.compile(r"^-?\$?[0-9OIl]{1,3}(?:,[0-9OIl]{3})*(?:\.[0-9OIl]{2})$")
BAJIO_ROW_RX = re.compile(r"^\d+$")

MONEX_SKIP_LINES = {
    "Descripcion",
    "Descripción",
    "Fecha",
    "oper.",
    "liq.",
    "Emisora",
    "Serie",
    "Instrument",
    "o",
    "Referenc",
    "ia",
    "Titulos",
    "Contrato",
    "Cantidad",
    "Plaz",
    "Tasa",
    "rend",
    "Prima",
    "unitaria",
    "Precio",
    "strike",
    "Importe",
    "Movimientos",
    "Sistema Corporativo Monex",
}

SPANISH_MONTHS = {
    "ENE": 1,
    "FEB": 2,
    "MAR": 3,
    "ABR": 4,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AGO": 8,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DIC": 12,
    "DEC": 12,
}


@dataclass
class TreasuryMovement:
    statement_id: str
    source_file: str
    bank: str
    sequence: int
    account_number: str | None = None
    account_holder: str | None = None
    currency: str | None = None
    movement_date: str | None = None
    settlement_date: str | None = None
    statement_date: str | None = None
    time: str | None = None
    branch: str | None = None
    description: str | None = None
    concept: str | None = None
    long_description: str | None = None
    reference: str | None = None
    counterparty: str | None = None
    movement_type: str | None = None
    category: str | None = None
    debit: float | None = None
    credit: float | None = None
    balance: float | None = None
    raw_text: str | None = None


@dataclass
class TreasuryStatement:
    id: str
    source_file: str
    bank: str
    ocr_used: bool
    account_holder: str | None = None
    account_number: str | None = None
    clabe: str | None = None
    contract: str | None = None
    alias: str | None = None
    currency: str | None = None
    period_start: str | None = None
    period_end: str | None = None
    period_label: str | None = None
    statement_date: str | None = None
    opening_balance: float | None = None
    closing_balance: float | None = None
    total_debits: float | None = None
    total_credits: float | None = None
    warnings: list[str] = field(default_factory=list)
    raw_text: str = ""
    movements: list[TreasuryMovement] = field(default_factory=list)


MOVEMENT_FIELD_LABELS = {
    "sheet_name": "Hoja destino",
    "movement_type": "Tipo de movimiento",
    "company": "Empresa",
    "payee": "A nombre de",
    "group": "Grupo",
    "business_unit": "Unidad de negocio",
    "project": "Obra",
    "reconciliation": "Conciliación",
    "specific_concept": "Concepto específico",
    "detailed_concept": "Concepto detallado",
    "observations": "Observaciones",
}

MOVEMENT_FIELD_ALIASES = {
    "movement_type": (
        "tipo de movimiento",
        "tipo de moviento",
    ),
    "date": ("fecha",),
    "company": ("empresa",),
    "cashbox": ("caja",),
    "check_number": ("n chq", "n chq.", "no cheque", "no. cheque"),
    "payee": ("a nombre de",),
    "group": ("grupo",),
    "business_unit": ("unidad de negocio",),
    "project": ("obra",),
    "reconciliation": ("conciliacion",),
    "specific_concept": ("concepto especifico",),
    "detailed_concept": ("concepto detallado",),
    "deposits": ("depositos",),
    "withdrawals": ("retiros",),
    "breakdown": ("desglose", "desglose de retiros"),
    "balance": ("saldo",),
    "observations": ("observaciones", "columna1"),
}

MOVEMENT_EDITABLE_FIELDS = (
    "movement_type",
    "company",
    "payee",
    "group",
    "business_unit",
    "project",
    "reconciliation",
    "specific_concept",
    "detailed_concept",
    "observations",
)

TREASURY_STOPWORDS = {
    "de",
    "del",
    "la",
    "las",
    "los",
    "por",
    "para",
    "con",
    "una",
    "uno",
    "que",
    "spei",
    "pago",
    "banco",
    "transferencia",
    "transferencias",
    "deposito",
    "depositos",
    "abono",
    "cargo",
    "com",
    "ref",
}

CATEGORY_TO_RECONCILIATION = {
    "iva_comision": "COMISIONES BANCARIAS",
    "comision": "COMISIONES BANCARIAS",
    "intereses_credito": "INTERESES",
    "intereses": "INTERESES",
    "nomina": "NÓMINA",
    "cheque": "CHEQUE",
    "divisas": "DIVISAS",
    "transferencia_entrada": "TRASPASO",
    "transferencia_salida": "TRASPASO",
    "deposito": "DEPÓSITO",
    "prestamo_credito": "PRÉSTAMO",
}

DEFAULT_MOVEMENT_TYPES = ("TRANSFERENCIA", "INVERSIÓN", "DEPÓSITO", "CHEQUE", "CARGO", "ABONO")


def _slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_only).strip("-").lower()
    return slug or "statement"


def _normalize(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "")
    value = value.encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"\s+", " ", value.strip().lower())
    return value


def _normalize_spaces(value: str | None) -> str | None:
    if value is None:
        return None
    return " ".join(value.replace("\xa0", " ").split()).strip() or None


def _clean_money_token(value: str) -> str:
    value = value.strip().replace("$", "").replace(" ", "")
    return value.translate(str.maketrans({"O": "0", "o": "0", "I": "1", "l": "1"}))


def _parse_money(value: str | None) -> float | None:
    if not value:
        return None
    try:
        return float(_clean_money_token(value).replace(",", ""))
    except Exception:
        return None


def _parse_date(value: str | None) -> str | None:
    raw = _normalize_spaces(value)
    if not raw:
        return None

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            pass

    mon_match = re.match(r"^(?P<day>\d{2})-(?P<mon>[A-Za-z]{3})-(?P<year>\d{4})$", raw)
    if mon_match:
        parts = mon_match.groupdict()
        month = SPANISH_MONTHS.get(parts["mon"].upper())
        if month:
            return datetime(int(parts["year"]), month, int(parts["day"])).date().isoformat()
    return None


def _parse_santander_date(part_a: str, part_b: str) -> str | None:
    digits = re.sub(r"\D", "", f"{part_a}{part_b}")
    if len(digits) != 8:
        return None
    return _parse_date(f"{digits[:2]}/{digits[2:4]}/{digits[4:]}")


def _find_label_value(lines: list[str], labels: tuple[str, ...], lookahead: int = 2) -> str | None:
    normalized_labels = tuple(_normalize(label) for label in labels)
    for idx, line in enumerate(lines):
        current = _normalize(line)
        for label in normalized_labels:
            if current.startswith(f"{label}:"):
                value = line.split(":", 1)[1].strip()
                return _normalize_spaces(value)
            if current == label and idx + 1 < len(lines):
                for probe in lines[idx + 1 : idx + 1 + lookahead]:
                    value = _normalize_spaces(probe)
                    if value:
                        return value
    return None


def _split_trailing_money(line: str) -> tuple[str, list[str]]:
    parts = line.split()
    tail: list[str] = []
    while parts and MONEY_ONLY_RX.match(parts[-1]):
        tail.insert(0, parts.pop())
        if len(tail) == 2:
            break
    return " ".join(parts).strip(), tail


def _extract_reference(text: str | None) -> str | None:
    raw = _normalize_spaces(text)
    if not raw:
        return None

    for pattern in (
        r"(CH-\d+)",
        r"(REF[A-Z0-9_/-]+)",
        r"(CRE_[A-Z0-9_]+)",
        r"(BB\d{10,})",
        r"(\b\d{8,18}\b)",
    ):
        match = re.search(pattern, raw, re.IGNORECASE)
        if match:
            return match.group(1)
    return None


def _extract_counterparty(text: str | None) -> str | None:
    raw = _normalize_spaces(text)
    if not raw:
        return None

    patterns = (
        r"(?:ordenante|nombre del ordenante|emisor)\s*:\s*(.+?)(?:\s+\|\s+| cuenta | rfc | clave | referencia[: ]| fecha[: ]| hora[: ]|$)",
        r"(?:beneficiario|nombre beneficiario|nombre receptor)\s*:\s*(.+?)(?:\s+\|\s+| cuenta | rfc | clave | referencia[: ]| fecha[: ]| hora[: ]|$)",
        r"(?:ordenante)\s*:\s*(.+?)(?:\s+cuenta|\s+rfc|\s+hora[: ]|\s+referencia[: ]|$)",
    )

    for pattern in patterns:
        match = re.search(pattern, raw, re.IGNORECASE)
        if match:
            value = _normalize_spaces(match.group(1))
            if value:
                value = re.split(r"\b(?:Cuenta|RFC|Clave|Referencia|Fecha|Hora)\b\s*:?", value, maxsplit=1, flags=re.IGNORECASE)[0].strip()
            if value:
                return value
    return None


def _movement_type(debit: float | None, credit: float | None) -> str:
    if debit and debit > 0:
        return "cargo"
    if credit and credit > 0:
        return "abono"
    return "informativo"


def _movement_category(description: str | None, concept: str | None, debit: float | None, credit: float | None) -> str:
    haystack = _normalize(" ".join(part for part in [description or "", concept or ""] if part))
    haystack = haystack.replace("i v a", "iva")

    if "iva" in haystack and "comision" in haystack:
        return "iva_comision"
    if "comision" in haystack or "membresia" in haystack or "administracion paquete" in haystack:
        return "comision"
    if "interes" in haystack and "credito" in haystack:
        return "intereses_credito"
    if "interes" in haystack:
        return "intereses"
    if "nomina" in haystack:
        return "nomina"
    if "cheque" in haystack:
        return "cheque"
    if "divisa" in haystack or "compra de divisas" in haystack:
        return "divisas"
    if "spei recibido" in haystack or "abono transferencia spei" in haystack:
        return "transferencia_entrada"
    if "spei enviado" in haystack or "spid" in haystack or ("transferencia" in haystack and debit):
        return "transferencia_salida"
    if "deposito" in haystack or "deposito de tercero" in haystack:
        return "deposito"
    if "prestamo" in haystack or "credito" in haystack:
        return "prestamo_credito"
    return _movement_type(debit, credit)


def _make_movement(
    statement: TreasuryStatement,
    sequence: int,
    *,
    currency: str | None = None,
    movement_date: str | None = None,
    settlement_date: str | None = None,
    time: str | None = None,
    branch: str | None = None,
    description: str | None = None,
    concept: str | None = None,
    long_description: str | None = None,
    reference: str | None = None,
    counterparty: str | None = None,
    debit: float | None = None,
    credit: float | None = None,
    balance: float | None = None,
    raw_text: str | None = None,
) -> TreasuryMovement:
    description = _normalize_spaces(description)
    concept = _normalize_spaces(concept)
    long_description = _normalize_spaces(long_description)
    raw_text = _normalize_spaces(raw_text)
    reference = _normalize_spaces(reference) or _extract_reference(" ".join(part for part in [description or "", concept or "", long_description or ""] if part))
    counterparty = _normalize_spaces(counterparty) or _extract_counterparty(" ".join(part for part in [description or "", concept or "", long_description or ""] if part))
    movement_type = _movement_type(debit, credit)
    category = _movement_category(description, concept or long_description, debit, credit)

    return TreasuryMovement(
        statement_id=statement.id,
        source_file=statement.source_file,
        bank=statement.bank,
        sequence=sequence,
        account_number=statement.account_number,
        account_holder=statement.account_holder,
        currency=currency or statement.currency,
        movement_date=movement_date,
        settlement_date=settlement_date,
        statement_date=statement.statement_date,
        time=time,
        branch=branch,
        description=description,
        concept=concept,
        long_description=long_description,
        reference=reference,
        counterparty=counterparty,
        movement_type=movement_type,
        category=category,
        debit=debit,
        credit=credit,
        balance=balance,
        raw_text=raw_text,
    )


def _detect_bank(text: str) -> str:
    normalized = _normalize(text[:2500])
    if "bbva net cash" in normalized or normalized.startswith("bbva "):
        return "BBVA"
    if "banregio" in normalized:
        return "Banregio"
    if "banbajio" in normalized or "cuenta conecta banbajio" in normalized:
        return "BanBajio"
    if "corporativo monex" in normalized or ("cliente:" in normalized and "movimientos de:" in normalized and "contrato:" in normalized):
        return "Monex"
    if "santander" in normalized or "contrato cmc" in normalized:
        return "Santander"
    return "Desconocido"


def _extract_pdf_text(path: Path) -> tuple[str, bool]:
    if fitz is None:
        raise RuntimeError("PyMuPDF no esta disponible en la API.")

    document = fitz.open(str(path))
    used_ocr = False
    page_texts: list[str] = []
    try:
        for page in document:
            plain_text = page.get_text("text").replace("\xa0", " ")
            compact_plain = re.sub(r"\s+", "", plain_text)
            final_text = plain_text

            if len(compact_plain) < 120:
                try:
                    textpage = page.get_textpage_ocr(language="spa+eng", dpi=300, full=True)
                    ocr_text = page.get_text("text", textpage=textpage).replace("\xa0", " ")
                    if len(re.sub(r"\s+", "", ocr_text)) > len(compact_plain):
                        final_text = ocr_text
                        used_ocr = True
                except Exception:
                    pass

            page_texts.append(final_text)
    finally:
        document.close()

    return "\n".join(page_texts), used_ocr


def _clean_lines(text: str) -> list[str]:
    return [line.strip() for line in text.splitlines() if line and line.strip()]


def _find_first_prefixed(lines: list[str], prefixes: tuple[str, ...]) -> str | None:
    for line in lines:
        for prefix in prefixes:
            if line.startswith(prefix):
                return _normalize_spaces(line.split(":", 1)[1])
    return None


def _parse_bbva(text: str, source_file: str, ocr_used: bool) -> TreasuryStatement:
    lines = _clean_lines(text)
    statement = TreasuryStatement(
        id=_slugify(Path(source_file).stem),
        source_file=source_file,
        bank="BBVA",
        ocr_used=ocr_used,
        account_holder=_find_label_value(lines, ("Nombre del Cliente",)),
        contract=_find_label_value(lines, ("No. Contrato",)),
        account_number=_find_label_value(lines, ("Cuenta",)),
        alias=_find_label_value(lines, ("Alias",)),
        currency=_find_label_value(lines, ("Divisa",)),
        statement_date=_parse_date(_find_label_value(lines, ("Fecha Consulta",))),
        period_label=_find_label_value(lines, ("Periodo de Consulta",)),
        raw_text=text,
    )
    if ocr_used:
        statement.warnings.append("Se uso OCR para leer este estado de cuenta.")

    start_idx = next((idx for idx, line in enumerate(lines) if _normalize(line) == "detalle de movimientos"), 0)
    body = lines[start_idx + 1 :]
    sequence = 1
    idx = 0
    while idx < len(body):
        line = body[idx]
        match = re.match(r"^(?P<date>\d{2}/\d{2}/\d{4})(?:\s+(?P<rest>.*))?$", line)
        if not match:
            idx += 1
            continue

        movement_date = _parse_date(match.group("date"))
        idx += 1
        desc_lines: list[str] = [match.group("rest").strip()] if match.group("rest") else []
        money_tokens: list[str] = []

        while idx < len(body):
            probe = body[idx]
            if re.match(r"^\d{2}/\d{2}/\d{4}\b", probe) or probe.startswith("Movimientos del dia:") or probe.startswith("Movimientos del día:") or probe.startswith("Fin dia") or probe.startswith("Fin día"):
                break
            if probe in {"Fecha", "Concepto/ Referencia", "Cargo", "Abono", "Saldo"}:
                idx += 1
                continue

            if MONEY_ONLY_RX.match(probe):
                money_tokens.append(probe)
            else:
                cleaned, trailing = _split_trailing_money(probe)
                if cleaned:
                    desc_lines.append(cleaned)
                money_tokens.extend(trailing)
            idx += 1

        if not money_tokens and not desc_lines:
            continue

        values = money_tokens[-2:] if len(money_tokens) >= 2 else money_tokens
        amount = _parse_money(values[0]) if values else None
        balance = _parse_money(values[1]) if len(values) > 1 else None
        debit = abs(amount) if amount is not None and amount < 0 else None
        credit = amount if amount is not None and amount > 0 else None
        description = _normalize_spaces(" ".join(desc_lines))
        if not description:
            continue

        statement.movements.append(
            _make_movement(
                statement,
                sequence,
                movement_date=movement_date,
                description=description,
                debit=debit,
                credit=credit,
                balance=balance,
                raw_text=" | ".join(desc_lines + money_tokens),
            )
        )
        sequence += 1

    credits = [item.credit for item in statement.movements if item.credit]
    debits = [item.debit for item in statement.movements if item.debit]
    if credits:
        statement.total_credits = round(sum(credits), 2)
    if debits:
        statement.total_debits = round(sum(debits), 2)
    if statement.movements:
        statement.closing_balance = statement.movements[-1].balance
    return statement


def _parse_banregio(text: str, source_file: str, ocr_used: bool) -> TreasuryStatement:
    lines = _clean_lines(text)
    period_match = re.search(r"Fecha\s+Inicio[: ]*([0-9/]+)\s*-\s*Fecha\s+Fin[: ]*([0-9/]+)", text, re.IGNORECASE)
    statement = TreasuryStatement(
        id=_slugify(Path(source_file).stem),
        source_file=source_file,
        bank="Banregio",
        ocr_used=ocr_used,
        account_holder=_normalize_spaces(lines[2]) if len(lines) > 2 else None,
        account_number=_find_label_value(lines, ("CUENTA", "Cuenta")),
        clabe=_find_label_value(lines, ("CLABE", "Clabe")),
        period_start=_parse_date(period_match.group(1)) if period_match else None,
        period_end=_parse_date(period_match.group(2)) if period_match else None,
        opening_balance=_parse_money(re.search(r"Saldo Inicial[: ]*\$?([0-9,]+\.[0-9]{2})", text, re.IGNORECASE).group(1)) if re.search(r"Saldo Inicial[: ]*\$?([0-9,]+\.[0-9]{2})", text, re.IGNORECASE) else None,
        closing_balance=_parse_money(re.search(r"Saldo Final[: ]*\$?([0-9,]+\.[0-9]{2})", text, re.IGNORECASE).group(1)) if re.search(r"Saldo Final[: ]*\$?([0-9,]+\.[0-9]{2})", text, re.IGNORECASE) else None,
        total_debits=_parse_money(re.search(r"Total Cargos[: ]*\$?([0-9,]+\.[0-9]{2})", text, re.IGNORECASE).group(1)) if re.search(r"Total Cargos[: ]*\$?([0-9,]+\.[0-9]{2})", text, re.IGNORECASE) else None,
        total_credits=_parse_money(re.search(r"Total Abonos[: ]*\$?([0-9,]+\.[0-9]{2})", text, re.IGNORECASE).group(1)) if re.search(r"Total Abonos[: ]*\$?([0-9,]+\.[0-9]{2})", text, re.IGNORECASE) else None,
        currency="MXN",
        raw_text=text,
    )
    if ocr_used:
        statement.warnings.append("Se uso OCR para leer este estado de cuenta.")

    if "_sin registros" in _normalize(text) or "sin registros" in _normalize(text):
        statement.warnings.append("El PDF de ejemplo no contiene movimientos.")
        return statement

    body = lines[next((idx for idx, line in enumerate(lines) if _normalize(line) == "saldo"), len(lines)) + 1 :]
    idx = 0
    sequence = 1
    while idx < len(body):
        line = body[idx]
        if not DATE_SLASH_RX.match(line):
            idx += 1
            continue
        movement_date = _parse_date(line)
        idx += 1
        bucket: list[str] = []
        while idx < len(body) and not DATE_SLASH_RX.match(body[idx]) and "total cargos" not in _normalize(body[idx]):
            bucket.append(body[idx])
            idx += 1
        text_block = " ".join(bucket)
        money = [part for part in bucket if MONEY_ONLY_RX.match(part)]
        debit = _parse_money(money[0]) if len(money) >= 1 else None
        credit = _parse_money(money[1]) if len(money) >= 2 else None
        balance = _parse_money(money[2]) if len(money) >= 3 else None
        statement.movements.append(
            _make_movement(
                statement,
                sequence,
                movement_date=movement_date,
                description=text_block,
                debit=debit if debit and debit > 0 else None,
                credit=credit if credit and credit > 0 else None,
                balance=balance,
                raw_text=text_block,
            )
        )
        sequence += 1
    return statement


def _parse_bajio(text: str, source_file: str, ocr_used: bool) -> TreasuryStatement:
    lines = _clean_lines(text)
    statement = TreasuryStatement(
        id=_slugify(Path(source_file).stem),
        source_file=source_file,
        bank="BanBajio",
        ocr_used=ocr_used,
        account_number=_find_label_value(lines, ("Cuenta",)),
        currency="MXN" if "Saldo Total" in text else None,
        closing_balance=_parse_money(re.search(r"Saldo Total[: ]*\$?([0-9,]+\.[0-9]{2})", text, re.IGNORECASE).group(1)) if re.search(r"Saldo Total[: ]*\$?([0-9,]+\.[0-9]{2})", text, re.IGNORECASE) else None,
        raw_text=text,
    )
    if ocr_used:
        statement.warnings.append("Se uso OCR para leer este estado de cuenta.")

    sequence = 1
    idx = 0
    while idx < len(lines):
        if not BAJIO_ROW_RX.match(lines[idx]):
            idx += 1
            continue
        if idx + 1 >= len(lines) or not _parse_date(lines[idx + 1]):
            idx += 1
            continue

        row_number = int(lines[idx])
        movement_date = _parse_date(lines[idx + 1])
        idx += 2
        block: list[str] = []
        while idx < len(lines):
            if BAJIO_ROW_RX.match(lines[idx]) and idx + 1 < len(lines) and _parse_date(lines[idx + 1]):
                break
            if "registros" in _normalize(lines[idx]):
                break
            block.append(lines[idx])
            idx += 1

        amount_tokens = [item for item in block if MONEY_ONLY_RX.match(item)]
        balance = _parse_money(amount_tokens[-1]) if amount_tokens else None
        amount = _parse_money(amount_tokens[0]) if amount_tokens else None
        text_block = " ".join(block)
        normalized_block = _normalize(text_block)
        is_credit = "spei recibido" in normalized_block
        debit = None
        credit = None
        if amount and amount > 0:
            if is_credit:
                credit = amount
            else:
                debit = amount

        description_parts = [
            item
            for item in block
            if not MONEY_ONLY_RX.match(item)
            and not item.startswith("Hora:")
            and not item.startswith("Referencia:")
            and not item.startswith("Numero de Referencia:")
            and not item.startswith("Número de Referencia:")
            and not item.startswith("Numero de Autorizacion:")
            and not item.startswith("Número de Autorización:")
            and not item.startswith("Recibo")
        ]
        description = _normalize_spaces(" ".join(description_parts))
        reference = _find_first_prefixed(block, ("Referencia:", "Número de Referencia:", "Numero de Referencia:"))
        if not reference:
            reference = _extract_reference(text_block)
        time = _find_first_prefixed(block, ("Hora:",))
        counterparty = _extract_counterparty(text_block)

        statement.movements.append(
            _make_movement(
                statement,
                row_number or sequence,
                movement_date=movement_date,
                time=time,
                description=description,
                reference=reference,
                counterparty=counterparty,
                debit=debit,
                credit=credit,
                balance=balance,
                raw_text=text_block,
            )
        )
        sequence = max(sequence, row_number + 1)

    if not statement.movements:
        statement.warnings.append("No se detectaron filas del detalle en el PDF de ejemplo.")
    else:
        incomplete = sum(1 for item in statement.movements if item.debit is None and item.credit is None)
        if incomplete:
            statement.warnings.append(f"{incomplete} movimientos requieren validacion manual de importe.")
    return statement


def _parse_monex(text: str, source_file: str, ocr_used: bool) -> TreasuryStatement:
    lines = _clean_lines(text)
    period_match = re.search(r"del dia\s+([0-9/]+)\s+al\s+([0-9/]+)", _normalize(text), re.IGNORECASE)
    statement = TreasuryStatement(
        id=_slugify(Path(source_file).stem),
        source_file=source_file,
        bank="Monex",
        ocr_used=ocr_used,
        account_holder=_find_label_value(lines, ("Cliente",)),
        contract=_find_label_value(lines, ("Contrato",)),
        clabe=_find_label_value(lines, ("Clabe",)),
        period_start=_parse_date(period_match.group(1)) if period_match else None,
        period_end=_parse_date(period_match.group(2)) if period_match else None,
        raw_text=text,
    )

    sequence = 1
    currency: str | None = None
    idx = 0

    def looks_new(start: int) -> bool:
        line = lines[start]
        if (
            not line
            or DATE_SLASH_RX.match(line)
            or MONEY_ONLY_RX.match(line)
            or line.startswith("MOVIMIENTOS DE:")
            or line.startswith("Movimientos del dia:")
            or line.startswith("Movimientos del día:")
            or line in {"Inicio dia", "Inicio día", "Fin dia", "Fin día"}
            or line.startswith("Fecha Emision:")
            or line.startswith("Fecha Emisión:")
            or line.startswith("Fecha Operacion:")
            or line.startswith("Fecha Operación:")
            or line.startswith("Pagina ")
            or line.startswith("Página ")
            or line in MONEX_SKIP_LINES
        ):
            return False
        for offset in range(1, 8):
            probe_a = start + offset
            probe_b = probe_a + 1
            if probe_b < len(lines) and DATE_SLASH_RX.match(lines[probe_a]) and DATE_SLASH_RX.match(lines[probe_b]):
                return True
        return True

    while idx < len(lines):
        line = lines[idx]
        if line.startswith("MOVIMIENTOS DE:"):
            currency = _normalize_spaces(line.split(":", 1)[1])
            idx += 1
            continue
        if line.startswith("Movimientos del dia:") or line.startswith("Movimientos del día:"):
            idx += 1
            continue
        if line in {"Inicio dia", "Inicio día", "Fin dia", "Fin día"}:
            idx += 2
            continue
        if line.startswith(("Fecha Emision:", "Fecha Emisión:", "Fecha Operacion:", "Fecha Operación:", "Pagina ", "Página ")) or line in MONEX_SKIP_LINES:
            idx += 1
            continue
        if line.isdigit() and len(line) <= 2:
            idx += 1
            continue
        if not looks_new(idx):
            idx += 1
            continue

        desc_lines: list[str] = []
        while idx < len(lines) and not DATE_SLASH_RX.match(lines[idx]):
            current = lines[idx]
            if current not in MONEX_SKIP_LINES and not current.startswith(("Fecha Emision:", "Fecha Emisión:", "Fecha Operacion:", "Fecha Operación:", "Pagina ", "Página ")):
                desc_lines.append(current)
            idx += 1

        if idx + 1 >= len(lines) or not DATE_SLASH_RX.match(lines[idx]) or not DATE_SLASH_RX.match(lines[idx + 1]):
            continue

        movement_date = _parse_date(lines[idx])
        settlement_date = _parse_date(lines[idx + 1])
        idx += 2
        detail_lines: list[str] = []
        while idx < len(lines):
            probe = lines[idx]
            if probe.startswith("MOVIMIENTOS DE:") or probe.startswith("Movimientos del dia:") or probe.startswith("Movimientos del día:") or probe in {"Inicio dia", "Inicio día", "Fin dia", "Fin día"}:
                break
            if probe.startswith(("Fecha Emision:", "Fecha Emisión:", "Fecha Operacion:", "Fecha Operación:", "Pagina ", "Página ")) or probe in MONEX_SKIP_LINES:
                idx += 1
                continue
            if DATE_SLASH_RX.match(probe) and idx + 1 < len(lines) and DATE_SLASH_RX.match(lines[idx + 1]):
                break
            detail_lines.append(probe)
            idx += 1

        amount = next((_parse_money(item) for item in reversed(detail_lines) if MONEY_ONLY_RX.match(item)), None)
        reference = next(
            (
                item
                for item in detail_lines
                if item not in {"0", "0.000000"}
                and (item.isdigit() or item.startswith("DIVISA"))
            ),
            None,
        )
        debit = abs(amount) if amount is not None and amount < 0 else None
        credit = amount if amount is not None and amount > 0 else None
        description = _normalize_spaces(" ".join(desc_lines))
        raw_text_block = _normalize_spaces(" ".join(desc_lines + detail_lines))
        concept = None
        if detail_lines:
            concept = _normalize_spaces(" ".join(item for item in detail_lines if not MONEY_ONLY_RX.match(item)))

        statement.movements.append(
            _make_movement(
                statement,
                sequence,
                currency=currency,
                movement_date=movement_date,
                settlement_date=settlement_date,
                description=description,
                concept=concept,
                reference=reference,
                debit=debit,
                credit=credit,
                raw_text=raw_text_block,
            )
        )
        sequence += 1

    if not statement.movements:
        statement.warnings.append("No se pudieron estructurar movimientos de Monex.")
    else:
        incomplete = sum(1 for item in statement.movements if item.debit is None and item.credit is None)
        if incomplete:
            statement.warnings.append(f"{incomplete} movimientos de Monex requieren revision manual de importe.")
        currencies = sorted({item.currency for item in statement.movements if item.currency})
        if len(currencies) > 1:
            statement.currency = "MULTI"
    return statement


def _parse_santander(text: str, source_file: str, ocr_used: bool) -> TreasuryStatement:
    lines = _clean_lines(text)
    contract_value = _find_label_value(lines, ("Contrato CMC",))
    statement = TreasuryStatement(
        id=_slugify(Path(source_file).stem),
        source_file=source_file,
        bank="Santander",
        ocr_used=ocr_used,
        contract=contract_value,
        account_number=_find_label_value(lines, ("Numero de Cuenta", "Número de Cuenta")),
        account_holder=_normalize_spaces(contract_value.split(" ", 1)[1]) if contract_value and " " in contract_value else None,
        period_start=_parse_date(re.search(r"Periodo:\s*([0-9/]+)", text).group(1)) if re.search(r"Periodo:\s*([0-9/]+)", text) else None,
        period_end=_parse_date(re.search(r"Periodo:\s*[0-9/]+\s+al\s+([0-9/]+)", text).group(1)) if re.search(r"Periodo:\s*[0-9/]+\s+al\s+([0-9/]+)", text) else None,
        opening_balance=_parse_money(re.search(r"Saldo Inicial:\s*\$([0-9,]+\.[0-9]{2})", text).group(1)) if re.search(r"Saldo Inicial:\s*\$([0-9,]+\.[0-9]{2})", text) else None,
        closing_balance=_parse_money(re.search(r"Saldo Final:\s*\$([0-9,]+\.[0-9]{2})", text).group(1)) if re.search(r"Saldo Final:\s*\$([0-9,]+\.[0-9]{2})", text) else None,
        total_credits=_parse_money(re.search(r"Importe Total Abonos:\s*\$([0-9,]+\.[0-9]{2})", text).group(1)) if re.search(r"Importe Total Abonos:\s*\$([0-9,]+\.[0-9]{2})", text) else None,
        total_debits=_parse_money(re.search(r"Importe Total Cargos:\s*\$([0-9,]+\.[0-9]{2})", text).group(1)) if re.search(r"Importe Total Cargos:\s*\$([0-9,]+\.[0-9]{2})", text) else None,
        currency="MXN",
        raw_text=text,
    )

    sequence = 1
    idx = 0
    while idx < len(lines):
        if not SANTANDER_ACCOUNT_RX.match(lines[idx]):
            idx += 1
            continue
        if idx + 7 >= len(lines) or not SANTANDER_DATE_PART_RX.match(lines[idx + 1]) or not TIME_RX.match(lines[idx + 3]) or not BRANCH_RX.match(lines[idx + 4]):
            idx += 1
            continue

        movement_date = _parse_santander_date(lines[idx + 1], lines[idx + 2])
        time = lines[idx + 3]
        branch = lines[idx + 4]
        cursor = idx + 5
        desc_lines: list[str] = []
        while cursor < len(lines) and not MONEY_ONLY_RX.match(lines[cursor]):
            if SANTANDER_ACCOUNT_RX.match(lines[cursor]):
                break
            desc_lines.append(lines[cursor])
            cursor += 1

        if cursor + 2 >= len(lines):
            idx = cursor
            continue

        debit = _parse_money(lines[cursor]) or None
        credit = _parse_money(lines[cursor + 1]) or None
        balance = _parse_money(lines[cursor + 2])
        cursor += 3
        reference = lines[cursor] if cursor < len(lines) else None
        cursor += 1
        extra_lines: list[str] = []
        while cursor < len(lines) and not SANTANDER_ACCOUNT_RX.match(lines[cursor]):
            normalized = _normalize(lines[cursor])
            if normalized.startswith("para dudas o aclaraciones") or normalized.startswith("banco santander") or normalized.startswith("enlace http") or normalized.startswith("este documento no es"):
                break
            extra_lines.append(lines[cursor])
            cursor += 1

        concept = extra_lines[0] if extra_lines else None
        long_description = " ".join(extra_lines[1:]) if len(extra_lines) > 1 else None
        statement.movements.append(
            _make_movement(
                statement,
                sequence,
                movement_date=movement_date,
                time=time,
                branch=branch,
                description=" ".join(desc_lines),
                concept=concept,
                long_description=long_description,
                reference=reference,
                debit=debit,
                credit=credit,
                balance=balance,
                raw_text=" | ".join(desc_lines + [reference or ""] + extra_lines),
            )
        )
        sequence += 1
        idx = cursor

    return statement


def _parse_statement(path: Path) -> TreasuryStatement:
    text, ocr_used = _extract_pdf_text(path)
    bank = _detect_bank(text)
    if bank == "BBVA":
        statement = _parse_bbva(text, path.name, ocr_used)
    elif bank == "Banregio":
        statement = _parse_banregio(text, path.name, ocr_used)
    elif bank == "BanBajio":
        statement = _parse_bajio(text, path.name, ocr_used)
    elif bank == "Monex":
        statement = _parse_monex(text, path.name, ocr_used)
    elif bank == "Santander":
        statement = _parse_santander(text, path.name, ocr_used)
    else:
        statement = TreasuryStatement(
            id=_slugify(path.stem),
            source_file=path.name,
            bank=bank,
            ocr_used=ocr_used,
            raw_text=text,
            warnings=["No se reconocio el banco del PDF."],
        )

    if not statement.movements:
        statement.warnings.append("No se detectaron movimientos estructurados en este archivo.")
    return statement


def _analysis_payload(statements: list[TreasuryStatement]) -> dict[str, Any]:
    flat_movements = [asdict(item) for statement in statements for item in statement.movements]
    banks = sorted({statement.bank for statement in statements})
    accounts = sorted({item["account_number"] for item in flat_movements if item.get("account_number")})

    return {
        "summary": {
            "statements": len(statements),
            "movements": len(flat_movements),
            "banks": banks,
            "accounts": len(accounts),
            "ocr_statements": sum(1 for item in statements if item.ocr_used),
        },
        "columns": [
            "bank",
            "source_file",
            "account_number",
            "account_holder",
            "currency",
            "movement_date",
            "settlement_date",
            "time",
            "movement_type",
            "category",
            "description",
            "concept",
            "reference",
            "counterparty",
            "debit",
            "credit",
            "balance",
        ],
        "statements": [asdict(item) for item in statements],
        "movements": flat_movements,
    }


def _template_ext_ok(filename: str | None) -> bool:
    return Path(filename or "").suffix.lower() in {".xlsx", ".xlsm"}


def _ensure_treasury_access(db: Session, user: User) -> None:
    available_keys = [
        app_key
        for app_key in TREASURY_APP_KEYS
        if (app := db.get(AppDefinition, app_key)) and app.enabled
    ]
    if not available_keys:
        raise HTTPException(status_code=404, detail="La herramienta de Tesorería no está disponible.")

    if user.is_admin:
        return

    has_permission = (
        db.query(UserAppPermission.id)
        .filter(UserAppPermission.user_id == user.id, UserAppPermission.app_key.in_(available_keys))
        .first()
    )
    if not has_permission:
        raise HTTPException(status_code=403, detail="No tienes permiso para esta herramienta de Tesorería.")


def _statement_from_analysis_payload(payload: dict[str, Any]) -> TreasuryStatement:
    movement_values = []
    for raw_movement in payload.get("movements") or []:
        if not isinstance(raw_movement, dict):
            continue
        movement_values.append(
            TreasuryMovement(
                **{
                    field_def.name: raw_movement.get(field_def.name)
                    for field_def in fields(TreasuryMovement)
                    if field_def.name in raw_movement
                }
            )
        )

    statement_values = {
        field_def.name: payload.get(field_def.name)
        for field_def in fields(TreasuryStatement)
        if field_def.name != "movements" and field_def.name in payload
    }
    return TreasuryStatement(**statement_values, movements=movement_values)


def _statements_from_analysis_json(analysis_json: str | None) -> list[TreasuryStatement] | None:
    if not analysis_json or not analysis_json.strip():
        return None

    try:
        payload = json.loads(analysis_json)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el análisis enviado: {exc}") from exc

    if not isinstance(payload, dict) or not isinstance(payload.get("statements"), list):
        raise HTTPException(status_code=400, detail="El análisis enviado no tiene el formato esperado.")

    statements = [
        _statement_from_analysis_payload(item)
        for item in payload["statements"]
        if isinstance(item, dict)
    ]
    if not statements:
        raise HTTPException(status_code=400, detail="El análisis enviado no contiene estados de cuenta.")
    return statements


def _as_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return _normalize_spaces(str(value))


def _iso_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return _parse_date(_as_text(value))


def _float_value(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    return _parse_money(str(value))


def _header_name(value: Any) -> str:
    text = _normalize_spaces(str(value or "")) or ""
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[°º#./_()-]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def _canonical_field_name(value: Any) -> str | None:
    normalized = _header_name(value)
    for field, aliases in MOVEMENT_FIELD_ALIASES.items():
        for alias in aliases:
            if normalized == alias:
                return field
    return None


def _extract_digit_tokens(*values: str | None) -> list[str]:
    tokens: list[str] = []
    for value in values:
        raw = re.sub(r"\D", "", value or "")
        if len(raw) >= 4:
            tokens.append(raw)
            tokens.append(raw[-4:])
            tokens.append(raw[-6:])
    seen: set[str] = set()
    ordered: list[str] = []
    for token in tokens:
        if token and token not in seen:
            seen.add(token)
            ordered.append(token)
    return ordered


def _search_tokens(*values: str | None) -> set[str]:
    text = " ".join(part for part in values if part)
    normalized = _normalize(text)
    tokens = set()
    for token in re.split(r"[^a-z0-9]+", normalized):
        if len(token) < 3 or token in TREASURY_STOPWORDS:
            continue
        tokens.add(token)
    return tokens


def _statement_label(statement: TreasuryStatement) -> str:
    return statement.account_number or statement.contract or statement.source_file


def _movement_amount(movement: TreasuryMovement) -> float:
    if movement.credit is not None:
        return round(float(movement.credit), 2)
    if movement.debit is not None:
        return round(float(-movement.debit), 2)
    return 0.0


def _source_movement_signature(
    *,
    movement_date: str | None,
    amount: float | None,
    detail: str | None,
) -> str:
    detail_norm = _normalize_spaces(detail) or ""
    amount_label = f"{round(float(amount or 0.0), 2):.2f}"
    return "|".join([movement_date or "", amount_label, _normalize(detail_norm)])


def _movement_signature(movement: TreasuryMovement) -> str:
    detail = " ".join(part for part in [movement.description, movement.concept, movement.long_description] if part)
    return _source_movement_signature(
        movement_date=movement.movement_date,
        amount=_movement_amount(movement),
        detail=detail,
    )


def _sheet_bank_hint(title: str) -> str | None:
    normalized = _normalize(title)
    if "bbva" in normalized:
        return "BBVA"
    if "banreg" in normalized:
        return "Banregio"
    if "bajio" in normalized:
        return "BanBajio"
    if "monex" in normalized:
        return "Monex"
    if "santander" in normalized:
        return "Santander"
    if "scotia" in normalized:
        return "Scotia"
    return None


def _sheet_kind(title: str) -> str:
    normalized = _normalize(title)
    if any(token in normalized for token in ("dll", "usd", "dolares", "dolares", "usd")):
        return "usd"
    if any(token in normalized for token in ("inv", "inver", "inversion", "fondo")):
        return "investment"
    return "operational"


def _statement_kind(statement: TreasuryStatement) -> str:
    normalized = _normalize(
        " ".join(
            part
            for part in [
                statement.source_file,
                statement.currency,
                statement.account_number,
                statement.contract,
                statement.raw_text[:1800],
            ]
            if part
        )
    )
    if any(token in normalized for token in ("usd", "dll", "dolar", "dolares", "dolares")):
        return "usd"
    if any(token in normalized for token in ("fondo", "inversion", "inver", "mesa dinero", "bursatil")):
        return "investment"
    return "operational"


def _movement_type_default(movement: TreasuryMovement, sheet_name: str | None = None) -> str:
    if sheet_name and _sheet_kind(sheet_name) == "investment":
        return "INVERSIÓN"
    if movement.category == "deposito":
        return "DEPÓSITO"
    if movement.category == "cheque":
        return "CHEQUE"
    return "TRANSFERENCIA"


def _default_reconciliation(movement: TreasuryMovement) -> str | None:
    if movement.category in CATEGORY_TO_RECONCILIATION:
        return CATEGORY_TO_RECONCILIATION[movement.category]
    if movement.movement_type == "cargo":
        return "EGRESO"
    if movement.movement_type == "abono":
        return "INGRESO"
    return None


def _default_payee(movement: TreasuryMovement) -> str | None:
    if movement.counterparty:
        return movement.counterparty
    if movement.reference and movement.reference.isdigit():
        return None
    if movement.description:
        return movement.description[:120]
    return movement.bank


def _field_value_counts(history_rows: list[dict[str, Any]]) -> dict[str, Counter[str]]:
    counts = {field: Counter() for field in MOVEMENT_EDITABLE_FIELDS}
    counts["movement_type"].update(DEFAULT_MOVEMENT_TYPES)
    for row in history_rows:
        values = row["values"]
        for field in MOVEMENT_EDITABLE_FIELDS:
            value = _normalize_spaces(values.get(field))
            if value:
                counts[field][value] += 1
    return counts


def _find_sheet_header(ws) -> tuple[int, int, int]:
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        row_values = [ws.cell(row_idx, col).value for col in range(1, min(ws.max_column, 20) + 1)]
        headers = [_canonical_field_name(value) for value in row_values]
        if {"date", "deposits", "withdrawals"} <= {item for item in headers if item}:
            return row_idx, 1, min(ws.max_column, 20)
    raise ValueError(f"No se pudo detectar el encabezado de la hoja {ws.title}.")


def _history_label(values: dict[str, str | None]) -> str | None:
    parts = [values.get("payee"), values.get("reconciliation"), values.get("detailed_concept")]
    label = " · ".join(part for part in parts if part)
    return label[:140] if label else None


def _read_sheet_profile(ws) -> dict[str, Any]:
    table_name = next(iter(ws.tables.keys()), None) if ws.tables else None
    if table_name:
        table = ws.tables[table_name]
        start_col, start_row, end_col, end_row = range_boundaries(table.ref)
    else:
        start_row, start_col, end_col = _find_sheet_header(ws)
        end_row = ws.max_row

    field_columns: dict[str, int] = {}
    for col in range(start_col, end_col + 1):
        field_name = _canonical_field_name(ws.cell(start_row, col).value)
        if field_name and field_name not in field_columns:
            field_columns[field_name] = col

    history_rows: list[dict[str, Any]] = []
    existing_signatures: set[str] = set()
    data_start_row = start_row + 1
    scan_end_row = max(end_row, ws.max_row)
    for row_idx in range(data_start_row, scan_end_row + 1):
        row_values: dict[str, Any] = {
            field: ws.cell(row_idx, column).value for field, column in field_columns.items()
        }
        has_content = any(
            row_values.get(field) not in (None, "")
            for field in ("date", "payee", "detailed_concept", "deposits", "withdrawals", "reconciliation", "group")
        )
        if not has_content:
            continue

        payload = {
            field: _normalize_spaces(_as_text(row_values.get(field))) for field in MOVEMENT_EDITABLE_FIELDS
        }
        payload["date"] = _iso_date(row_values.get("date"))
        payload["deposits"] = _float_value(row_values.get("deposits"))
        payload["withdrawals"] = _float_value(row_values.get("withdrawals"))
        payload["row_number"] = row_idx
        payload["search_tokens"] = _search_tokens(
            payload.get("payee"),
            payload.get("reconciliation"),
            payload.get("specific_concept"),
            payload.get("detailed_concept"),
            payload.get("observations"),
        )
        history_rows.append(
            {
                "row_number": row_idx,
                "values": payload,
                "label": _history_label(payload),
            }
        )

        amount = (payload.get("deposits") or 0.0) - (payload.get("withdrawals") or 0.0)
        existing_signatures.add(
            _source_movement_signature(
                movement_date=payload.get("date"),
                amount=amount,
                detail=payload.get("detailed_concept"),
            )
        )

    value_counts = _field_value_counts(history_rows)
    field_options = {
        field: [value for value, _ in value_counts[field].most_common(24)] for field in MOVEMENT_EDITABLE_FIELDS
    }
    defaults = {
        field: (field_options[field][0] if field_options[field] else None) for field in MOVEMENT_EDITABLE_FIELDS
    }
    defaults["movement_type"] = defaults["movement_type"] or _movement_type_default(
        TreasuryMovement(statement_id="", source_file="", bank="", sequence=0),
        ws.title,
    )

    return {
        "name": ws.title,
        "bank_hint": _sheet_bank_hint(ws.title),
        "sheet_kind": _sheet_kind(ws.title),
        "table_name": table_name,
        "table_start_col": start_col,
        "table_end_col": end_col,
        "table_end_row": end_row,
        "header_row": start_row,
        "data_start_row": data_start_row,
        "field_columns": field_columns,
        "field_options": field_options,
        "defaults": defaults,
        "history_rows": history_rows,
        "existing_signatures": existing_signatures,
    }


def _score_statement_sheet(statement: TreasuryStatement, profile: dict[str, Any]) -> int:
    score = 0
    if profile["bank_hint"] == statement.bank:
        score += 20
    elif profile["bank_hint"]:
        return -100

    statement_digits = _extract_digit_tokens(statement.account_number, statement.clabe, statement.contract, statement.source_file)
    title_normalized = _normalize(profile["name"])
    if any(token and token in re.sub(r"\D", "", title_normalized) for token in statement_digits if len(token) >= 4):
        score += 12
    elif any(token and token in title_normalized for token in statement_digits if len(token) == 4):
        score += 10

    statement_kind = _statement_kind(statement)
    if profile["sheet_kind"] == statement_kind:
        score += 6
    elif profile["sheet_kind"] != "operational" and statement_kind != "operational":
        score -= 4

    raw_hint = _normalize(" ".join(part for part in [statement.source_file, statement.raw_text[:600]] if part))
    if "obra publica" in raw_hint and "op" in title_normalized:
        score += 4
    return score


def _rank_statement_sheets(statement: TreasuryStatement, profiles: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ranked = [
        {
            "name": profile["name"],
            "score": _score_statement_sheet(statement, profile),
            "bank_hint": profile["bank_hint"],
        }
        for profile in profiles
    ]
    ranked.sort(key=lambda item: (-item["score"], item["name"]))
    bank_matches = [item for item in ranked if item["bank_hint"] == statement.bank]
    if bank_matches:
        return bank_matches
    useful = [item for item in ranked if item["score"] > -100]
    return useful or ranked


def _match_history_row(profile: dict[str, Any], movement: TreasuryMovement) -> tuple[dict[str, Any] | None, int]:
    movement_tokens = _search_tokens(
        movement.counterparty,
        movement.description,
        movement.concept,
        movement.long_description,
        movement.reference,
    )
    if not movement_tokens:
        return None, 0

    best_row: dict[str, Any] | None = None
    best_score = 0
    for row in profile["history_rows"]:
        row_tokens = row["values"]["search_tokens"]
        if not row_tokens:
            continue
        overlap = len(movement_tokens & row_tokens)
        if overlap == 0:
            continue
        score = overlap * 2
        row_payee = _normalize(row["values"].get("payee") or "")
        if movement.counterparty and row_payee and _normalize(movement.counterparty) == row_payee:
            score += 8
        if movement.reference and movement.reference in (row["values"].get("detailed_concept") or ""):
            score += 4
        amount_delta = abs(
            ((row["values"].get("deposits") or 0.0) - (row["values"].get("withdrawals") or 0.0))
            - _movement_amount(movement)
        )
        if amount_delta <= 0.01:
            score += 3
        if score > best_score:
            best_row = row
            best_score = score
    return best_row, best_score


def _initial_draft_values(
    statement: TreasuryStatement,
    movement: TreasuryMovement,
    sheet_name: str | None,
    profile: dict[str, Any] | None,
    history_row: dict[str, Any] | None,
) -> dict[str, Any]:
    values = {
        "movement_type": None,
        "date": movement.movement_date,
        "company": None,
        "payee": None,
        "group": None,
        "business_unit": None,
        "project": None,
        "reconciliation": None,
        "specific_concept": None,
        "detailed_concept": None,
        "deposits": movement.credit,
        "withdrawals": movement.debit,
        "breakdown": movement.debit if movement.debit else None,
        "observations": None,
    }

    if history_row:
        for field in MOVEMENT_EDITABLE_FIELDS:
            values[field] = history_row["values"].get(field)

    if profile:
        for field in MOVEMENT_EDITABLE_FIELDS:
            if not values.get(field):
                values[field] = profile["defaults"].get(field)

    values["movement_type"] = values["movement_type"] or _movement_type_default(movement, sheet_name)
    values["company"] = values["company"] or statement.account_holder or (profile["defaults"].get("company") if profile else None) or "DEESA"
    values["payee"] = values["payee"] or _default_payee(movement)
    values["reconciliation"] = values["reconciliation"] or _default_reconciliation(movement)
    values["specific_concept"] = values["specific_concept"] or values["reconciliation"]
    values["detailed_concept"] = values["detailed_concept"] or _normalize_spaces(
        " ".join(part for part in [movement.description, movement.concept, movement.long_description] if part)
    )
    return values


def _missing_draft_fields(values: dict[str, Any], profile: dict[str, Any] | None, sheet_name: str | None) -> list[str]:
    missing: list[str] = []
    if not sheet_name:
        missing.append("sheet_name")

    review_fields = ("movement_type", "company", "payee", "reconciliation", "detailed_concept")
    if profile:
        review_fields += tuple(
            field
            for field in ("group", "business_unit", "project")
            if field in profile["field_columns"]
        )

    seen: set[str] = set()
    for field in review_fields:
        if field in seen:
            continue
        seen.add(field)
        if not _normalize_spaces(_as_text(values.get(field))):
            missing.append(field)
    return missing


def _build_movement_draft(
    statement: TreasuryStatement,
    movement: TreasuryMovement,
    sheet_name: str | None,
    sheet_options: list[str],
    profile: dict[str, Any] | None,
) -> dict[str, Any]:
    history_row, history_score = _match_history_row(profile, movement) if profile else (None, 0)
    values = _initial_draft_values(statement, movement, sheet_name, profile, history_row)
    missing_fields = _missing_draft_fields(values, profile, sheet_name)
    suggestion_source = "historial" if history_row and history_score >= 4 else "automático"

    return {
        "draft_id": f"{statement.id}-{movement.sequence}",
        "statement_id": statement.id,
        "statement_label": _statement_label(statement),
        "sheet_name": sheet_name,
        "sheet_options": sheet_options,
        "suggestion_source": suggestion_source,
        "suggestion_score": history_score,
        "matched_history_label": history_row.get("label") if history_row else None,
        "missing_fields": missing_fields,
        "needs_review": bool(missing_fields) or suggestion_source != "historial",
        "movement": {
            "sequence": movement.sequence,
            "bank": movement.bank,
            "account_number": movement.account_number,
            "movement_date": movement.movement_date,
            "description": movement.description,
            "concept": movement.concept,
            "reference": movement.reference,
            "counterparty": movement.counterparty,
            "debit": movement.debit,
            "credit": movement.credit,
            "balance": movement.balance,
        },
        "values": values,
    }


def _prepare_movement_template(template_path: Path, statements: list[TreasuryStatement]) -> dict[str, Any]:
    workbook = load_workbook(template_path)
    profiles = [_read_sheet_profile(ws) for ws in workbook.worksheets]
    profile_map = {profile["name"]: profile for profile in profiles}

    drafts: list[dict[str, Any]] = []
    unmatched_statements: list[str] = []
    skipped_duplicates = 0
    review_count = 0

    for statement in statements:
        ranked = _rank_statement_sheets(statement, profiles)
        sheet_options = [item["name"] for item in ranked if item["score"] > -100]
        if not sheet_options:
            sheet_options = [profile["name"] for profile in profiles]
        best_score = ranked[0]["score"] if ranked else -100
        selected_sheet = sheet_options[0] if sheet_options and best_score > -100 else None
        if not selected_sheet:
            unmatched_statements.append(_statement_label(statement))
        profile = profile_map.get(selected_sheet) if selected_sheet else None

        for movement in statement.movements:
            source_signature = _movement_signature(movement)
            if profile and source_signature in profile["existing_signatures"]:
                skipped_duplicates += 1
                continue

            draft = _build_movement_draft(statement, movement, selected_sheet, sheet_options, profile)
            drafts.append(draft)
            if draft["needs_review"]:
                review_count += 1
            if profile:
                profile["existing_signatures"].add(source_signature)

    return {
        "filename": template_path.name,
        "sheets": [
            {
                "name": profile["name"],
                "bank_hint": profile["bank_hint"],
                "sheet_kind": profile["sheet_kind"],
                "field_options": profile["field_options"],
                "defaults": profile["defaults"],
            }
            for profile in profiles
        ],
        "drafts": drafts,
        "review_count": review_count,
        "skipped_duplicates": skipped_duplicates,
        "unmatched_statements": unmatched_statements,
    }


def _score_balance_row(statement: TreasuryStatement, bank_value: str | None, account_label: str | None, clabe_value: str | None) -> tuple[int, str]:
    row_text = " ".join(part for part in [bank_value, account_label, clabe_value] if part)
    normalized_row = _normalize(row_text)
    score = 0
    reason_parts: list[str] = []

    bank_hint = _sheet_bank_hint(bank_value or "")
    if bank_hint == statement.bank:
        score += 18
        reason_parts.append("banco")

    statement_digits = _extract_digit_tokens(statement.account_number, statement.clabe, statement.contract, statement.source_file)
    row_digits = _extract_digit_tokens(account_label, clabe_value, bank_value)
    if any(token in row_digits for token in statement_digits if len(token) >= 4):
        score += 12
        reason_parts.append("cuenta")
    elif any(token in normalized_row for token in statement_digits if len(token) == 4):
        score += 8
        reason_parts.append("últimos 4 dígitos")

    row_kind = _sheet_kind(row_text)
    statement_kind = _statement_kind(statement)
    if row_kind == statement_kind:
        score += 4
        reason_parts.append("tipo de cuenta")
    elif row_kind != "operational" and statement_kind != "operational":
        score -= 3

    return score, ", ".join(reason_parts) or "coincidencia general"


def _prepare_balance_template(template_path: Path, statements: list[TreasuryStatement]) -> dict[str, Any]:
    workbook = load_workbook(template_path)
    ws = workbook[workbook.sheetnames[0]]

    updates: list[dict[str, Any]] = []
    matched_statement_ids: set[str] = set()
    for row_idx in range(1, ws.max_row + 1):
        bank_value = _normalize_spaces(_as_text(ws.cell(row_idx, 4).value))
        account_label = _normalize_spaces(_as_text(ws.cell(row_idx, 5).value))
        clabe_value = _normalize_spaces(_as_text(ws.cell(row_idx, 6).value))
        if not bank_value and not account_label:
            continue

        candidates = []
        for statement in statements:
            if statement.closing_balance is None:
                continue
            score, reason = _score_balance_row(statement, bank_value, account_label, clabe_value)
            if score <= 0:
                continue
            candidates.append((score, reason, statement))
        if not candidates:
            continue

        score, reason, statement = sorted(candidates, key=lambda item: (-item[0], item[2].source_file))[0]
        if score < 18:
            continue

        is_dollar_row = _sheet_kind(" ".join(part for part in [bank_value, account_label] if part)) == "usd"
        column_key = "dolares" if is_dollar_row else "pesos"
        current_value = _float_value(ws.cell(row_idx, 9 if is_dollar_row else 8).value)
        updates.append(
            {
                "id": f"balance-{row_idx}",
                "row_number": row_idx,
                "bank": bank_value,
                "account_label": account_label,
                "column_key": column_key,
                "current_value": current_value,
                "new_value": round(float(statement.closing_balance), 2),
                "statement_id": statement.id,
                "statement_label": _statement_label(statement),
                "confidence": round(min(score / 34, 0.99), 2),
                "reason": reason,
                "enabled": True,
            }
        )
        matched_statement_ids.add(statement.id)

    return {
        "filename": template_path.name,
        "sheet_name": ws.title,
        "updates": updates,
        "unmatched_statements": [
            _statement_label(statement) for statement in statements if statement.id not in matched_statement_ids
        ],
    }


def _row_is_blank_for_insert(ws, profile: dict[str, Any], row_idx: int) -> bool:
    for field in ("date", "payee", "detailed_concept", "deposits", "withdrawals"):
        column = profile["field_columns"].get(field)
        if column is None:
            continue
        value = ws.cell(row_idx, column).value
        if value not in (None, ""):
            return False
    return True


def _clone_template_row(ws, profile: dict[str, Any], source_row: int, target_row: int) -> None:
    for col_idx in range(profile["table_start_col"], profile["table_end_col"] + 1):
        source = ws.cell(source_row, col_idx)
        target = ws.cell(target_row, col_idx)
        target._style = copy.copy(source._style)
        target.number_format = source.number_format
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.protection = copy.copy(source.protection)
        if isinstance(source.value, str) and source.value.startswith("="):
            target.value = Translator(source.value, origin=source.coordinate).translate_formula(target.coordinate)
        else:
            target.value = None


def _ensure_target_row(ws, profile: dict[str, Any]) -> int:
    for row_idx in range(profile["data_start_row"], profile["table_end_row"] + 1):
        if _row_is_blank_for_insert(ws, profile, row_idx):
            if row_idx > ws.max_row:
                source_row = max(profile["data_start_row"], min(ws.max_row, row_idx - 1))
                _clone_template_row(ws, profile, source_row, row_idx)
            return row_idx

    source_row = max(profile["data_start_row"], min(ws.max_row, profile["table_end_row"]))
    target_row = profile["table_end_row"] + 1
    _clone_template_row(ws, profile, source_row, target_row)
    if profile["table_name"]:
        table = ws.tables[profile["table_name"]]
        start_col, start_row, end_col, _ = range_boundaries(table.ref)
        table.ref = f"{ws.cell(start_row, start_col).coordinate}:{ws.cell(target_row, end_col).coordinate}"
    profile["table_end_row"] = target_row
    return target_row


def _write_draft_row(ws, profile: dict[str, Any], row_idx: int, draft: dict[str, Any]) -> None:
    values = draft["values"]
    for field, column in profile["field_columns"].items():
        cell = ws.cell(row_idx, column)
        if field == "date":
            if values.get("date"):
                cell.value = datetime.strptime(values["date"], "%Y-%m-%d")
        elif field == "deposits":
            if values.get("deposits") is not None:
                cell.value = round(float(values["deposits"]), 2)
        elif field == "withdrawals":
            if values.get("withdrawals") is not None:
                cell.value = round(float(values["withdrawals"]), 2)
        elif field == "breakdown":
            if values.get("breakdown") is not None:
                cell.value = round(float(values["breakdown"]), 2)
        elif field == "balance":
            if not (isinstance(cell.value, str) and cell.value.startswith("=")) and draft["movement"].get("balance") is not None:
                cell.value = round(float(draft["movement"]["balance"]), 2)
        elif field in values and values.get(field):
            cell.value = values[field]


def _render_movement_workbook(template_path: Path, drafts: list[dict[str, Any]]) -> bytes:
    workbook = load_workbook(template_path)
    profiles = {ws.title: _read_sheet_profile(ws) for ws in workbook.worksheets}

    grouped: dict[str, list[dict[str, Any]]] = {}
    for draft in drafts:
        if not draft.get("sheet_name"):
            continue
        grouped.setdefault(draft["sheet_name"], []).append(draft)

    for sheet_name, sheet_drafts in grouped.items():
        if sheet_name not in profiles:
            continue
        ws = workbook[sheet_name]
        profile = profiles[sheet_name]
        for draft in sorted(sheet_drafts, key=lambda item: (item["values"].get("date") or "", item["draft_id"])):
            row_idx = _ensure_target_row(ws, profile)
            _write_draft_row(ws, profile, row_idx, draft)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _render_balance_workbook(template_path: Path, updates: list[dict[str, Any]]) -> bytes:
    workbook = load_workbook(template_path)
    ws = workbook[workbook.sheetnames[0]]
    for update in updates:
        if not update.get("enabled", True):
            continue
        row_idx = int(update["row_number"])
        column_idx = 9 if update.get("column_key") == "dolares" else 8
        ws.cell(row_idx, column_idx).value = round(float(update["new_value"]), 2)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _save_upload(upload: UploadFile, root: Path, fallback_name: str) -> Path:
    safe_name = re.sub(r"[^A-Za-z0-9._ -]+", "_", Path(upload.filename or fallback_name).name)
    target = root / safe_name
    target.write_bytes(upload.file.read())
    upload.file.seek(0)
    return target


async def _parse_uploaded_statements(files: list[UploadFile], root: Path) -> list[TreasuryStatement]:
    statements: list[TreasuryStatement] = []
    for index, upload in enumerate(files):
        safe_name = re.sub(r"[^A-Za-z0-9._ -]+", "_", Path(upload.filename or f"file_{index}.pdf").name)
        target = root / f"{index:02d}-{safe_name}"
        target.write_bytes(await upload.read())
        statements.append(_parse_statement(target))
    return statements


async def _resolve_input_statements(
    files: list[UploadFile] | None,
    analysis_json: str | None,
    root: Path,
) -> list[TreasuryStatement]:
    statements = _statements_from_analysis_json(analysis_json)
    if statements is not None:
        return statements

    valid_files = [item for item in (files or []) if item.filename]
    if not valid_files:
        raise HTTPException(status_code=400, detail="Debes subir al menos un PDF.")

    for item in valid_files:
        if Path(item.filename or "").suffix.lower() != ".pdf":
            raise HTTPException(status_code=400, detail=f"Solo se permiten archivos PDF. Recibí: {item.filename}")

    return await _parse_uploaded_statements(valid_files, root)


@router.post("/analyze")
async def analyze_bank_movements(
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    _ensure_treasury_access(db, user)

    valid_files = [item for item in files if item.filename]
    if not valid_files:
        raise HTTPException(status_code=400, detail="Debes subir al menos un PDF.")

    for item in valid_files:
        if Path(item.filename or "").suffix.lower() != ".pdf":
            raise HTTPException(status_code=400, detail=f"Solo se permiten archivos PDF. Recibí: {item.filename}")

    with tempfile.TemporaryDirectory(prefix="bank-movements-") as temp_dir:
        root = Path(temp_dir)
        statements = await _parse_uploaded_statements(valid_files, root)

    return _analysis_payload(statements)


@router.post("/prepare")
async def prepare_bank_templates(
    files: list[UploadFile] | None = File(default=None),
    movements_template: UploadFile | None = File(None),
    balances_template: UploadFile | None = File(None),
    analysis_json: str | None = Form(default=None),
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    _ensure_treasury_access(db, user)

    if movements_template and not _template_ext_ok(movements_template.filename):
        raise HTTPException(status_code=400, detail="El Excel de movimientos debe ser .xlsx o .xlsm.")
    if balances_template and not _template_ext_ok(balances_template.filename):
        raise HTTPException(status_code=400, detail="El Excel de saldos debe ser .xlsx o .xlsm.")

    with tempfile.TemporaryDirectory(prefix="bank-movements-prepare-") as temp_dir:
        root = Path(temp_dir)
        statements = await _resolve_input_statements(files, analysis_json, root)

        movement_template_data = None
        balance_template_data = None
        if movements_template and movements_template.filename:
            movement_path = _save_upload(movements_template, root, "movimientos.xlsx")
            movement_template_data = _prepare_movement_template(movement_path, statements)
        if balances_template and balances_template.filename:
            balance_path = _save_upload(balances_template, root, "saldos.xlsx")
            balance_template_data = _prepare_balance_template(balance_path, statements)

    return {
        "analysis": _analysis_payload(statements),
        "movement_template": movement_template_data,
        "balances_template": balance_template_data,
    }


@router.post("/export")
async def export_bank_templates(
    movements_template: UploadFile | None = File(None),
    balances_template: UploadFile | None = File(None),
    drafts_json: str = Form("[]"),
    balance_updates_json: str = Form("[]"),
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    _ensure_treasury_access(db, user)

    if not movements_template and not balances_template:
        raise HTTPException(status_code=400, detail="Debes subir al menos uno de los Excel para exportar.")
    if movements_template and not _template_ext_ok(movements_template.filename):
        raise HTTPException(status_code=400, detail="El Excel de movimientos debe ser .xlsx o .xlsm.")
    if balances_template and not _template_ext_ok(balances_template.filename):
        raise HTTPException(status_code=400, detail="El Excel de saldos debe ser .xlsx o .xlsm.")

    try:
        drafts = json.loads(drafts_json or "[]")
        balance_updates = json.loads(balance_updates_json or "[]")
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo leer la configuración enviada: {exc}") from exc

    with tempfile.TemporaryDirectory(prefix="bank-movements-export-") as temp_dir:
        root = Path(temp_dir)
        rendered_files: list[tuple[str, bytes]] = []

        if movements_template and movements_template.filename:
            movement_path = _save_upload(movements_template, root, "movimientos.xlsx")
            movement_bytes = _render_movement_workbook(movement_path, drafts if isinstance(drafts, list) else [])
            movement_name = f"{movement_path.stem}_actualizado{movement_path.suffix}"
            rendered_files.append((movement_name, movement_bytes))

        if balances_template and balances_template.filename:
            balance_path = _save_upload(balances_template, root, "saldos.xlsx")
            balance_bytes = _render_balance_workbook(balance_path, balance_updates if isinstance(balance_updates, list) else [])
            balance_name = f"{balance_path.stem}_actualizado{balance_path.suffix}"
            rendered_files.append((balance_name, balance_bytes))

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for filename, content in rendered_files:
            archive.writestr(filename, content)
    zip_buffer.seek(0)

    headers = {"Content-Disposition": 'attachment; filename="tesoreria_actualizada.zip"'}
    return StreamingResponse(zip_buffer, media_type="application/zip", headers=headers)
