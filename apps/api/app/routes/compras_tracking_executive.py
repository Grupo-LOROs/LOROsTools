from __future__ import annotations

from collections import Counter
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import openpyxl
import re
import tempfile
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.utils import column_index_from_string
from reportlab.graphics.shapes import Circle, Drawing, Line, Path, Rect, String
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.core.config import settings
from app.db.models import AppDefinition, Job, User
from app.db.session import get_db
from app.deps import ensure_app_access, require_user
from app.routes.compras_tracking import (
    IMPORTACIONES_SOURCE_APP_KEY,
    TRACKING_APP_KEY,
    ShipmentTracking,
    _latest_tracking_date,
    _load_importaciones_history,
    _normalize_header,
    _save_upload,
    _shipment_identifiers,
    _to_iso_date,
)


router = APIRouter(
    prefix="/tools/compras/importaciones-tracking/executive",
    tags=["tools-compras-tracking-executive"],
)

EXECUTIVE_STAGE_META = (
    ("planned", "Programación", "#94a3b8"),
    ("in_transit", "En tránsito", "#0ea5e9"),
    ("arrival", "Arribo a puerto", "#14b8a6"),
    ("pedimento", "Pedimento", "#f59e0b"),
    ("dispatch", "Despacho", "#f97316"),
    ("delivered", "Entregado", "#22c55e"),
)
EXECUTIVE_STAGE_INDEX = {key: idx for idx, (key, _, _) in enumerate(EXECUTIVE_STAGE_META)}

PORT_GEO_RULES = (
    (("qingdao",), ("Qingdao, China", 36.0671, 120.3826)),
    (("shanghai",), ("Shanghái, China", 31.2304, 121.4737)),
    (("ningbo",), ("Ningbo, China", 29.8683, 121.5440)),
    (("shenzhen", "yantian"), ("Shenzhen, China", 22.5431, 114.0579)),
    (("xiamen",), ("Xiamen, China", 24.4798, 118.0894)),
    (("tianjin",), ("Tianjin, China", 39.0842, 117.2000)),
    (("lazaro cardenas", "lazaro"), ("Lázaro Cárdenas, México", 17.9583, -102.1947)),
    (("manzanillo",), ("Manzanillo, México", 19.0501, -104.3188)),
    (("veracruz",), ("Veracruz, México", 19.1738, -96.1342)),
    (("altamira",), ("Altamira, México", 22.3928, -97.9381)),
    (("ensenada",), ("Ensenada, México", 31.8667, -116.5967)),
    (("long beach",), ("Long Beach, EUA", 33.7701, -118.1937)),
    (("los angeles",), ("Los Ángeles, EUA", 34.0522, -118.2437)),
)


@dataclass
class ExecutiveLocation:
    label: str
    lat: float
    lng: float


@dataclass
class ExecutiveMilestone:
    key: str
    label: str
    date: str | None
    status: str


@dataclass
class ExecutiveRoute:
    key: str
    origin_label: str
    destination_label: str
    count: int
    active: int
    delivered: int
    color: str
    origin: ExecutiveLocation
    destination: ExecutiveLocation


@dataclass
class ExecutiveAlert:
    shipment_id: str
    level: str
    title: str
    detail: str


@dataclass
class ExecutiveShipment:
    id: str
    order_number: str | None = None
    general_po: str | None = None
    invoice_number: str | None = None
    supplier_display: str | None = None
    supplier_name: str | None = None
    container: str | None = None
    visa_reference: str | None = None
    terminal: str | None = None
    forwarder: str | None = None
    transportista: str | None = None
    despacho: str | None = None
    goods_summary: str | None = None
    pedimento: str | None = None
    iva: str | None = None
    warehouse: str | None = None
    container_status: str | None = None
    provider_payment_due: str | None = None
    provider_payment_status: str | None = None
    forwarder_payment_due: str | None = None
    forwarder_payment_status: str | None = None
    etd: str | None = None
    eta: str | None = None
    storage_deadline: str | None = None
    pedimento_charge_date: str | None = None
    dispatch_date: str | None = None
    warehouse_request_date: str | None = None
    delay_reference: str | None = None
    origin_port: str | None = None
    destination_port: str | None = None
    incoterm: str | None = None
    total_usd: float | None = None
    source_updated_at: str | None = None
    stage_key: str = "planned"
    stage_label: str = "Programación"
    stage_color: str = "#94a3b8"
    progress_pct: int = 0
    attention_level: str = "ok"
    attention_reason: str | None = None
    next_event_label: str | None = None
    next_event_date: str | None = None
    days_to_eta: int | None = None
    origin_geo: ExecutiveLocation | None = None
    destination_geo: ExecutiveLocation | None = None
    milestones: list[ExecutiveMilestone] | None = None


def _coerce_date(value: Any) -> date | None:
    iso = _to_iso_date(value)
    if not iso:
        return None
    try:
        return date.fromisoformat(iso)
    except ValueError:
        return None


def _status_style(level: str) -> tuple[str, str]:
    if level == "risk":
        return ("Riesgo", "#dc2626")
    if level == "watch":
        return ("Atención", "#d97706")
    return ("En curso", "#0f766e")


def _stage_meta(stage_key: str) -> tuple[str, str]:
    for key, label, color in EXECUTIVE_STAGE_META:
        if key == stage_key:
            return label, color
    return ("Programación", "#94a3b8")


def _geocode_port(label: str | None) -> ExecutiveLocation | None:
    normalized = _normalize_header(label)
    if not normalized:
        return None
    for aliases, (resolved_label, lat, lng) in PORT_GEO_RULES:
        if any(alias in normalized for alias in aliases):
            return ExecutiveLocation(label=resolved_label, lat=lat, lng=lng)
    return None


def _infer_destination_from_terminal(terminal: str | None) -> str | None:
    normalized = _normalize_header(terminal)
    if "apm" in normalized:
        return "Lázaro Cárdenas, México"
    if "manzanillo" in normalized:
        return "Manzanillo, México"
    return None


def _excel_value(sheet, row: int, column: int, cache: dict[str, Any]) -> Any:
    cell = sheet.cell(row=row, column=column)
    coordinate = cell.coordinate
    if coordinate in cache:
        return cache[coordinate]

    raw = cell.value
    if not isinstance(raw, str) or not raw.startswith("="):
        cache[coordinate] = raw
        return raw

    cache[coordinate] = None
    formula = raw.replace(" ", "").upper()

    offset_match = re.match(r"^=\+?([A-Z]{1,3})(\d+)([+-])(\d+)$", formula)
    if offset_match:
        ref_col, ref_row, sign, delta = offset_match.groups()
        base = _excel_value(sheet, int(ref_row), column_index_from_string(ref_col), cache)
        if base is None:
            return None
        amount = int(delta) * (1 if sign == "+" else -1)
        if isinstance(base, datetime):
            resolved = base + timedelta(days=amount)
        elif isinstance(base, date):
            resolved = datetime.combine(base + timedelta(days=amount), datetime.min.time())
        elif isinstance(base, (int, float)):
            resolved = base + amount
        else:
            resolved = None
        cache[coordinate] = resolved
        return resolved

    weeknum_match = re.match(r"^=\+?WEEKNUM\(([A-Z]{1,3})(\d+)\)$", formula)
    if weeknum_match:
        ref_col, ref_row = weeknum_match.groups()
        base = _excel_value(sheet, int(ref_row), column_index_from_string(ref_col), cache)
        base_date = _coerce_date(base)
        resolved = base_date.isocalendar().week if base_date else None
        cache[coordinate] = resolved
        return resolved

    ref_match = re.match(r"^=\+?([A-Z]{1,3})(\d+)$", formula)
    if ref_match:
        ref_col, ref_row = ref_match.groups()
        resolved = _excel_value(sheet, int(ref_row), column_index_from_string(ref_col), cache)
        cache[coordinate] = resolved
        return resolved

    cache[coordinate] = raw
    return raw


def _find_plan_sheet(workbook: openpyxl.Workbook):
    for name in workbook.sheetnames:
        if "plan de entregas" in _normalize_header(name):
            return workbook[name]
    return workbook[workbook.sheetnames[0]]


def _find_header_row(sheet) -> int:
    max_scan = min(sheet.max_row, 15)
    for row in range(1, max_scan + 1):
        values = [_normalize_header(sheet.cell(row=row, column=col).value) for col in range(1, 9)]
        if "num pedido" in values and "contenedor" in values:
            return row
    raise HTTPException(status_code=400, detail="No se encontró la hoja PLAN DE ENTREGAS con el formato esperado.")


def _text_or_none(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_executive_template(path: Path, source_updated_at: str | None = None) -> list[ExecutiveShipment]:
    workbook = openpyxl.load_workbook(str(path), data_only=False)
    sheet = _find_plan_sheet(workbook)
    header_row = _find_header_row(sheet)
    cache: dict[str, Any] = {}
    shipments: list[ExecutiveShipment] = []
    empty_streak = 0

    for row in range(header_row + 1, sheet.max_row + 1):
        order_number = _text_or_none(_excel_value(sheet, row, 2, cache))
        container = _text_or_none(_excel_value(sheet, row, 5, cache))
        supplier_display = _text_or_none(_excel_value(sheet, row, 6, cache))

        if not any((order_number, container, supplier_display)):
            empty_streak += 1
            if empty_streak >= 12 and shipments:
                break
            continue
        empty_streak = 0

        shipments.append(
            ExecutiveShipment(
                id=container or order_number or f"fila-{row}",
                order_number=order_number,
                visa_reference=_text_or_none(_excel_value(sheet, row, 3, cache)),
                terminal=_text_or_none(_excel_value(sheet, row, 4, cache)),
                container=container,
                supplier_display=supplier_display,
                forwarder=_text_or_none(_excel_value(sheet, row, 7, cache)),
                provider_payment_due=_to_iso_date(_excel_value(sheet, row, 8, cache)),
                provider_payment_status=_text_or_none(_excel_value(sheet, row, 9, cache)),
                forwarder_payment_due=_to_iso_date(_excel_value(sheet, row, 10, cache)),
                forwarder_payment_status=_text_or_none(_excel_value(sheet, row, 11, cache)),
                etd=_to_iso_date(_excel_value(sheet, row, 12, cache)),
                eta=_to_iso_date(_excel_value(sheet, row, 13, cache)),
                storage_deadline=_to_iso_date(_excel_value(sheet, row, 14, cache)),
                pedimento_charge_date=_to_iso_date(_excel_value(sheet, row, 15, cache)),
                dispatch_date=_to_iso_date(_excel_value(sheet, row, 16, cache)),
                warehouse_request_date=_to_iso_date(_excel_value(sheet, row, 17, cache)),
                delay_reference=_text_or_none(_excel_value(sheet, row, 18, cache)),
                goods_summary=_text_or_none(_excel_value(sheet, row, 19, cache)),
                transportista=_text_or_none(_excel_value(sheet, row, 20, cache)),
                despacho=_text_or_none(_excel_value(sheet, row, 21, cache)),
                pedimento=_text_or_none(_excel_value(sheet, row, 22, cache)),
                iva=_text_or_none(_excel_value(sheet, row, 23, cache)),
                warehouse=_text_or_none(_excel_value(sheet, row, 24, cache)),
                container_status=_text_or_none(_excel_value(sheet, row, 1, cache)),
                source_updated_at=source_updated_at,
            )
        )

    return shipments


def _merge_with_history(shipments: list[ExecutiveShipment], history: list[ShipmentTracking]) -> None:
    index: dict[str, ShipmentTracking] = {}
    for record in history:
        for identifier in _shipment_identifiers(record):
            index[identifier] = record

    for shipment in shipments:
        matched = None
        for identifier in _shipment_identifiers(shipment):
            matched = index.get(identifier)
            if matched:
                break
        if not matched:
            continue
        shipment.general_po = shipment.general_po or matched.general_po
        shipment.invoice_number = shipment.invoice_number or matched.invoice_number
        shipment.supplier_name = shipment.supplier_name or matched.supplier_name
        shipment.origin_port = shipment.origin_port or matched.origin_port
        shipment.destination_port = shipment.destination_port or matched.destination_port
        shipment.incoterm = shipment.incoterm or matched.incoterm
        shipment.total_usd = shipment.total_usd or matched.total_usd
        shipment.source_updated_at = shipment.source_updated_at or matched.source_updated_at


def _status_is_pending(value: str | None) -> bool:
    normalized = _normalize_header(value)
    return "pendiente" in normalized or "pending" in normalized


def _delivered_flag(shipment: ExecutiveShipment) -> bool:
    normalized = _normalize_header(shipment.container_status)
    return "entregado" in normalized


def _milestone_status(milestone_date: str | None, today: date) -> str:
    resolved = _coerce_date(milestone_date)
    if not resolved:
        return "upcoming"
    return "completed" if resolved <= today else "scheduled"


def _next_future_event(shipment: ExecutiveShipment, today: date) -> tuple[str | None, str | None]:
    candidates = [
        ("Pago a proveedor", shipment.provider_payment_due),
        ("Pago a forwarder", shipment.forwarder_payment_due),
        ("ETA", shipment.eta),
        ("Cobro de pedimento", shipment.pedimento_charge_date),
        ("Despacho", shipment.dispatch_date),
        ("Solicitud en almacén", shipment.warehouse_request_date),
    ]
    dated = []
    for label, value in candidates:
        resolved = _coerce_date(value)
        if resolved and resolved >= today:
            dated.append((resolved, label, value))
    if not dated:
        return None, None
    dated.sort(key=lambda item: item[0])
    _, label, raw = dated[0]
    return label, raw


def _attention_for(shipment: ExecutiveShipment, today: date) -> tuple[str, str | None]:
    if _status_is_pending(shipment.provider_payment_status):
        due = _coerce_date(shipment.provider_payment_due)
        if due and due < today:
            return "risk", "Pago a proveedor vencido"
    if _status_is_pending(shipment.forwarder_payment_status):
        due = _coerce_date(shipment.forwarder_payment_due)
        if due and due < today:
            return "risk", "Pago a forwarder vencido"
    storage_deadline = _coerce_date(shipment.storage_deadline)
    if storage_deadline and storage_deadline < today and not _delivered_flag(shipment):
        return "risk", "Último día de almacenaje vencido"
    eta = _coerce_date(shipment.eta)
    if eta and eta < today and not shipment.dispatch_date and not _delivered_flag(shipment):
        return "watch", "Ya arribó o venció ETA sin despacho registrado"
    if eta:
        delta = (eta - today).days
        if 0 <= delta <= 7 and not _delivered_flag(shipment):
            return "watch", "Arribo próximo en los siguientes 7 días"
    pedimento_charge = _coerce_date(shipment.pedimento_charge_date)
    if pedimento_charge and pedimento_charge < today and not shipment.dispatch_date and not _delivered_flag(shipment):
        return "watch", "Cobro de pedimento vencido sin despacho"
    return "ok", None


def _enrich_executive_status(shipment: ExecutiveShipment) -> None:
    today = date.today()
    delivered = _delivered_flag(shipment)
    if not shipment.destination_port:
        shipment.destination_port = _infer_destination_from_terminal(shipment.terminal)

    stages = [
        ("planned", "Programación", None),
        ("in_transit", "En tránsito", shipment.etd),
        ("arrival", "Arribo a puerto", shipment.eta),
        ("pedimento", "Pedimento", shipment.pedimento_charge_date),
        ("dispatch", "Despacho", shipment.dispatch_date),
        ("delivered", "Entregado", shipment.warehouse_request_date),
    ]

    current_index = 0
    if delivered:
        current_index = len(stages) - 1
    elif shipment.dispatch_date:
        current_index = 4
    elif shipment.pedimento_charge_date:
        current_index = 3
    elif shipment.eta:
        current_index = 2
    elif shipment.etd:
        current_index = 1

    milestones: list[ExecutiveMilestone] = []
    for idx, (key, label, milestone_date) in enumerate(stages):
        status = _milestone_status(milestone_date, today)
        if idx < current_index and status == "upcoming":
            status = "completed"
        elif idx == current_index:
            status = "completed" if delivered and idx == len(stages) - 1 else "current"
        elif idx > current_index and status == "scheduled":
            status = "upcoming"
        milestones.append(ExecutiveMilestone(key=key, label=label, date=milestone_date, status=status))

    shipment.stage_key = stages[current_index][0]
    shipment.stage_label, shipment.stage_color = _stage_meta(shipment.stage_key)
    shipment.progress_pct = int(round((current_index / max(len(stages) - 1, 1)) * 100))
    shipment.days_to_eta = None
    eta = _coerce_date(shipment.eta)
    if eta:
        shipment.days_to_eta = (eta - today).days

    shipment.origin_geo = _geocode_port(shipment.origin_port)
    shipment.destination_geo = _geocode_port(shipment.destination_port)
    shipment.milestones = milestones
    shipment.next_event_label, shipment.next_event_date = _next_future_event(shipment, today)
    shipment.attention_level, shipment.attention_reason = _attention_for(shipment, today)


def _route_summary(shipments: list[ExecutiveShipment]) -> list[ExecutiveRoute]:
    grouped: dict[str, ExecutiveRoute] = {}
    for shipment in shipments:
        if not shipment.origin_geo or not shipment.destination_geo:
            continue
        key = f"{shipment.origin_geo.label}|{shipment.destination_geo.label}"
        if key not in grouped:
            grouped[key] = ExecutiveRoute(
                key=key,
                origin_label=shipment.origin_geo.label,
                destination_label=shipment.destination_geo.label,
                count=0,
                active=0,
                delivered=0,
                color=shipment.stage_color,
                origin=shipment.origin_geo,
                destination=shipment.destination_geo,
            )
        grouped[key].count += 1
        grouped[key].active += 0 if shipment.stage_key == "delivered" else 1
        grouped[key].delivered += 1 if shipment.stage_key == "delivered" else 0
    return sorted(grouped.values(), key=lambda item: (-item.count, item.origin_label, item.destination_label))


def _build_stage_breakdown(shipments: list[ExecutiveShipment]) -> list[dict[str, Any]]:
    counts = Counter(item.stage_key for item in shipments)
    return [
        {"key": key, "label": label, "count": counts.get(key, 0), "color": color}
        for key, label, color in EXECUTIVE_STAGE_META
    ]


def _build_alerts(shipments: list[ExecutiveShipment]) -> list[ExecutiveAlert]:
    alerts: list[ExecutiveAlert] = []
    for shipment in shipments:
        if shipment.attention_level == "ok" or not shipment.attention_reason:
            continue
        alerts.append(
            ExecutiveAlert(
                shipment_id=shipment.id,
                level=shipment.attention_level,
                title=shipment.attention_reason,
                detail=f"{shipment.order_number or shipment.id} · {shipment.supplier_display or 'Proveedor sin identificar'}",
            )
        )
    alerts.sort(key=lambda item: (0 if item.level == "risk" else 1, item.shipment_id))
    return alerts[:8]


def _build_supplier_breakdown(shipments: list[ExecutiveShipment]) -> list[dict[str, Any]]:
    counts = Counter((item.supplier_display or item.supplier_name or "Sin proveedor").strip() for item in shipments)
    return [{"label": label, "count": count} for label, count in counts.most_common(6)]


def _build_terminal_breakdown(shipments: list[ExecutiveShipment]) -> list[dict[str, Any]]:
    counts = Counter((item.terminal or "Sin terminal").strip() for item in shipments)
    return [{"label": label, "count": count} for label, count in counts.most_common(4)]


def _build_overview(shipments: list[ExecutiveShipment], routes: list[ExecutiveRoute]) -> dict[str, Any]:
    total_usd = round(sum(item.total_usd or 0 for item in shipments), 2)
    total_usd_count = len([item for item in shipments if item.total_usd is not None])
    upcoming_arrivals = len(
        [
            item
            for item in shipments
            if item.days_to_eta is not None and 0 <= item.days_to_eta <= 14 and item.stage_key != "delivered"
        ]
    )
    overdue = len([item for item in shipments if item.attention_level == "risk"])
    return {
        "shipments": len(shipments),
        "delivered": len([item for item in shipments if item.stage_key == "delivered"]),
        "active": len([item for item in shipments if item.stage_key != "delivered"]),
        "with_route_data": len([item for item in shipments if item.origin_geo and item.destination_geo]),
        "pending_provider_payment": len([item for item in shipments if _status_is_pending(item.provider_payment_status)]),
        "pending_forwarder_payment": len([item for item in shipments if _status_is_pending(item.forwarder_payment_status)]),
        "upcoming_arrivals": upcoming_arrivals,
        "at_risk": overdue,
        "total_usd": total_usd if total_usd_count else None,
        "total_usd_count": total_usd_count,
        "coverage_pct": int(round((len(routes) / max(len(shipments), 1)) * 100)),
    }


def _sort_shipments(shipments: list[ExecutiveShipment]) -> list[ExecutiveShipment]:
    return sorted(
        shipments,
        key=lambda item: (
            0 if item.attention_level == "risk" else 1 if item.attention_level == "watch" else 2,
            EXECUTIVE_STAGE_INDEX.get(item.stage_key, 0),
            _latest_tracking_date(
                ShipmentTracking(
                    id=item.id,
                    source="executive",
                    order_number=item.order_number,
                    general_po=item.general_po,
                    invoice_number=item.invoice_number,
                    container=item.container,
                    etd=item.etd,
                    eta=item.eta,
                    customs_release=item.dispatch_date,
                    warehouse_arrival=item.warehouse_request_date,
                    source_updated_at=item.source_updated_at,
                )
            ),
        ),
    )


def _load_latest_importaciones_workbook(db: Session, user: User) -> tuple[Job, Path]:
    files_root = Path(settings.files_root).resolve()
    query = (
        db.query(Job)
        .filter(
            Job.app_key == IMPORTACIONES_SOURCE_APP_KEY,
            Job.status == "succeeded",
            Job.output_path.isnot(None),
        )
        .order_by(Job.created_at.desc())
    )
    if not user.is_admin:
        query = query.filter(Job.created_by == user.username)

    for job in query.all():
        if not job.output_path:
            continue
        abs_path = (files_root / job.output_path).resolve()
        if abs_path.exists() and abs_path.suffix.lower() in {".xlsx", ".xlsm"}:
            return job, abs_path

    raise HTTPException(
        status_code=404,
        detail="No se encontró un Excel reciente de Importaciones para construir la vista ejecutiva.",
    )


def _build_executive_payload(
    db: Session,
    user: User,
    workbook_path: Path,
    source_label: str,
    source_updated_at: str | None,
    use_importaciones_history: bool,
) -> dict[str, Any]:
    shipments = _parse_executive_template(workbook_path, source_updated_at=source_updated_at)
    if not shipments:
        raise HTTPException(status_code=400, detail="El Excel no contiene embarques utilizables para la vista ejecutiva.")

    history_shipments = _load_importaciones_history(db, user) if use_importaciones_history else []
    if history_shipments:
        _merge_with_history(shipments, history_shipments)

    for shipment in shipments:
        _enrich_executive_status(shipment)

    ordered = _sort_shipments(shipments)
    routes = _route_summary(ordered)
    stage_breakdown = _build_stage_breakdown(ordered)
    alerts = _build_alerts(ordered)

    return {
        "generated_at": datetime.now(ZoneInfo("America/Mexico_City")).isoformat(),
        "data_source": {
            "label": source_label,
            "updated_at": source_updated_at,
            "used_history": use_importaciones_history,
        },
        "overview": _build_overview(ordered, routes),
        "stage_breakdown": stage_breakdown,
        "supplier_breakdown": _build_supplier_breakdown(ordered),
        "terminal_breakdown": _build_terminal_breakdown(ordered),
        "routes": [asdict(item) for item in routes],
        "alerts": [asdict(item) for item in alerts],
        "shipments": [asdict(item) for item in ordered],
    }


def _styles():
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="ExecHeroTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=22,
            leading=26,
            textColor=colors.white,
            spaceAfter=6,
        )
    )
    styles.add(
        ParagraphStyle(
            name="ExecHeroBody",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#dbeafe"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="ExecSection",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=16,
            textColor=colors.HexColor("#0f172a"),
            spaceAfter=8,
            spaceBefore=10,
        )
    )
    styles.add(
        ParagraphStyle(
            name="ExecBody",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#334155"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="ExecMuted",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=8,
            leading=11,
            textColor=colors.HexColor("#64748b"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="ExecMetric",
            parent=styles["Heading3"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=18,
            textColor=colors.HexColor("#0f172a"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="ExecMetricLabel",
            parent=styles["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#64748b"),
        )
    )
    return styles


def _metric_card(label: str, value: str, styles) -> Table:
    table = Table(
        [[Paragraph(label, styles["ExecMetricLabel"]), Paragraph(value, styles["ExecMetric"])]],
        colWidths=[1.55 * inch],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#dbeafe")),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    return table


def _stage_bar_drawing(stage_breakdown: list[dict[str, Any]], width: float = 520, height: float = 54) -> Drawing:
    drawing = Drawing(width, height)
    total = sum(item["count"] for item in stage_breakdown) or 1
    x = 0
    for item in stage_breakdown:
        segment_width = width * (item["count"] / total)
        drawing.add(Rect(x, 24, segment_width, 12, fillColor=colors.HexColor(item["color"]), strokeWidth=0))
        if segment_width > 52:
            drawing.add(
                String(
                    x + 4,
                    6,
                    f'{item["label"]}: {item["count"]}',
                    fontName="Helvetica",
                    fontSize=7,
                    fillColor=colors.HexColor("#334155"),
                )
            )
        x += segment_width
    drawing.add(Rect(0, 24, width, 12, strokeColor=colors.HexColor("#cbd5e1"), fillColor=None, strokeWidth=1))
    return drawing


def _project_to_map(location: ExecutiveLocation, width: float, height: float) -> tuple[float, float]:
    x = (location.lng + 180.0) / 360.0 * width
    y = height - ((location.lat + 90.0) / 180.0 * height)
    return x, y


def _route_map_drawing(routes: list[dict[str, Any]], width: float = 520, height: float = 220) -> Drawing:
    drawing = Drawing(width, height)
    drawing.add(Rect(0, 0, width, height, rx=18, ry=18, fillColor=colors.HexColor("#eff6ff"), strokeWidth=0))
    for step in range(1, 5):
        y = height * step / 5
        drawing.add(Line(0, y, width, y, strokeColor=colors.HexColor("#dbeafe"), strokeWidth=0.7))
    for step in range(1, 6):
        x = width * step / 6
        drawing.add(Line(x, 0, x, height, strokeColor=colors.HexColor("#dbeafe"), strokeWidth=0.7))

    labeled: set[str] = set()
    for item in routes[:8]:
        origin = item.get("origin")
        destination = item.get("destination")
        if not origin or not destination:
            continue
        start = ExecutiveLocation(**origin)
        end = ExecutiveLocation(**destination)
        sx, sy = _project_to_map(start, width, height)
        ex, ey = _project_to_map(end, width, height)
        mid_x = (sx + ex) / 2
        control_y = min(height - 18, max(20, (sy + ey) / 2 + 32))
        path = Path(strokeColor=colors.HexColor(item["color"]), strokeWidth=max(1.4, min(item["count"], 4)), fillColor=None)
        path.moveTo(sx, sy)
        path.curveTo(mid_x, control_y, mid_x, control_y, ex, ey)
        drawing.add(path)
        drawing.add(Circle(sx, sy, 3.2, fillColor=colors.HexColor("#0f172a"), strokeColor=None))
        drawing.add(Circle(ex, ey, 3.2, fillColor=colors.HexColor("#0f172a"), strokeColor=None))
        if start.label not in labeled:
            labeled.add(start.label)
            drawing.add(String(sx + 5, sy + 5, start.label, fontName="Helvetica", fontSize=7, fillColor=colors.HexColor("#0f172a")))
        if end.label not in labeled:
            labeled.add(end.label)
            drawing.add(String(ex + 5, ey - 10, end.label, fontName="Helvetica", fontSize=7, fillColor=colors.HexColor("#0f172a")))
    return drawing


def _shipment_card(shipment: dict[str, Any], styles) -> Table:
    status_label, _ = _status_style(shipment["attention_level"])
    title = shipment.get("order_number") or shipment.get("container") or shipment["id"]
    subtitle = shipment.get("supplier_display") or shipment.get("supplier_name") or "Proveedor sin identificar"
    next_event = "Sin siguiente fecha"
    if shipment.get("next_event_label") and shipment.get("next_event_date"):
        next_event = f'{shipment["next_event_label"]}: {shipment["next_event_date"]}'
    route = "Ruta no disponible"
    if shipment.get("origin_port") and shipment.get("destination_port"):
        route = f'{shipment["origin_port"]} -> {shipment["destination_port"]}'

    left_lines = [
        Paragraph(f"<b>{title}</b>", styles["ExecBody"]),
        Paragraph(subtitle, styles["ExecMuted"]),
        Paragraph(f'Etapa: <b>{shipment["stage_label"]}</b>', styles["ExecBody"]),
        Paragraph(f'Contenedor: {shipment.get("container") or "Sin dato"}', styles["ExecMuted"]),
        Paragraph(f'Terminal: {shipment.get("terminal") or "Sin dato"}', styles["ExecMuted"]),
        Paragraph(f'Ruta: {route}', styles["ExecMuted"]),
    ]
    right_lines = [
        Paragraph(f"<b>{status_label}</b>", styles["ExecBody"]),
        Paragraph(shipment.get("attention_reason") or "Sin alertas críticas registradas", styles["ExecMuted"]),
        Paragraph(f'Progreso: {shipment["progress_pct"]}%', styles["ExecBody"]),
        Paragraph(f"Siguiente: {next_event}", styles["ExecMuted"]),
        Paragraph(f'ETA: {shipment.get("eta") or "Sin ETA"}', styles["ExecMuted"]),
        Paragraph(f'Despacho: {shipment.get("dispatch_date") or "Sin fecha"}', styles["ExecMuted"]),
    ]
    table = Table([[left_lines, right_lines]], colWidths=[4.55 * inch, 4.2 * inch])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#dbeafe")),
                ("LINEBEFORE", (1, 0), (1, 0), 1, colors.HexColor("#e2e8f0")),
                ("TOPPADDING", (0, 0), (-1, -1), 12),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#f8fafc")),
            ]
        )
    )
    return table


def _build_pdf(payload: dict[str, Any]) -> bytes:
    styles = _styles()
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=0.45 * inch,
        rightMargin=0.45 * inch,
        topMargin=0.45 * inch,
        bottomMargin=0.45 * inch,
    )
    story: list[Any] = []
    overview = payload["overview"]
    data_source = payload["data_source"]
    routes = payload["routes"]
    alerts = payload["alerts"]
    shipments = payload["shipments"]

    hero = Table(
        [
            [
                [
                    Paragraph("Seguimiento ejecutivo de importaciones", styles["ExecHeroTitle"]),
                    Paragraph(
                        "Vista de dirección con hitos, riesgos, pagos, cobertura de rutas y estado consolidado del portafolio.",
                        styles["ExecHeroBody"],
                    ),
                    Paragraph(
                        f'Fuente: {data_source["label"]} | Generado: {payload["generated_at"][:16].replace("T", " ")}',
                        styles["ExecHeroBody"],
                    ),
                ]
            ]
        ],
        colWidths=[10 * inch],
    )
    hero.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0f172a")),
                ("BOX", (0, 0), (-1, -1), 0, colors.white),
                ("LEFTPADDING", (0, 0), (-1, -1), 18),
                ("RIGHTPADDING", (0, 0), (-1, -1), 18),
                ("TOPPADDING", (0, 0), (-1, -1), 18),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 18),
            ]
        )
    )
    story.append(hero)
    story.append(Spacer(1, 0.18 * inch))

    metrics = [
        _metric_card("Embarques", str(overview["shipments"]), styles),
        _metric_card("Entregados", str(overview["delivered"]), styles),
        _metric_card("Activos", str(overview["active"]), styles),
        _metric_card("Riesgo", str(overview["at_risk"]), styles),
        _metric_card("Arribos 14 días", str(overview["upcoming_arrivals"]), styles),
        _metric_card("Rutas visibles", str(overview["with_route_data"]), styles),
    ]
    metrics_table = Table([metrics[:3], metrics[3:]], colWidths=[3.2 * inch, 3.2 * inch, 3.2 * inch])
    metrics_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story.append(metrics_table)

    story.append(Paragraph("Distribución por etapa", styles["ExecSection"]))
    story.append(_stage_bar_drawing(payload["stage_breakdown"]))
    story.append(Spacer(1, 0.14 * inch))

    if routes:
        story.append(Paragraph("Mapa ejecutivo de rutas", styles["ExecSection"]))
        story.append(_route_map_drawing(routes))

    if alerts:
        story.append(Paragraph("Alertas prioritarias", styles["ExecSection"]))
        alert_rows = [["Nivel", "Embarque", "Detalle"]]
        for item in alerts[:6]:
            label, _ = _status_style(item["level"])
            alert_rows.append([label, item["shipment_id"], item["title"]])
        alert_table = Table(alert_rows, colWidths=[1.2 * inch, 1.8 * inch, 6.8 * inch], repeatRows=1)
        alert_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e0f2fe")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                    ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(alert_table)

    story.append(PageBreak())
    story.append(Paragraph("Portafolio de embarques", styles["ExecSection"]))
    for shipment in shipments:
        story.append(_shipment_card(shipment, styles))
        story.append(Spacer(1, 0.12 * inch))

    doc.build(story)
    return buffer.getvalue()


def _resolve_workbook_input(
    tracking_file: UploadFile | None,
    use_latest_template: bool,
    temp_root: Path,
    db: Session,
    user: User,
) -> tuple[Path, str, str | None]:
    if tracking_file is not None and tracking_file.filename:
        filename = Path(tracking_file.filename).name
        if Path(filename).suffix.lower() not in {".xlsx", ".xlsm"}:
            raise HTTPException(status_code=400, detail="La vista ejecutiva solo acepta archivos Excel .xlsx o .xlsm.")
        uploaded_path = _save_upload(temp_root / filename, tracking_file)
        return uploaded_path, f"Archivo cargado: {filename}", None

    if not use_latest_template:
        raise HTTPException(status_code=400, detail="Debes subir el Excel de seguimiento o activar el uso del último archivo generado.")

    job, workbook_path = _load_latest_importaciones_workbook(db, user)
    source_label = f"Último Excel generado por Importaciones ({workbook_path.name})"
    updated_at = job.created_at.astimezone(ZoneInfo("America/Mexico_City")).isoformat() if job.created_at else None
    return workbook_path, source_label, updated_at


@router.post("/analyze")
async def analyze_executive_tracking(
    tracking_file: UploadFile | None = File(default=None),
    use_latest_template: bool = Form(default=True),
    use_importaciones_history: bool = Form(default=True),
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    app = db.get(AppDefinition, TRACKING_APP_KEY)
    if not app or not app.enabled:
        raise HTTPException(status_code=404, detail="App de seguimiento no disponible.")

    ensure_app_access(user, TRACKING_APP_KEY, db)

    with tempfile.TemporaryDirectory(prefix="compras-executive-") as temp_dir:
        workbook_path, source_label, source_updated_at = _resolve_workbook_input(
            tracking_file=tracking_file,
            use_latest_template=use_latest_template,
            temp_root=Path(temp_dir),
            db=db,
            user=user,
        )
        return _build_executive_payload(
            db=db,
            user=user,
            workbook_path=workbook_path,
            source_label=source_label,
            source_updated_at=source_updated_at,
            use_importaciones_history=use_importaciones_history,
        )


@router.post("/export-pdf")
async def export_executive_tracking_pdf(
    tracking_file: UploadFile | None = File(default=None),
    use_latest_template: bool = Form(default=True),
    use_importaciones_history: bool = Form(default=True),
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    app = db.get(AppDefinition, TRACKING_APP_KEY)
    if not app or not app.enabled:
        raise HTTPException(status_code=404, detail="App de seguimiento no disponible.")

    ensure_app_access(user, TRACKING_APP_KEY, db)

    with tempfile.TemporaryDirectory(prefix="compras-executive-pdf-") as temp_dir:
        workbook_path, source_label, source_updated_at = _resolve_workbook_input(
            tracking_file=tracking_file,
            use_latest_template=use_latest_template,
            temp_root=Path(temp_dir),
            db=db,
            user=user,
        )
        payload = _build_executive_payload(
            db=db,
            user=user,
            workbook_path=workbook_path,
            source_label=source_label,
            source_updated_at=source_updated_at,
            use_importaciones_history=use_importaciones_history,
        )

    pdf_bytes = _build_pdf(payload)
    filename = f"seguimiento-importaciones-direccion-{date.today().isoformat()}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )
