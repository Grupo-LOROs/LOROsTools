import os
import shutil
import time
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import DateTime, Integer, String, Text, create_engine, select
from sqlalchemy.dialects.postgresql import JSONB, UUID
from sqlalchemy.orm import DeclarativeBase, Mapped, Session, mapped_column


DATABASE_URL = os.getenv("DATABASE_URL", "")
FILES_ROOT = os.getenv("FILES_ROOT", "/data/files")
POLL_SECONDS = float(os.getenv("WORKER_POLL_SECONDS", "2"))


class Base(DeclarativeBase):
    pass


class Job(Base):
    __tablename__ = "jobs"

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True)
    app_key: Mapped[str] = mapped_column(String(64), nullable=False)
    created_by: Mapped[str] = mapped_column(String(64), nullable=False)

    status: Mapped[str] = mapped_column(String(20), nullable=False)
    progress: Mapped[int] = mapped_column(Integer, nullable=False)
    message: Mapped[str | None] = mapped_column(Text)
    params: Mapped[dict] = mapped_column(JSONB, nullable=False)

    template_path: Mapped[str | None] = mapped_column(String(512))
    output_path: Mapped[str | None] = mapped_column(String(512))

    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True))
    updated_at: Mapped[datetime] = mapped_column(DateTime(timezone=True))


engine = create_engine(DATABASE_URL, pool_pre_ping=True)


def _make_placeholder_xlsx(dest: Path, job: Job) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "result"
    ws["A1"] = "status"
    ws["B1"] = job.status
    ws["A2"] = "app_key"
    ws["B2"] = job.app_key
    ws["A3"] = "job_id"
    ws["B3"] = str(job.id)
    ws["A4"] = "note"
    ws["B4"] = "Placeholder output (worker stub). Replace with real parser/processor." 
    wb.save(dest)


def _process_job(job: Job) -> str:
    """Returns relative output path."""
    files_root = Path(FILES_ROOT)
    job_root_rel = f"jobs/{job.id}"
    out_rel = f"{job_root_rel}/output/output.xlsx"
    out_abs = (files_root / out_rel).resolve()

    # If there is a template, copy it as a starting point
    if job.template_path:
        tpl_abs = (files_root / job.template_path).resolve()
        out_abs.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(tpl_abs, out_abs)
        return out_rel

    _make_placeholder_xlsx(out_abs, job)
    return out_rel


def main() -> None:
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL is not set")

    print(f"[worker] starting | poll={POLL_SECONDS}s | files_root={FILES_ROOT}")

    while True:
        try:
            with Session(engine) as db:
                job = db.execute(
                    select(Job).where(Job.status == "queued").order_by(Job.created_at.asc()).limit(1)
                ).scalar_one_or_none()

                if not job:
                    time.sleep(POLL_SECONDS)
                    continue

                job.status = "running"
                job.progress = 5
                job.message = "Processing"
                db.commit()

                try:
                    out_rel = _process_job(job)
                    job.output_path = out_rel
                    job.status = "succeeded"
                    job.progress = 100
                    job.message = "Done"
                except Exception as e:
                    job.status = "failed"
                    job.progress = 100
                    job.message = f"Failed: {e}"

                db.commit()

        except Exception as e:
            # don't die on transient errors
            print(f"[worker] loop error: {e}")
            time.sleep(POLL_SECONDS)


if __name__ == "__main__":
    main()
