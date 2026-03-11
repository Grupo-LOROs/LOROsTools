import tempfile
import unittest
import uuid
from datetime import date, datetime
from pathlib import Path

import pdfplumber
from openpyxl import Workbook

from processors.base import JobContext
from processors.cxp_autorizacion_pagos import (
    APP_KEY,
    _load_snapshot,
    process,
)


class CxpAutorizacionPagosTests(unittest.TestCase):
    def _build_workbook(self, path: Path) -> None:
        wb = Workbook()
        bd = wb.active
        bd.title = "BD.CXP"
        resumen = wb.create_sheet("RESUMEN POR U.N")
        conceptos = wb.create_sheet("DESLGOSE POR TIPO DE GAST ")

        bd["B4"] = 9
        bd["B5"] = datetime(2026, 3, 2)
        bd["B6"] = datetime(2026, 3, 6)
        headers = [
            "RAZON ",
            "UNIDAD DE NEGOCIO",
            "FAMILIA CORPORATIVO",
            "PROYECTO",
            "PROVEEDOR",
            "CONCEPTO ",
            "CONCEPTO DETALLADO",
            "FORMA DE PAGO",
            "IMPORTE PROVISIONADO MXN",
            "PAGADO ",
            "LUNES",
            "MARTES",
            "MIÉRCOLES",
            "JUEVES",
            "VIERNES",
            "FECHA REAL DE PAGO",
        ]
        for col, value in enumerate(headers, start=1):
            bd.cell(11, col).value = value

        rows = [
            ["DEESA", "COMERCIALIZACION", "COMER, EDIF, URBA", "RC COMER", "Proveedor A", "COMISIONES", "COMISIONES DE VENTA", "EFECTIVO", 100.0, 100.0, None, None, None, None, None, datetime(2026, 3, 6)],
            ["DEESA", "COMERCIALIZACION", "COMER, EDIF, URBA", "RC COMER", "Proveedor B", "COMISIONES", "NO DEBE ENTRAR", "EFECTIVO", 25.0, 25.0, None, None, None, None, None, datetime(2026, 3, 4)],
            ["DEESA", "OBRA PUBLICA", "OBRA PUBLICA", "RC URBA", "Proveedor C", "COMBUSTIBLE", "GASOLINA", "TRANSFERENCIA", 200.0, 200.0, None, None, None, None, None, datetime(2026, 2, 6)],
            ["CORPORATIVO", "SOCIOS", "OFICINA CENTRAL, SOCIOS", "25. SOCIOS", "Proveedor D", "PRESTAMOS", "NO DEBE ENTRAR", "COMPROBACION", 50.0, 50.0, None, None, None, None, None, datetime(2026, 2, 5)],
        ]
        for row_idx, row_values in enumerate(rows, start=12):
            for col_idx, value in enumerate(row_values, start=1):
                bd.cell(row_idx, col_idx).value = value

        resumen["A9"] = "UNIDAD DE NEGOCIO"
        resumen["B9"] = "Suma de VIERNES"
        resumen["A10"] = "COMERCIALIZACION"
        resumen["B10"] = 100.0
        resumen["A11"] = "OBRA PUBLICA"
        resumen["B11"] = 200.0
        resumen["A12"] = "TOTAL"
        resumen["B12"] = 300.0

        conceptos["A4"] = "CONCEPTO"
        conceptos["D4"] = "LUNES"
        conceptos["E4"] = "MARTES"
        conceptos["F4"] = "MIÉRCOLES"
        conceptos["G4"] = "JUEVES"
        conceptos["H4"] = "VIERNES"
        conceptos["A5"] = "COMISIONES"
        conceptos["H5"] = 100.0
        conceptos["A6"] = "COMBUSTIBLE"
        conceptos["H6"] = 200.0

        wb.save(path)

    def test_load_snapshot_filters_entries_by_target_weekday(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "provision.xlsx"
            self._build_workbook(workbook_path)

            snapshot = _load_snapshot(workbook_path)

            self.assertEqual(snapshot.week_number, 9)
            self.assertEqual(snapshot.start_date, date(2026, 3, 2))
            self.assertEqual(snapshot.end_date, date(2026, 3, 6))
            self.assertEqual(snapshot.target_label, "VIERNES")
            self.assertEqual(len(snapshot.entries), 2)
            self.assertEqual(snapshot.total, 300.0)
            self.assertEqual([row[0] for row in snapshot.unit_summary], ["COMERCIALIZACION", "OBRA PUBLICA", "TOTAL"])
            self.assertEqual(snapshot.concept_summary, [("COMISIONES", 100.0), ("COMBUSTIBLE", 200.0)])

    def test_process_generates_pdf_with_expected_text(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            job_id = uuid.uuid4()
            inputs_dir = root / "jobs" / str(job_id) / "inputs"
            output_dir = root / "jobs" / str(job_id) / "output"
            inputs_dir.mkdir(parents=True, exist_ok=True)
            output_dir.mkdir(parents=True, exist_ok=True)

            workbook_path = inputs_dir / "provision.xlsx"
            self._build_workbook(workbook_path)

            ctx = JobContext(
                job_id=job_id,
                app_key=APP_KEY,
                params={},
                files_root=root,
                inputs_dir=inputs_dir,
                output_dir=output_dir,
                template_abs=None,
                report_progress=lambda percent, message: None,
            )

            rel_output = process(ctx)
            abs_output = root / rel_output

            self.assertTrue(abs_output.exists())
            self.assertEqual(abs_output.name, "AUTORIZACION_DE_PAGOS_060326.pdf")
            self.assertEqual(ctx.params["target_weekday"], "VIERNES")
            self.assertEqual(ctx.params["detail_rows"], 2)
            self.assertEqual(ctx.params["total_amount"], 300.0)

            with pdfplumber.open(abs_output) as pdf:
                text = "\n".join(page.extract_text() or "" for page in pdf.pages)

            self.assertIn("AUTORIZACIÓN DE PAGOS", text)
            self.assertIn("COMERCIALIZACION", text)
            self.assertIn("OBRA PUBLICA", text)
            self.assertIn("$300.00", text)
            self.assertNotIn("NO DEBE ENTRAR", text)


if __name__ == "__main__":
    unittest.main()
