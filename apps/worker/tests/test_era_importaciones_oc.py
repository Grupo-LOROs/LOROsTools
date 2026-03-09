import unittest
from datetime import datetime

from openpyxl import Workbook

from processors.era_importaciones_oc import (
    OrderItem,
    OrderRecord,
    _append_record_to_workbook,
    _display_source_name,
    _extract_invoice_from_lines,
    _parse_packing_items,
)


class EraImportacionesProcessorTests(unittest.TestCase):
    def test_display_source_name_removes_upload_prefix(self):
        self.assertEqual(_display_source_name("00-factura.pdf"), "factura.pdf")
        self.assertEqual(_display_source_name("factura.pdf"), "factura.pdf")

    def test_extract_invoice_from_lines_reads_number_and_date(self):
        invoice_number, invoice_date = _extract_invoice_from_lines(
            [
                "PACKING LIST",
                "TO INVOICE NO. DATE",
                "26L05 31-DEC2025",
            ]
        )

        self.assertEqual(invoice_number, "26L05")
        self.assertEqual(invoice_date, datetime(2025, 12, 31))

    def test_parse_packing_items_extracts_model_qty_and_weights(self):
        items = _parse_packing_items(
            [
                "SOLAR VACUUM TUBE TUG-08 12 374.40KGS 346.40KGS 3.92M3",
                "SOLAR FRAME BG-10C 170 2720.00KGS 2550.00KGS 14.25M3",
            ]
        )

        self.assertEqual(len(items), 2)
        self.assertEqual(items[0].model, "TUG-08")
        self.assertEqual(items[0].quantity, 12)
        self.assertEqual(items[0].gross_weight_kg, 374.40)
        self.assertEqual(items[1].model, "BG-10C")
        self.assertEqual(items[1].volume_m3, 14.25)

    def test_goods_summary_groups_same_model_and_provider_label(self):
        record = OrderRecord(
            source_file="packing-list.pdf",
            supplier_name="HELIOS POWER LTD",
            items=[
                OrderItem(model="TUG-08", description="Solar vacuum tube", quantity=12),
                OrderItem(model="TUG-08", description="Solar vacuum tube", quantity=220),
                OrderItem(model="BG-10C", description="Solar frame", quantity=170),
            ],
        )

        self.assertEqual(record.provider_label(), "HELIOS")
        self.assertEqual(record.goods_summary(), "TUG-08 (232) BG-10C (170)")

    def test_append_record_to_workbook_writes_expected_values(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "PLAN DE ENTREGAS "

        for col in range(1, 26):
            ws.cell(3, col).value = f"base-{col}"

        record = OrderRecord(
            source_file="26L05_PL.pdf",
            order_number="26L05",
            supplier_name="HELIOS POWER LTD",
            container="TEMU8352600",
            terminal="APM",
            etd=datetime(2025, 12, 20),
            eta=datetime(2026, 1, 10),
            items=[
                OrderItem(model="TUG-08", description="Solar vacuum tube", quantity=232),
                OrderItem(model="BG-10C", description="Solar frame", quantity=170),
            ],
        )

        _append_record_to_workbook(
            ws,
            record,
            row=4,
            params={
                "referencia_visa": "RV-001",
                "forwarder": "KUEHNE",
                "transportista": "TRAXION",
                "despacho": "ALMACÉN",
            },
        )

        self.assertEqual(ws.cell(4, 1).value, "DOCUMENTACIÓN ENVIADA")
        self.assertEqual(ws.cell(4, 2).value, "26L05")
        self.assertEqual(ws.cell(4, 3).value, "RV-001")
        self.assertEqual(ws.cell(4, 4).value, "APM")
        self.assertEqual(ws.cell(4, 5).value, "TEMU8352600")
        self.assertEqual(ws.cell(4, 6).value, "HELIOS")
        self.assertEqual(ws.cell(4, 7).value, "KUEHNE")
        self.assertEqual(ws.cell(4, 8).value, "=+M4-5")
        self.assertEqual(ws.cell(4, 9).value, "PENDIENTE")
        self.assertEqual(ws.cell(4, 10).value, "=+M4+30")
        self.assertEqual(ws.cell(4, 11).value, "PENDIENTE")
        self.assertEqual(ws.cell(4, 12).value, datetime(2025, 12, 20))
        self.assertEqual(ws.cell(4, 13).value, datetime(2026, 1, 10))
        self.assertEqual(ws.cell(4, 14).value, "=+M4+6")
        self.assertEqual(ws.cell(4, 15).value, "=+M4+1")
        self.assertEqual(ws.cell(4, 16).value, "=+O4+1")
        self.assertEqual(ws.cell(4, 17).value, "=+P4+1")
        self.assertEqual(ws.cell(4, 18).value, "=+M4+20")
        self.assertEqual(ws.cell(4, 19).value, "TUG-08 (232) BG-10C (170)")
        self.assertEqual(ws.cell(4, 20).value, "TRAXION")
        self.assertEqual(ws.cell(4, 21).value, "ALMACÉN")


if __name__ == "__main__":
    unittest.main()
