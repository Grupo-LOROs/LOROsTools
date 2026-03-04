"""
ERA Ventas – Comisionador 2026
───────────────────────────────
Calculates sales commissions based on tiered pricing schema.

Input:  1 XLSX base_comisiones with sheets:
        - ResultadosVentascomisionesporc (sales data)
        - Hoja2 (OV validation: ov==cruce)
        + 1 XLSM schema file with:
        - COMISIONES 2026 (commission tiers by sales volume)
        - NUEVAS LISTAS (product price lists P1-P4)

Output: 1 XLSX detail + 1 PDF carátula (summary)

Adapted from: ComisionadorERA/comisionador.py
"""

import os
import re
import warnings
from dataclasses import dataclass
from datetime import date, datetime
from typing import Optional, Dict, Tuple

import numpy as np
import openpyxl
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from .base import JobContext

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
    category=UserWarning,
)

# ─── Config ──────────────────────────────────────────────────

BASE_SHEET_VENTAS = "ResultadosVentascomisionesporc"
BASE_SHEET_FILTRO = "Hoja2"
IVA_FACTOR = 1.16
MIN_COMMISSION_TOTAL = 1.0

NON_COMMISSION_PRODUCT_REGEX = re.compile(r"(?i)\b(iva|i\.?v\.?a\.?|impuesto|tax)\b")

DISPLAY_COLS = [
    "Fecha", "Asesor", "Cliente", "OV", "Producto",
    "Cantidad", "Precio Bruto", "Precio Unitario Neto", "Venta Total",
    "Precio 4", "Precio 3", "Precio 2", "Precio 1",
    "Comisión", "Total comisión",
]


# ─── Helpers ─────────────────────────────────────────────────

def _safe_float(x, default=float("nan")):
    try:
        if pd.isna(x):
            return default
        return float(x)
    except Exception:
        return default


def _money(x):
    try:
        if pd.isna(x):
            return ""
        return f"{float(x):,.2f}"
    except Exception:
        return str(x)


def _norm_ov(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v).strip()
    return str(v).strip()


def _is_tax_line(product_value) -> bool:
    s = "" if product_value is None else str(product_value).strip()
    if not s:
        return False
    return NON_COMMISSION_PRODUCT_REGEX.search(s) is not None


# ─── OV Validation (Hoja2) ───────────────────────────────────

def _extract_valid_ovs_from_hoja2(xlsx_path: str) -> Tuple[set, int, int]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if BASE_SHEET_FILTRO not in wb.sheetnames:
        raise ValueError(f"No se encontró la hoja '{BASE_SHEET_FILTRO}' en la base de comisiones.")
    ws = wb[BASE_SHEET_FILTRO]

    header_row = None
    ov_col = cruce_col = obs_col = None

    for r in range(1, 80):
        vals = []
        for c in range(1, 80):
            v = ws.cell(r, c).value
            vals.append(v.strip().lower() if isinstance(v, str) else v)
        if "ov" in vals and "cruce" in vals:
            header_row = r
            ov_col = vals.index("ov") + 1
            cruce_col = vals.index("cruce") + 1
            if "obs" in vals:
                obs_col = vals.index("obs") + 1
            break

    if header_row is None:
        raise ValueError("No se encontraron encabezados 'ov' y 'cruce' en Hoja2.")

    valid = set()
    ok = 0
    not_equal = 0

    for r in range(header_row + 1, ws.max_row + 1):
        ov = _norm_ov(ws.cell(r, ov_col).value)
        cr = _norm_ov(ws.cell(r, cruce_col).value)
        obs = ""
        if obs_col is not None:
            v = ws.cell(r, obs_col).value
            obs = str(v).strip() if v is not None else ""

        if not ov or not cr:
            continue
        if "no factur" in obs.lower():
            continue

        if ov == cr:
            valid.add(ov)
            ok += 1
        else:
            not_equal += 1

    return valid, ok, not_equal


# ─── Commission Rules from Schema ────────────────────────────

class Rules2026:
    def __init__(self, schema_file: str):
        self.schema_file = schema_file
        self.comm_table: Optional[pd.DataFrame] = None
        self.price_map: Dict[str, Dict[str, float]] = {}

    def load(self):
        if not os.path.exists(self.schema_file):
            raise FileNotFoundError(f"No se encontró el esquema: {self.schema_file}")

        wb = openpyxl.load_workbook(self.schema_file, data_only=True, keep_vba=True)
        self._load_comm_table(wb)
        self._load_price_map(wb)

    def _load_comm_table(self, wb):
        if "COMISIONES 2026" not in wb.sheetnames:
            raise ValueError("No se encontró la hoja 'COMISIONES 2026' en el esquema.")
        ws = wb["COMISIONES 2026"]

        header_row = None
        for r in range(1, 80):
            b = ws.cell(r, 2).value
            c = ws.cell(r, 3).value
            if isinstance(b, str) and isinstance(c, str):
                if b.strip().lower() == "limite inf" and c.strip().lower() == "limite sup":
                    header_row = r
                    break
        if header_row is None:
            raise ValueError("No se encontraron encabezados 'Limite inf' / 'Limite sup' en COMISIONES 2026.")

        rows = []
        for r in range(header_row + 1, header_row + 500):
            li = ws.cell(r, 2).value
            ls = ws.cell(r, 3).value
            p4 = ws.cell(r, 4).value
            p3 = ws.cell(r, 5).value
            p2 = ws.cell(r, 6).value
            p1 = ws.cell(r, 7).value

            if li is None or ls is None:
                if rows:
                    break
                continue

            try:
                rows.append({
                    "lim_inf": float(li), "lim_sup": float(ls),
                    "p4": float(p4), "p3": float(p3),
                    "p2": float(p2), "p1": float(p1),
                })
            except Exception:
                continue

        if not rows:
            raise ValueError("No se pudo leer la tabla de 'COMISIONES 2026'.")

        self.comm_table = pd.DataFrame(rows).sort_values("lim_inf").reset_index(drop=True)

    def _load_price_map(self, wb):
        if "NUEVAS LISTAS" not in wb.sheetnames:
            raise ValueError("No se encontró la hoja 'NUEVAS LISTAS' en el esquema.")
        ws = wb["NUEVAS LISTAS"]

        header_row = None
        for r in range(1, 80):
            v = ws.cell(r, 2).value
            if isinstance(v, str) and v.strip().upper() == "MODELO":
                header_row = r
                break
        if header_row is None:
            raise ValueError("No se encontró encabezado 'MODELO' en NUEVAS LISTAS.")

        p4_col, p3_col, p2_col, p1_col = 5, 8, 11, 14

        for r in range(header_row + 1, ws.max_row + 1):
            prod = ws.cell(r, 2).value
            prod_s = str(prod).strip() if prod is not None else ""
            if not prod_s:
                continue

            key = prod_s.strip().upper()
            p4 = ws.cell(r, p4_col).value
            p3 = ws.cell(r, p3_col).value
            p2 = ws.cell(r, p2_col).value
            p1 = ws.cell(r, p1_col).value

            if all(v is None for v in (p4, p3, p2, p1)):
                continue

            self.price_map[key] = {
                "p4": _safe_float(p4), "p3": _safe_float(p3),
                "p2": _safe_float(p2), "p1": _safe_float(p1),
            }


# ─── Commission Calculation ──────────────────────────────────

def _pick_commission_row(comm_table: pd.DataFrame, total_sales: float):
    min_inf = float(comm_table["lim_inf"].min())
    max_sup = float(comm_table["lim_sup"].max())

    if total_sales < MIN_COMMISSION_TOTAL:
        return None
    if total_sales < min_inf:
        return comm_table.iloc[0]
    if total_sales > max_sup:
        return comm_table.iloc[-1]

    m = (comm_table["lim_inf"] <= total_sales) & (total_sales <= comm_table["lim_sup"])
    if not m.any():
        return comm_table.iloc[-1]
    return comm_table[m].iloc[0]


def _build_asesor_rate_maps(comm_table: pd.DataFrame, total_by_asesor: Dict[str, float]):
    rate_p1, rate_p2, rate_p3, rate_p4 = {}, {}, {}, {}
    for asesor, total in total_by_asesor.items():
        row = _pick_commission_row(comm_table, float(total))
        if row is None:
            rate_p1[asesor] = rate_p2[asesor] = rate_p3[asesor] = rate_p4[asesor] = 0.0
        else:
            rate_p1[asesor] = float(row["p1"])
            rate_p2[asesor] = float(row["p2"])
            rate_p3[asesor] = float(row["p3"])
            rate_p4[asesor] = float(row["p4"])
    return rate_p1, rate_p2, rate_p3, rate_p4


def _infer_tier_vector(neto, p4, p3, p2, p1):
    missing = p1.isna() | p2.isna() | p3.isna() | p4.isna() | neto.isna()
    tier = np.select(
        [~missing & (neto >= p1), ~missing & (neto >= p2), ~missing & (neto >= p3), ~missing],
        [1, 2, 3, 4],
        default=4,
    )
    return pd.Series(tier, index=neto.index)


# ─── PDF Carátula ────────────────────────────────────────────

def _export_caratula_pdf(path_pdf: str, resumen_df: pd.DataFrame, fecha_ini, fecha_fin):
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(path_pdf, pagesize=LETTER, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)

    story = []
    titulo = f"CALCULO DE COMISIONES DEL {fecha_ini:%d-%b-%Y} AL {fecha_fin:%d-%b-%Y}"
    story.append(Paragraph(titulo.upper(), styles["Title"]))
    story.append(Spacer(1, 14))

    data = [["NOMBRE ASESOR", "VENTAS", "TOTAL COMISION $"]]
    for _, r in resumen_df.iterrows():
        data.append([str(r["Asesor"]), _money(r["Venta Total"]), _money(r["Total comisión"])])
    data.append(["TOTALES", _money(resumen_df["Venta Total"].sum()), _money(resumen_df["Total comisión"].sum())])

    table = Table(data, hAlign="LEFT", colWidths=[280, 120, 120])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.whitesmoke),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(table)
    doc.build(story)


# ─── Excel detail export ─────────────────────────────────────

def _export_detail_xlsx(out_path: str, detail_df: pd.DataFrame, resumen_df: pd.DataFrame):
    """Write detail and summary to a clean XLSX."""
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        detail_df.to_excel(writer, sheet_name="Detalle", index=False)
        resumen_df.to_excel(writer, sheet_name="Resumen", index=False)


# ─── Worker processor entry point ────────────────────────────

def process(ctx: JobContext) -> str:
    """Process ERA Ventas commission job. Returns relative output path."""

    ctx.report_progress(5, "Identificando archivos...")

    # Input files: we expect 1 XLSX (base_comisiones) as input
    # and 1 XLSM (schema) as template
    xlsx_inputs = ctx.input_files(".xlsx") + ctx.input_files(".xls")
    xlsm_inputs = ctx.input_files(".xlsm")

    # The schema can come as template OR as an xlsm input file
    schema_path = None
    base_path = None

    if ctx.template_abs and ctx.template_abs.exists():
        tpl_name = ctx.template_abs.name.lower()
        if tpl_name.endswith(".xlsm") or "comisiones" in tpl_name or "esquema" in tpl_name:
            schema_path = ctx.template_abs
        else:
            # Template might be the base_comisiones
            base_path = ctx.template_abs

    # Check xlsm inputs for schema
    if not schema_path and xlsm_inputs:
        schema_path = xlsm_inputs[0]

    # Check xlsx inputs for base
    if not base_path and xlsx_inputs:
        # If we have multiple xlsx, pick the one that's NOT the schema
        for f in xlsx_inputs:
            if schema_path and f.name == schema_path.name:
                continue
            base_path = f
            break

    # Fallback: if schema not found, maybe it's among xlsx files
    if not schema_path:
        for f in xlsx_inputs:
            if f != base_path:
                schema_path = f
                break

    if not base_path:
        raise ValueError(
            "No se encontró el archivo base_comisiones (.xlsx). "
            "Sube el archivo de base de comisiones como input."
        )

    if not schema_path:
        raise ValueError(
            "No se encontró el esquema de comisiones (.xlsm). "
            "Sube el archivo de esquema como template o como input adicional."
        )

    # ── Load rules ───────────────────────────────────────────
    ctx.report_progress(10, "Cargando esquema de comisiones...")
    rules = Rules2026(str(schema_path))
    rules.load()

    # ── Extract valid OVs from Hoja2 ────────────────────────
    ctx.report_progress(20, "Validando OVs (Hoja2: ov==cruce)...")
    valid_ovs, ok_rows, neq_rows = _extract_valid_ovs_from_hoja2(str(base_path))
    if not valid_ovs:
        raise ValueError("No se encontraron OVs válidas en Hoja2 (ov y cruce empatadas).")

    # ── Read sales data ──────────────────────────────────────
    ctx.report_progress(30, "Leyendo datos de ventas...")
    df = pd.read_excel(
        str(base_path),
        sheet_name=BASE_SHEET_VENTAS,
        engine="openpyxl",
        usecols="A,D,E,H,I,S,T",
    )
    df = df.rename(columns={
        df.columns[0]: "Fecha",
        df.columns[1]: "Asesor",
        df.columns[2]: "Cliente",
        df.columns[3]: "Producto",
        df.columns[4]: "Cantidad",
        df.columns[5]: "Precio Bruto",
        df.columns[6]: "OV",
    })

    # ── Apply filters ────────────────────────────────────────
    ctx.report_progress(40, "Aplicando filtros...")
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df = df.dropna(subset=["Fecha"]).copy()
    df = df[~df["Producto"].apply(_is_tax_line)].copy()
    df["OV"] = df["OV"].apply(_norm_ov)
    df = df[df["OV"].isin(valid_ovs)].copy()
    df["Cantidad"] = pd.to_numeric(df["Cantidad"], errors="coerce")
    df["Precio Bruto"] = pd.to_numeric(df["Precio Bruto"], errors="coerce")
    df = df.dropna(subset=["Producto", "Cantidad", "Precio Bruto"]).copy()

    # Match products against NUEVAS LISTAS
    df["Producto_key"] = df["Producto"].astype(str).str.strip().str.upper()
    valid_products = set(rules.price_map.keys())
    df = df[df["Producto_key"].isin(valid_products)].copy()

    if df.empty:
        raise ValueError(
            "Después de filtrar por ov==cruce + excluir IVA + productos en NUEVAS LISTAS, "
            "no quedaron filas. Revisa que Artículo (base) coincida con MODELO (NUEVAS LISTAS)."
        )

    # ── Use full date range (no date filter — process all data) ──
    ctx.report_progress(50, "Calculando comisiones...")

    # Determine date range from data
    dates = sorted(df["Fecha"].dt.date.unique())
    d_ini = dates[0]
    d_fin = dates[-1]

    # If params specify date range, use them
    params = ctx.params or {}
    if params.get("fecha_inicio"):
        try:
            d_ini = pd.to_datetime(params["fecha_inicio"]).date()
        except Exception:
            pass
    if params.get("fecha_fin"):
        try:
            d_fin = pd.to_datetime(params["fecha_fin"]).date()
        except Exception:
            pass

    if d_ini > d_fin:
        d_ini, d_fin = d_fin, d_ini

    # Filter by period
    df = df[(df["Fecha"].dt.date >= d_ini) & (df["Fecha"].dt.date <= d_fin)].copy()
    if df.empty:
        raise ValueError("No hay filas en el periodo seleccionado.")

    # ── Calculate net price and total sales ──────────────────
    ctx.report_progress(60, "Calculando neto y venta total...")
    df["Asesor"] = df["Asesor"].fillna("").astype(str).str.strip()
    df["Cliente"] = df["Cliente"].fillna("").astype(str).str.strip()
    df["Precio Unitario Neto"] = df["Precio Bruto"] * IVA_FACTOR
    df["Venta Total"] = df["Precio Unitario Neto"] * df["Cantidad"]

    # ── Merge price lists ────────────────────────────────────
    ctx.report_progress(70, "Buscando precios P4–P1...")
    prices_df = pd.DataFrame.from_dict(rules.price_map, orient="index")
    prices_df.index.name = "Producto_key"
    prices_df = prices_df.rename(columns={"p4": "Precio 4", "p3": "Precio 3", "p2": "Precio 2", "p1": "Precio 1"})
    df = df.merge(prices_df, left_on="Producto_key", right_index=True, how="left")

    # ── Commission calculation ───────────────────────────────
    ctx.report_progress(80, "Calculando comisión por asesor...")
    total_by_asesor = df.groupby("Asesor", dropna=False)["Venta Total"].sum().to_dict()
    comm_table = rules.comm_table
    rate_p1, rate_p2, rate_p3, rate_p4 = _build_asesor_rate_maps(comm_table, total_by_asesor)

    tier = _infer_tier_vector(
        df["Precio Unitario Neto"],
        df["Precio 4"], df["Precio 3"], df["Precio 2"], df["Precio 1"],
    )

    r1 = df["Asesor"].map(rate_p1).fillna(0.0)
    r2 = df["Asesor"].map(rate_p2).fillna(0.0)
    r3 = df["Asesor"].map(rate_p3).fillna(0.0)
    r4 = df["Asesor"].map(rate_p4).fillna(0.0)

    df["Comisión"] = np.select(
        [tier == 1, tier == 2, tier == 3, tier == 4],
        [r1, r2, r3, r4],
        default=0.0,
    )
    df["Total comisión"] = df["Comisión"] * df["Venta Total"]

    # ── Build output ─────────────────────────────────────────
    ctx.report_progress(90, "Generando archivos de salida...")
    out_detail = df[DISPLAY_COLS].copy()
    resumen = out_detail.groupby(["Asesor"], dropna=False).agg({
        "Venta Total": "sum",
        "Total comisión": "sum",
    }).reset_index().sort_values("Asesor")

    # Write detail XLSX
    detail_path = ctx.output_path("comisiones_detalle.xlsx")
    _export_detail_xlsx(str(detail_path), out_detail, resumen)

    # Write carátula PDF
    pdf_path = ctx.output_path("caratula_comisiones.pdf")
    d_ini_dt = datetime.combine(d_ini, datetime.min.time())
    d_fin_dt = datetime.combine(d_fin, datetime.min.time())
    _export_caratula_pdf(str(pdf_path), resumen, d_ini_dt, d_fin_dt)

    ctx.report_progress(95, "Comisiones calculadas exitosamente")
    return ctx.output_rel("comisiones_detalle.xlsx")
