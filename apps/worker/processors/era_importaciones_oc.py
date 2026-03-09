from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable
from zoneinfo import ZoneInfo

import openpyxl
import pdfplumber
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .base import JobContext

APP_KEY = "era_importaciones_generador_oc"
WORKSHEET_NAME = "PLAN DE ENTREGAS "

BUYER_NAME = "ENERGÍA RENOVABLE DE AMÉRICA SA DE CV"
BUYER_ADDRESS = (
    "MATÍAS DE BOCANEGRA #42 COL. EL MIRADOR DEL PUNHUATO C.P. 58249 "
    "MORELIA, MICHOACÁN, MÉXICO"
)
BUYER_RFC = "ERA080725618"
LETTER_CITY = "Morelia, Michoacán"
LETTER_COMPANY_LABEL = "RAZÓN SOCIAL ERA"
LETTER_BRAND = "ERA Energía Renovable de América"
LETTER_ADDRESSEE = (
    "ING. MARCO ANTONIO RAMOS TORRES TITULAR DE LA ADUANA DE LÁZARO CÁRDENAS "
    "DE LA AGENCIA NACIONAL DE ADUANAS DE MÉXICO"
)
LETTER_SUBJECT = "ASUNTO: CARTA COMPLEMENTARIA (FACTURA)"
LETTER_LEGAL_REP = "LIZBETH LÓPEZ RODRÍGUEZ"
LETTER_SIGN_OFF = "REPRESENTANTE LEGAL"
LETTER_IMPORTER_LINES = [
    "ENERGÍA RENOVABLE DE AMÉRICA SA DE CV",
    "MATÍAS DE BOCANEGRA #34-A COL. EL MIRADOR DEL PUNHUATO C.P. 58249",
    "MORELIA, MICHOACÁN DE OCAMPO, MÉXICO",
    "RFC: ERA080725618",
]

CONTAINER_RX = re.compile(r"\b[A-Z]{4}\d{7}\b")
MODEL_RX = re.compile(r"\b[A-Z]{2,6}-\d{2,3}[A-Z]?\b")
PACKING_ITEM_RX = re.compile(
    r"^(?P<description>.+?)\s+(?P<qty>\d+)\s+(?P<gross>\d+(?:\.\d+)?)KGS\s+"
    r"(?P<net>\d+(?:\.\d+)?)KGS\s+(?P<m3>\d+(?:\.\d+)?)M3$",
    re.IGNORECASE,
)
PRICE_ITEM_RX = re.compile(
    r"^\d+(?:\.\d+)?\s+(?P<model>[A-Z0-9-]+)\s+(?P<description>.+?)\s+"
    r"(?P<qty>\d+)\s+(?P<serial>[A-Z0-9-]+)\s+\$(?P<unit>[\d,]+\.\d{2})\s+"
    r"\$(?P<total>[\d,]+\.\d{2})$",
    re.IGNORECASE,
)
UPLOAD_PREFIX_RX = re.compile(r"^\d{2}-(?P<name>.+)$")
STOP_BLOCK_MARKERS = (
    "CONSIGNEE",
    "IMPORTER",
    "IMPORTADOR",
    "TO INVOICE",
    "PACKING LIST",
    "PORT ",
    "VESSEL",
    "COUNTRY OF ORIGIN",
    "PAIS DE ORIGEN",
    "INCOTERM",
    "FLETE",
    "TYPE OF CONTAINER",
    "TIPE OF CONTAINER",
)
MONTHS = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}
MONTHS_ES = {
    1: "enero",
    2: "febrero",
    3: "marzo",
    4: "abril",
    5: "mayo",
    6: "junio",
    7: "julio",
    8: "agosto",
    9: "septiembre",
    10: "octubre",
    11: "noviembre",
    12: "diciembre",
}


@dataclass
class OrderItem:
    model: str | None
    description: str
    quantity: int | None
    gross_weight_kg: float | None = None
    net_weight_kg: float | None = None
    volume_m3: float | None = None
    unit_price_usd: float | None = None
    total_price_usd: float | None = None
    serial: str | None = None


@dataclass
class OrderRecord:
    source_file: str
    order_number: str | None = None
    general_po: str | None = None
    invoice_number: str | None = None
    order_date: datetime | None = None
    supplier_name: str | None = None
    supplier_address_lines: list[str] = field(default_factory=list)
    provider_alias: str | None = None
    container: str | None = None
    incoterm: str | None = None
    origin_port: str | None = None
    destination_port: str | None = None
    tax_id: str | None = None
    country_origin: str | None = None
    freight_usd: float | None = None
    etd: datetime | None = None
    eta: datetime | None = None
    terminal: str | None = None
    forwarder: str | None = None
    transportista: str | None = None
    despacho: str | None = "ALMACÉN"
    items: list[OrderItem] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def preferred_number(self) -> str:
        return self.order_number or self.invoice_number or Path(self.source_file).stem

    def provider_label(self) -> str:
        if self.provider_alias:
            return self.provider_alias
        if not self.supplier_name:
            return ""
        parts = [part for part in re.split(r"\s+", self.supplier_name.upper()) if part]
        for part in parts:
            if part not in {"LTD", "LIMITED", "CO", "COMPANY", "SA", "DE", "CV", "POWER"}:
                return part
        return parts[0] if parts else ""

    def goods_summary(self) -> str:
        grouped: dict[str, int | None] = {}
        ordered_labels: list[str] = []
        for item in self.items:
            label = item.model or item.description
            if label not in grouped:
                ordered_labels.append(label)
                grouped[label] = 0 if item.quantity is not None else None
            if item.quantity is None or grouped[label] is None:
                grouped[label] = None
            else:
                grouped[label] += item.quantity

        summary: list[str] = []
        for label in ordered_labels:
            quantity = grouped[label]
            summary.append(f"{label} ({quantity})" if quantity is not None else label)

        joined = " ".join(summary).strip()
        return joined[:500]

    def total_usd(self) -> float | None:
        totals = [item.total_price_usd for item in self.items if item.total_price_usd is not None]
        return round(sum(totals), 2) if totals else None



def _safe_name(raw: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", raw.strip())
    return cleaned[:120] or "output"



def _display_source_name(filename: str) -> str:
    match = UPLOAD_PREFIX_RX.match(filename)
    return match.group("name") if match else filename



def _display_provider_name(record: OrderRecord, params: dict | None = None) -> str:
    params = params or {}
    manual = (params.get("provider_alias") or "").strip()
    return manual or record.provider_label()



def _parse_float(raw: str | None) -> float | None:
    if not raw:
        return None
    try:
        return float(raw.replace(",", "").strip())
    except Exception:
        return None



def _parse_date_any(raw: str | None) -> datetime | None:
    if not raw:
        return None

    text = raw.strip().replace("/", "-").replace(".", "-")

    month_match = re.search(r"\b(\d{1,2})-([A-Za-z]{3})-?(20\d{2})\b", text)
    if month_match:
        day, month, year = month_match.groups()
        month_number = MONTHS.get(month.upper())
        if month_number:
            try:
                return datetime(int(year), month_number, int(day))
            except ValueError:
                return None

    for fmt in ("%d-%m-%Y", "%m-%d-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass

    numeric = re.search(r"\b(20\d{2})-(\d{1,2})-(\d{1,2})\b", text)
    if numeric:
        year, month, day = numeric.groups()
        try:
            return datetime(int(year), int(month), int(day))
        except ValueError:
            return None

    return None



def _non_empty_lines(text: str) -> list[str]:
    return [line.strip() for line in text.splitlines() if line and line.strip()]



def _extract_pdf_text(path: Path) -> tuple[str, list[str]]:
    pages: list[str] = []
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            pages.append(page.extract_text() or "")
    text = "\n".join(pages).replace("\u00a0", " ")
    lines = _non_empty_lines(text)
    return text, lines



def _extract_supplier_name(lines: Iterable[str], labels: tuple[str, ...]) -> str | None:
    line_list = list(lines)
    for idx, line in enumerate(line_list):
        upper = line.upper()
        if any(label in upper for label in labels):
            tail = re.sub(r"^EXPORTER\s*(\([^)]*\))?\s*", "", line, flags=re.IGNORECASE).strip(" :.-")
            if tail and tail.upper() not in labels:
                return tail
            for next_line in line_list[idx + 1 : idx + 4]:
                if next_line and next_line.upper() not in labels:
                    return next_line.strip()
    return None



def _extract_labeled_block(lines: Iterable[str], labels: tuple[str, ...], max_lines: int = 5) -> list[str]:
    line_list = list(lines)
    for idx, line in enumerate(line_list):
        upper = line.upper()
        if not any(label in upper for label in labels):
            continue

        block: list[str] = []
        tail = re.sub(r"^EXPORTER\s*(\([^)]*\))?\s*", "", line, flags=re.IGNORECASE)
        tail = re.sub(r"^PROVEEDOR\s*:?\s*", "", tail, flags=re.IGNORECASE)
        tail = tail.strip(" :.-")
        if tail:
            block.append(tail)

        for next_line in line_list[idx + 1 : idx + 1 + max_lines]:
            candidate = next_line.strip()
            upper_candidate = candidate.upper()
            if not candidate:
                break
            if any(upper_candidate.startswith(marker) for marker in STOP_BLOCK_MARKERS):
                break
            block.append(candidate)

        return block

    return []



def _extract_invoice_from_lines(lines: list[str]) -> tuple[str | None, datetime | None]:
    for index, line in enumerate(lines):
        if "TO INVOICE NO. DATE" not in line.upper():
            continue

        candidate = lines[index + 1] if index + 1 < len(lines) else ""
        match = re.search(
            r"([A-Z0-9-]{4,})\s+(\d{1,2}-[A-Z]{3}-?20\d{2}|\d{1,2}[/-]\d{1,2}[/-]20\d{2})$",
            candidate,
            re.IGNORECASE,
        )
        if match:
            return match.group(1).strip(), _parse_date_any(match.group(2))

    for line in lines:
        match = re.search(
            r"FACTURA\s+NO\.?\s*([A-Z0-9-]+)\s+DE\s+FECHA\s+(\d{1,2}[/-]\d{1,2}[/-]20\d{2})",
            line,
            re.IGNORECASE,
        )
        if match:
            return match.group(1).strip(), _parse_date_any(match.group(2))

    return None, None



def _parse_packing_items(lines: Iterable[str]) -> list[OrderItem]:
    items: list[OrderItem] = []
    for line in lines:
        match = PACKING_ITEM_RX.match(line.strip())
        if not match:
            continue
        description = match.group("description").strip()
        models = MODEL_RX.findall(description)
        model = models[-1] if models else None
        qty = int(match.group("qty"))
        items.append(
            OrderItem(
                model=model,
                description=description,
                quantity=qty,
                gross_weight_kg=_parse_float(match.group("gross")),
                net_weight_kg=_parse_float(match.group("net")),
                volume_m3=_parse_float(match.group("m3")),
            )
        )
    return items



def _parse_price_items(lines: Iterable[str]) -> list[OrderItem]:
    items: list[OrderItem] = []
    for line in lines:
        match = PRICE_ITEM_RX.match(line.strip())
        if not match:
            continue
        items.append(
            OrderItem(
                model=match.group("model").strip(),
                description=match.group("description").strip(),
                quantity=int(match.group("qty")),
                serial=match.group("serial").strip(),
                unit_price_usd=_parse_float(match.group("unit")),
                total_price_usd=_parse_float(match.group("total")),
            )
        )
    return items



def _parse_order_pdf(path: Path) -> OrderRecord:
    text, lines = _extract_pdf_text(path)
    upper = text.upper()

    record = OrderRecord(source_file=_display_source_name(path.name))
    record.container = CONTAINER_RX.search(text).group(0) if CONTAINER_RX.search(text) else None

    order_match = re.search(r"NO\.\s*PO:\s*([A-Z0-9+/\-]+)(?:\s+DATE:|\b)", text, re.IGNORECASE)
    if order_match:
        record.order_number = order_match.group(1).strip()

    general_match = re.search(r"NO\.\s*GENERAL\s*PO:\s*([A-Z0-9\-]+)", text, re.IGNORECASE)
    if general_match:
        record.general_po = general_match.group(1).strip()

    invoice_number, invoice_date = _extract_invoice_from_lines(lines)
    if invoice_number:
        record.invoice_number = invoice_number
        record.order_date = invoice_date

    po_date_match = re.search(r"DATE:\s*(\d{1,2}[/-]\d{1,2}[/-]20\d{2})", text, re.IGNORECASE)
    if po_date_match and record.order_date is None:
        record.order_date = _parse_date_any(po_date_match.group(1))

    provider_match = re.search(r"PROVE?DOR\s+([A-Z/& ]+)", text, re.IGNORECASE)
    if provider_match:
        record.provider_alias = provider_match.group(1).strip()

    supplier_block = _extract_labeled_block(lines, ("EXPORTER", "PROVEEDOR"))
    if supplier_block:
        record.supplier_name = supplier_block[0]
        record.supplier_address_lines = supplier_block[1:]
    else:
        record.supplier_name = _extract_supplier_name(lines, ("EXPORTER",))

    if not record.supplier_name:
        supplier_match = re.search(
            r"TOTAL,\s*USD\s*\$?[\d,]+\.\d{2}\s+([A-Z][A-Z0-9 .,&/-]{3,})\s+PAIS DE ORIGEN",
            text,
            re.IGNORECASE | re.DOTALL,
        )
        if supplier_match:
            record.supplier_name = " ".join(supplier_match.group(1).split())

    tax_id_match = re.search(r"TAX\s*ID[:\s]*([A-Z0-9-]+)", text, re.IGNORECASE)
    if tax_id_match:
        record.tax_id = tax_id_match.group(1).strip()

    incoterm_match = re.search(r"INCOTERM:\s*([A-Z0-9]+)", text, re.IGNORECASE)
    if incoterm_match:
        record.incoterm = incoterm_match.group(1).strip().upper()

    country_origin_match = re.search(r"PA[IÍ]S\s+DE\s+ORIGEN:\s*([A-ZÁÉÍÓÚ .,-]+)", text, re.IGNORECASE)
    if country_origin_match:
        record.country_origin = " ".join(country_origin_match.group(1).split())

    freight_match = re.search(r"FLETE\s+MAR[ÍI]TIMO:\s*\$?([\d,]+(?:\.\d+)?)", text, re.IGNORECASE)
    if freight_match:
        record.freight_usd = _parse_float(freight_match.group(1))

    etd_match = re.search(r"ETD:\s*([0-9./-]{8,10})", text, re.IGNORECASE)
    if etd_match:
        record.etd = _parse_date_any(etd_match.group(1))

    eta_match = re.search(r"ETA:\s*([0-9./-]{8,10})", text, re.IGNORECASE)
    if eta_match:
        record.eta = _parse_date_any(eta_match.group(1))

    from_match = re.search(
        r"PORT\s+OR\s+LOADING:\s*(.+?)(?:\s+PORT\s+OF\s+DISCHARGE:|$)",
        text,
        re.IGNORECASE | re.DOTALL,
    )
    if from_match:
        record.origin_port = " ".join(from_match.group(1).split())

    to_match = re.search(
        r"PORT\s+OF\s+DISCHARGE:\s*(.+?)(?:\s+TIPE\s+OF\s+CONTAINER:|\s+TYPE\s+OF\s+CONTAINER:|$)",
        text,
        re.IGNORECASE | re.DOTALL,
    )
    if to_match:
        record.destination_port = " ".join(to_match.group(1).split())

    if not record.origin_port or not record.destination_port:
        vessel_match = re.search(
            r"VESSEL\s+([A-ZÁÉÍÓÚ ,.-]+)\s+TO\s+([A-ZÁÉÍÓÚ ,.-]+)",
            upper,
            re.IGNORECASE,
        )
        if vessel_match:
            record.origin_port = record.origin_port or " ".join(vessel_match.group(1).split())
            record.destination_port = record.destination_port or " ".join(vessel_match.group(2).split())

    price_items = _parse_price_items(lines)
    packing_items = _parse_packing_items(lines)
    if price_items:
        record.items = price_items
    elif packing_items:
        record.items = packing_items
    else:
        record.warnings.append("No se detectaron partidas en el PDF")

    if not record.order_number and record.invoice_number:
        record.order_number = record.invoice_number

    if record.destination_port and "LAZARO" in record.destination_port.upper():
        record.terminal = "APM"
    elif record.destination_port and "MANZANILLO" in record.destination_port.upper():
        record.terminal = "MANZANILLO"

    if not record.country_origin and record.origin_port and "," in record.origin_port:
        record.country_origin = record.origin_port.split(",")[-1].strip()

    if "PACKING LIST" in upper and record.container is None:
        record.warnings.append("No se detectó contenedor en el packing list")

    return record



def _copy_row_style(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.protection = copy(src.protection)
            dst.number_format = src.number_format
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height



def _next_available_row(ws) -> int:
    last_row = 2
    for row in range(ws.max_row, 2, -1):
        if any(ws.cell(row, col).value not in (None, "") for col in range(1, 26)):
            last_row = row
            break
    return last_row + 1



def _formula_from_eta(row: int, offset: int) -> str:
    sign = "+" if offset >= 0 else "-"
    return f"=+M{row}{sign}{abs(offset)}"



def _append_record_to_workbook(ws, record: OrderRecord, row: int, params: dict) -> None:
    source_row = max(3, row - 1)
    _copy_row_style(ws, source_row, row)

    forwarder = params.get("forwarder") or record.forwarder
    transportista = params.get("transportista") or record.transportista
    despacho = params.get("despacho") or record.despacho or "ALMACÉN"
    terminal = params.get("terminal") or record.terminal
    provider = _display_provider_name(record, params)

    ws.cell(row, 1).value = "DOCUMENTACIÓN ENVIADA" if record.eta or record.etd else None
    ws.cell(row, 2).value = record.preferred_number()
    ws.cell(row, 3).value = params.get("referencia_visa")
    ws.cell(row, 4).value = terminal
    ws.cell(row, 5).value = record.container
    ws.cell(row, 6).value = provider
    ws.cell(row, 7).value = forwarder
    ws.cell(row, 8).value = _formula_from_eta(row, -5) if record.eta else None
    ws.cell(row, 9).value = "PENDIENTE" if record.eta else None
    ws.cell(row, 10).value = _formula_from_eta(row, 30) if record.eta else None
    ws.cell(row, 11).value = "PENDIENTE" if record.eta else None
    ws.cell(row, 12).value = record.etd
    ws.cell(row, 13).value = record.eta
    ws.cell(row, 14).value = _formula_from_eta(row, 6) if record.eta else None
    ws.cell(row, 15).value = _formula_from_eta(row, 1) if record.eta else None
    ws.cell(row, 16).value = f"=+O{row}+1" if record.eta else None
    ws.cell(row, 17).value = f"=+P{row}+1" if record.eta else None
    ws.cell(row, 18).value = _formula_from_eta(row, 20) if record.eta else None
    ws.cell(row, 19).value = record.goods_summary() or None
    ws.cell(row, 20).value = transportista
    ws.cell(row, 21).value = despacho



def _build_pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="Meta",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#334155"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="SectionTitle",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=13,
            textColor=colors.HexColor("#0f172a"),
            spaceAfter=6,
            spaceBefore=8,
        )
    )
    styles.add(
        ParagraphStyle(
            name="LetterBody",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            textColor=colors.black,
            alignment=4,
        )
    )
    styles.add(
        ParagraphStyle(
            name="LetterHeading",
            parent=styles["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=13,
            textColor=colors.black,
        )
    )
    styles.add(
        ParagraphStyle(
            name="LetterSmall",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.black,
        )
    )
    styles.add(
        ParagraphStyle(
            name="LetterSmallBold",
            parent=styles["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.black,
        )
    )
    return styles



def _paragraph(text: str, style_name: str, styles) -> Paragraph:
    return Paragraph(text.replace("\n", "<br/>"), styles[style_name])



def _format_letter_date(value: datetime | None = None) -> str:
    date_value = value or datetime.now(ZoneInfo("America/Mexico_City"))
    month_name = MONTHS_ES[date_value.month]
    return f"{LETTER_CITY} a {date_value.day} de {month_name} de {date_value.year}"



def _format_invoice_date(value: datetime | None) -> str:
    return value.strftime("%d/%m/%Y") if value else "SIN FECHA"



def _format_money_usd(value: float | None, digits: int = 2) -> str:
    if value is None:
        return ""
    return f"${value:,.{digits}f}"



def _table_paragraph(text: str, styles, bold: bool = False) -> Paragraph:
    style_name = "LetterSmallBold" if bold else "LetterSmall"
    return _paragraph(text or "", style_name, styles)



def _item_description(item: OrderItem) -> str:
    parts = [item.description]
    if item.model and item.model not in item.description:
        parts.append(f"MODELO: {item.model}")
    if item.serial:
        parts.append(f"No. de serie: {item.serial}")
    return "<br/>".join(parts)



def _build_letter_items_table(record: OrderRecord, styles) -> Table:
    rows = [
        [
            _table_paragraph("CONTENEDOR", styles, bold=True),
            _table_paragraph("DESCRIPCIÓN DE LA MERCANCÍA", styles, bold=True),
            _table_paragraph("MARCA", styles, bold=True),
            _table_paragraph("CANTIDAD/UMC", styles, bold=True),
            _table_paragraph("VALOR UNITARIO USD", styles, bold=True),
            _table_paragraph("VALOR TOTAL USD", styles, bold=True),
        ]
    ]

    if record.items:
        for index, item in enumerate(record.items):
            rows.append(
                [
                    _table_paragraph(record.container if index == 0 else "", styles),
                    _table_paragraph(_item_description(item), styles),
                    _table_paragraph(LETTER_BRAND, styles),
                    _table_paragraph(f"{item.quantity} PZS" if item.quantity is not None else "", styles),
                    _table_paragraph(_format_money_usd(item.unit_price_usd, digits=4), styles),
                    _table_paragraph(_format_money_usd(item.total_price_usd), styles),
                ]
            )
    else:
        rows.append(
            [
                _table_paragraph(record.container or "", styles),
                _table_paragraph("No se detectaron partidas en el PDF.", styles),
                _table_paragraph(LETTER_BRAND, styles),
                _table_paragraph("", styles),
                _table_paragraph("", styles),
                _table_paragraph("", styles),
            ]
        )

    table = Table(
        rows,
        colWidths=[25 * mm, 68 * mm, 28 * mm, 22 * mm, 24 * mm, 24 * mm],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.7, colors.black),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#475569")),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f8fafc")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table



def _render_order_pdf(record: OrderRecord, out_path: Path, params: dict | None = None) -> None:
    styles = _build_pdf_styles()
    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )

    story: list = []
    invoice_number = record.invoice_number or record.preferred_number()
    provider_name = _display_provider_name(record, params)

    story.append(_paragraph(LETTER_COMPANY_LABEL, "LetterHeading", styles))
    story.append(_paragraph(f"PROVEEDOR {provider_name}", "LetterHeading", styles))
    story.append(_paragraph(_format_letter_date(), "LetterHeading", styles))
    story.append(Spacer(1, 4 * mm))

    story.append(_paragraph(LETTER_ADDRESSEE, "LetterHeading", styles))
    story.append(_paragraph("P R E S E N T E", "LetterHeading", styles))
    story.append(Spacer(1, 3 * mm))
    story.append(_paragraph(LETTER_SUBJECT, "LetterHeading", styles))
    story.append(Spacer(1, 3 * mm))

    letter_body = (
        f"{LETTER_LEGAL_REP}, EN REPRESENTACIÓN DE {BUYER_NAME} Y DE CONFORMIDAD CON LO "
        "DISPUESTO EN EL ARTÍCULO 36 Y 36-A FRACCIÓN I INCISO A) Y ARTÍCULO 65 DE LA LEY "
        "ADUANERA Y REGLA 3.1.8 DE LAS REGLAS GENERALES DE COMERCIO EXTERIOR VIGENTES, "
        "BAJO PROTESTA DE DECIR VERDAD NOS PERMITIMOS EFECTUAR LA ACLARACIÓN DE:"
    )
    story.append(_paragraph(letter_body, "LetterBody", styles))
    story.append(Spacer(1, 2 * mm))
    story.append(
        _paragraph(
            f"LA FACTURA No. {invoice_number} DE FECHA {_format_invoice_date(record.order_date)} "
            "CONFORME LO SIGUIENTE:",
            "LetterHeading",
            styles,
        )
    )
    story.append(Spacer(1, 3 * mm))
    story.append(_paragraph("IMPORTADOR:", "LetterHeading", styles))
    story.append(_paragraph("<br/>".join(LETTER_IMPORTER_LINES), "Meta", styles))
    story.append(Spacer(1, 4 * mm))

    story.append(_build_letter_items_table(record, styles))
    story.append(Spacer(1, 3 * mm))

    total_usd = record.total_usd()
    if total_usd is not None:
        story.append(_paragraph(f"TOTAL, USD {_format_money_usd(total_usd)}", "LetterHeading", styles))
        story.append(Spacer(1, 4 * mm))

    supplier_lines = [record.supplier_name or ""]
    supplier_lines.extend(record.supplier_address_lines)
    supplier_lines = [line for line in supplier_lines if line]
    if supplier_lines:
        story.append(_paragraph("PROVEEDOR:", "LetterHeading", styles))
        story.append(_paragraph("<br/>".join(supplier_lines), "Meta", styles))
        story.append(Spacer(1, 2 * mm))

    if record.tax_id:
        story.append(_paragraph(f"TAX ID: {record.tax_id}", "Meta", styles))
    if record.country_origin:
        story.append(_paragraph(f"PAÍS DE ORIGEN: {record.country_origin}", "Meta", styles))
    if record.incoterm:
        story.append(_paragraph(f"INCOTERM: {record.incoterm}", "Meta", styles))
    if record.freight_usd is not None:
        story.append(_paragraph(f"FLETE MARÍTIMO: {_format_money_usd(record.freight_usd)} USD", "Meta", styles))

    story.append(Spacer(1, 4 * mm))
    if record.warnings:
        story.append(_paragraph("Observaciones", "SectionTitle", styles))
        story.append(_paragraph("<br/>".join(record.warnings), "Meta", styles))
        story.append(Spacer(1, 3 * mm))

    story.append(
        _paragraph(
            "SIN MÁS POR EL MOMENTO, AGRADEZCO SU AMABLE ATENCIÓN Y, ASÍ MISMO, "
            "ME PONGO A SU DISPOSICIÓN PARA CUALQUIER ACLARACIÓN.",
            "LetterBody",
            styles,
        )
    )
    story.append(Spacer(1, 8 * mm))
    story.append(_paragraph("ATENTAMENTE", "LetterHeading", styles))
    story.append(Spacer(1, 10 * mm))
    story.append(_paragraph(LETTER_LEGAL_REP, "LetterHeading", styles))
    story.append(_paragraph(LETTER_SIGN_OFF, "Meta", styles))

    doc.build(story)



def process(ctx: JobContext) -> str:
    pdf_inputs = ctx.input_files(".pdf")
    if not pdf_inputs:
        raise ValueError("Se requiere al menos un PDF de proveedor.")

    if not ctx.template_abs or not ctx.template_abs.exists():
        raise ValueError("Se requiere el archivo de Excel de programación de entregas.")

    ctx.report_progress(8, "Leyendo PDFs de proveedor...")

    records: list[OrderRecord] = []
    for index, pdf_path in enumerate(pdf_inputs, start=1):
        percent = 8 + int(32 * index / len(pdf_inputs))
        ctx.report_progress(percent, f"Extrayendo información de {pdf_path.name}...")
        record = _parse_order_pdf(pdf_path)
        if not record.order_number and not record.invoice_number and not record.items:
            raise ValueError(f"No se pudo extraer información utilizable de {pdf_path.name}")
        records.append(record)

    ctx.report_progress(48, "Actualizando programación de entregas...")

    workbook = openpyxl.load_workbook(str(ctx.template_abs))
    sheet = workbook[WORKSHEET_NAME] if WORKSHEET_NAME in workbook.sheetnames else workbook.active

    next_row = _next_available_row(sheet)
    for record in records:
        _append_record_to_workbook(sheet, record, next_row, ctx.params or {})
        next_row += 1

    workbook_name = f"{_safe_name(Path(ctx.template_abs.name).stem)}_actualizado.xlsx"
    workbook_out = ctx.output_path(workbook_name)
    workbook.save(str(workbook_out))

    ctx.report_progress(72, "Generando cartas complementarias...")

    for index, record in enumerate(records, start=1):
        percent = 72 + int(23 * index / len(records))
        ctx.report_progress(percent, f"Generando carta para {record.preferred_number()}...")
        pdf_name = f"CARTA_{_safe_name(record.preferred_number())}.pdf"
        pdf_out = ctx.output_path(pdf_name)
        _render_order_pdf(record, pdf_out, ctx.params or {})

    ctx.report_progress(98, f"Generados {len(records)} PDF(s) tipo carta y un Excel actualizado.")
    return ctx.output_rel(workbook_out.name)
