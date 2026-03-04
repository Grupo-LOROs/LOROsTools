"""
ERA Compras – Generador desde Órdenes de Compra (FacturaChnToMx)
────────────────────────────────────────────────────────────────
Converts Chinese commercial invoice (CI) PDFs into Mexican
"Carta de Porte" Excel files, filling a template with extracted fields.

Input:  One XLSX template (the Carta template) via template upload
        + one or more PDF files (Chinese invoices)
Output: One XLSX per PDF (or per model if multi-model invoice).

Adapted from: FacturaChnToMx/ci_pdf_to_carta.py
"""

import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List

import pdfplumber
import openpyxl

from .base import JobContext


# ─── Date parsing ────────────────────────────────────────────

def _parse_date_any(s: str) -> Optional[datetime]:
    s = s.strip().replace("/", "-").replace(".", "-")
    patterns = [
        (re.compile(r"\b(20\d{2})-(\d{1,2})-(\d{1,2})\b"), "YMD"),
        (re.compile(r"\b(\d{1,2})-(\d{1,2})-(20\d{2})\b"), "DMY_or_MDY"),
    ]
    for rx, kind in patterns:
        m = rx.search(s)
        if not m:
            continue
        a, b, c = m.groups()
        if kind == "YMD":
            try:
                return datetime(int(a), int(b), int(c))
            except ValueError:
                return None
        if kind == "DMY_or_MDY":
            x, y, yyyy = int(a), int(b), int(c)
            if x > 12:
                dd, mm = x, y
            elif y > 12:
                mm, dd = x, y
            else:
                mm, dd = x, y
            try:
                return datetime(yyyy, mm, dd)
            except ValueError:
                return None
    candidates = re.findall(r"\b\d{1,2}-\d{1,2}-20\d{2}\b|\b20\d{2}-\d{1,2}-\d{1,2}\b", s)
    for c in candidates:
        for fmt in ("%Y-%m-%d", "%m-%d-%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(c, fmt)
            except ValueError:
                pass
    return None


# ─── PDF text extraction ─────────────────────────────────────

def _normalize_spaces(text: str) -> str:
    return re.sub(r"[ \t]+", " ", text.replace("\u00a0", " ")).strip()


def _extract_pdf_text(pdf_path: Path) -> str:
    chunks = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            chunks.append(t)
    return _normalize_spaces("\n".join(chunks))


# ─── Field extraction from CI PDF ────────────────────────────

def _extract_fields_from_ci_pdf(text: str) -> Dict[str, Optional[str]]:
    t = text

    invoice_no = None
    m = re.search(r"BUYER\s+INVOICE\s+PO\s+([A-Z0-9\-]+)", t, flags=re.IGNORECASE)
    if m:
        invoice_no = m.group(1).strip()

    invoice_date = None
    m = re.search(r"INVOICE\s+DATE[:\s]*(\d{1,2}/\d{1,2}/20\d{2})", t, flags=re.IGNORECASE)
    if m:
        try:
            invoice_date = datetime.strptime(m.group(1), "%m/%d/%Y")
        except ValueError:
            invoice_date = None

    container = None
    m = re.search(r"\b([A-Z]{4}\d{7})\b", t)
    if m:
        container = m.group(1)

    models = re.findall(r"\b([A-Z]{2,5}-\d{2,3})\b", t)
    models = list(dict.fromkeys(models))

    qty = None
    m = re.search(r"\b(QTY|QUANTITY)[:\s]*([0-9]+)\b", t, flags=re.IGNORECASE)
    if m:
        qty = m.group(2)

    unit_price = None
    m = re.search(r"\b(UNIT\s+PRICE|PRICE)[:\s]*([0-9]+(?:\.[0-9]+)?)\b", t, flags=re.IGNORECASE)
    if m:
        unit_price = m.group(2)

    serie_del = None
    serie_al = None
    m = re.search(r"\bDEL[:\s]*([A-Z0-9\-]+)\b", t, flags=re.IGNORECASE)
    if m:
        serie_del = m.group(1).strip()
    m = re.search(r"\bAL[:\s]*([A-Z0-9\-]+)\b", t, flags=re.IGNORECASE)
    if m:
        serie_al = m.group(1).strip()

    umc = None
    m = re.search(r"\bUMC[:\s]*([A-Z]{2,10})\b", t, flags=re.IGNORECASE)
    if m:
        umc = m.group(1).strip()

    return {
        "invoice_no": invoice_no,
        "invoice_date": invoice_date.strftime("%Y-%m-%d") if invoice_date else None,
        "container": container,
        "models": models or None,
        "qty": qty,
        "umc": umc,
        "unit_price": unit_price,
        "serie_del": serie_del,
        "serie_al": serie_al,
    }


# ─── Excel template filling ──────────────────────────────────

def _set_or_clear(cell, value):
    if value is None or value == "":
        cell.value = None
    else:
        cell.value = value


def _fill_carta_template(
    carta_template_path: Path,
    out_path: Path,
    fields: Dict[str, Optional[str]],
    model: Optional[str] = None,
):
    wb = openpyxl.load_workbook(str(carta_template_path))
    ws = wb["MODELO"]

    _set_or_clear(ws["C14"], fields.get("invoice_no"))

    date_str = fields.get("invoice_date")
    if date_str:
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            ws["E14"].value = dt
        except Exception:
            ws["E14"].value = None
    else:
        ws["E14"].value = None

    _set_or_clear(ws["B24"], fields.get("container"))

    chosen_model = model
    if not chosen_model:
        models_list = fields.get("models") or []
        chosen_model = models_list[0] if models_list else None
    _set_or_clear(ws["D25"], chosen_model)

    _set_or_clear(ws["F24"], int(fields["qty"]) if fields.get("qty") and fields["qty"].isdigit() else None)
    _set_or_clear(ws["G24"], fields.get("umc"))
    _set_or_clear(ws["H24"], float(fields["unit_price"]) if fields.get("unit_price") else None)
    _set_or_clear(ws["D27"], fields.get("serie_del"))
    _set_or_clear(ws["D28"], fields.get("serie_al"))

    wb.save(str(out_path))


def _generate_from_pdf(
    pdf_path: Path,
    carta_template_path: Path,
    out_dir: Path,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    text = _extract_pdf_text(pdf_path)
    fields = _extract_fields_from_ci_pdf(text)

    models = fields.get("models") or []
    if len(models) <= 1:
        out_path = out_dir / f"CARTA_{pdf_path.stem}.xlsx"
        _fill_carta_template(carta_template_path, out_path, fields)
        return [out_path]

    outputs = []
    for m in models:
        out_path = out_dir / f"CARTA_{pdf_path.stem}_{m}.xlsx"
        _fill_carta_template(carta_template_path, out_path, fields, model=m)
        outputs.append(out_path)
    return outputs


# ─── Worker processor entry point ────────────────────────────

def process(ctx: JobContext) -> str:
    """Process ERA Compras OC generation job. Returns relative output path."""
    ctx.report_progress(5, "Leyendo archivos de entrada...")

    # Input: PDFs (commercial invoices) — could be XLSX too, but primarily PDFs
    input_files = ctx.input_files(".pdf")
    xlsx_inputs = ctx.input_files(".xlsx") + ctx.input_files(".xls")

    # The template is the Carta de Porte Excel template
    if not ctx.template_abs or not ctx.template_abs.exists():
        # If no template provided, check if an xlsx was uploaded as input
        if xlsx_inputs:
            template_path = xlsx_inputs[0]
            # Remaining xlsx files are also inputs (unlikely but handle)
        else:
            raise ValueError(
                "Se requiere la plantilla de Carta de Porte (archivo XLSX). "
                "Súbela como template al crear el job."
            )
    else:
        template_path = ctx.template_abs

    if not input_files:
        raise ValueError("No se encontraron archivos PDF de facturas comerciales (CI)")

    ctx.report_progress(10, f"Procesando {len(input_files)} factura(s)...")

    all_outputs = []
    for i, pdf in enumerate(input_files):
        pct = 10 + int(80 * (i + 1) / len(input_files))
        ctx.report_progress(pct, f"Procesando {pdf.name}...")

        outputs = _generate_from_pdf(pdf, template_path, ctx.output_dir)
        all_outputs.extend(outputs)

    ctx.report_progress(95, f"Generados {len(all_outputs)} archivo(s) de Carta de Porte")

    # If single output, return it directly
    if len(all_outputs) == 1:
        return ctx.output_rel(all_outputs[0].name)

    # If multiple outputs, return the first one (all are in output_dir)
    # The portal can list all files in the output directory
    return ctx.output_rel(all_outputs[0].name)
