"""
ERA Proyectos – Comisionador CFE (CFEDataExtraction)
─────────────────────────────────────────────────────
Extracts consumption and tariff data from CFE PDF receipts,
enriches with live CFE tariffs and NASA solar irradiance data,
and generates an Excel workbook with formulas for solar PV ROI analysis.

Input:  Multiple PDFs (CFE electrical bill receipts)
        + Optional XLSX template (CONSUMO.xlsx)
Output: Single XLSX with extracted data, tariffs, NASA hours, and formulas.

Adapted from: CFEDataExtraction/extract_pdf_data.py (latest version)
"""

import os
import re
import traceback
import unicodedata
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional, Dict, List

import pdfplumber
import openpyxl

from .base import JobContext

# ─── We import from the cfe_lib package (copied from CFEDataExtraction) ──

try:
    from cfe_lib.cfe_tariffs import get_tariffs_for_period_start, TariffCache
except Exception:
    get_tariffs_for_period_start = None
    TariffCache = None

try:
    from cfe_lib.nasa_power_hours import compute_period_hours_and_solar_hours
except Exception:
    compute_period_hours_and_solar_hours = None

try:
    from cfe_lib.geo_utils import geocode_address
except Exception:
    geocode_address = None

# Bundled template path (inside Docker image)
BUNDLED_TEMPLATE = Path(__file__).resolve().parent.parent / "templates" / "CONSUMO.xlsx"


# ─── Date parsing (Spanish months) ───────────────────────────

MONTH_MAP = {
    'ENE': 'JAN', 'FEB': 'FEB', 'MAR': 'MAR', 'ABR': 'APR',
    'MAY': 'MAY', 'JUN': 'JUN', 'JUL': 'JUL', 'AGO': 'AUG',
    'SEP': 'SEP', 'OCT': 'OCT', 'NOV': 'NOV', 'DIC': 'DEC',
}


def _convert_spanish_date(date_str: str) -> date:
    parts = date_str.strip().split()
    if len(parts) == 3:
        day, mon, year = parts
        mon_eng = MONTH_MAP.get(mon.upper(), mon)
        if len(year) == 2:
            year = '20' + year
        return datetime.strptime(f"{day} {mon_eng} {year}", "%d %b %Y").date()
    raise ValueError(f"Date format not recognized: {date_str}")


def _extract_period_fechas(text: str):
    period_re = re.compile(r'PERIODO FACTURADO:?\s*([\dA-Z ]+)-([\dA-Z ]+)', re.IGNORECASE)
    match = period_re.search(text)
    if match:
        try:
            ini = _convert_spanish_date(match.group(1).strip())
            fin = _convert_spanish_date(match.group(2).strip())
            return ini, fin
        except Exception:
            pass

    date_re = re.compile(r'(\d{1,2} [A-Z]{3} \d{2})-(\d{1,2} [A-Z]{3} \d{2})')
    match = date_re.search(text)
    if match:
        try:
            return _convert_spanish_date(match.group(1)), _convert_spanish_date(match.group(2))
        except Exception:
            pass

    return None, None


# ─── Address extraction ──────────────────────────────────────

def _extract_address_from_lines(lines: List[str]) -> Optional[str]:
    if not lines:
        return None

    def _clean_line(s):
        s = (s or "").strip()
        if not s:
            return ""
        if "$" in s:
            s = s.split("$", 1)[0].strip()
        s = re.sub(r"\(.*$", "", s).strip()
        s = re.sub(r"TOTAL\s+A\s+PAGAR.*$", "", s, flags=re.IGNORECASE).strip()
        s = re.sub(r"\s+", " ", s).strip()
        return s

    # 1) Explicit label
    label_re = re.compile(
        r"(DOMICILIO(?:\sFISCAL)?|DOMICILIO/SERVICIO|DIRECCI[ÓO]N)\s*:?\s*(.*)$",
        re.IGNORECASE,
    )
    for i, line in enumerate(lines):
        m = label_re.search(line or "")
        if not m:
            continue
        tail = _clean_line(m.group(2) or "")
        parts = [tail] if tail else []
        for j in range(1, 6):
            if i + j >= len(lines):
                break
            nxt_raw = lines[i + j] or ""
            nxt = _clean_line(nxt_raw)
            if not nxt:
                continue
            if re.search(
                r"NO\.?\s*DE\s*SERVICIO|PERIODO\s*FACTURADO|RMU:|CUENTA:|TARIFA",
                nxt_raw,
                re.IGNORECASE,
            ):
                break
            if re.match(r"^[A-ZÁÉÍÓÚÑ ]{6,}:\s*$", nxt_raw.strip()):
                break
            parts.append(nxt)
        cand = re.sub(r"\s+", " ", " ".join([p for p in parts if p])).strip()
        if cand and len(cand) >= 10 and "paseo de la reforma" not in cand.lower():
            return cand

    # 2) Block between TOTAL A PAGAR and NO. DE SERVICIO
    idx_total = None
    for i, line in enumerate(lines):
        if "TOTAL A PAGAR" in (line or "").upper():
            idx_total = i
            break
    idx_serv = None
    for i, line in enumerate(lines):
        if re.search(r"NO\.?\s*DE\s*SERVICIO", line or "", re.IGNORECASE):
            idx_serv = i
            break

    if idx_total is not None and idx_serv is not None and idx_serv > idx_total:
        cand_lines = []
        for raw in lines[idx_total + 1 : idx_serv]:
            raw = raw or ""
            if re.search(r"\bNETMET\b", raw, re.IGNORECASE):
                continue
            if re.search(r"\b(peso|pesos|m\.n\.)\b", raw, re.IGNORECASE):
                continue
            cleaned = _clean_line(raw)
            if not cleaned:
                continue
            if re.fullmatch(r"[\d\.,]+", cleaned):
                continue
            cand_lines.append(cleaned)
        cand = re.sub(r"\s+", " ", " ".join(cand_lines)).strip()
        if cand and len(cand) >= 10 and "paseo de la reforma" not in cand.lower():
            return cand

    # 3) Weak fallback: find a block that looks like street + colony + state/city
    joined = " ".join([str(x) for x in lines])
    m2 = re.search(
        r"(CALLE|COL\.|COLONIA|AV\.|AVENIDA)\s+(.{20,120})",
        joined,
        flags=re.IGNORECASE,
    )
    if m2:
        cand = (m2.group(0) or "").strip()
        cand = re.sub(r"\s+", " ", cand).strip()
        if cand and "paseo de la reforma" not in cand.lower() and len(cand) >= 15:
            return cand

    return None


def _clean_address_for_geocode(addr: str) -> str:
    if not addr:
        return ""
    a = re.split(r"\bR\.?F\.?C\.?\b", addr, flags=re.IGNORECASE)[0]
    a = re.sub(r"\bRFC[:\s].*$", "", a, flags=re.IGNORECASE)
    a = re.sub(r"\bCFE\d+\w*\b", "", a, flags=re.IGNORECASE)
    a = re.sub(r"\s+", " ", a).strip()
    # Common accent-stripped replacements from PDF extraction
    a = a.replace("Cdigo Postal", "Codigo Postal")
    a = a.replace("Ciudad de Mxico", "Ciudad de Mexico")
    a = a.replace("Alcalda", "Alcaldia")
    a = a.replace("Jurez", "Juarez")
    a = a.replace("Cuauhtmoc", "Cuauhtemoc")
    if "mexico" not in a.lower():
        a = a + ", Mexico"
    return a


# ─── Excel template columns ─────────────────────────────────

EXCEL_COLUMNS = [
    'PERIODO', 'PERIODO', 'DÍAS', 'kWh BASE', 'kWh INTERMEDIA', 'kWh PUNTA',
    'kWh total', 'kW BASE', 'kW INTERMEDIA', 'kW PUNTA', 'kW MÁX',
    'POTENCIA CALCULADA', 'kVARh', 'FP',
    'CARGO FIJO ($/MES)', 'kWh BASE ($/kWh)', 'kWh INTERMEDIA ($/kWh)',
    'kWh PUNTA ($/kWh)', 'DISTRIBUCIÓN ($/kW)', 'CAPACIDAD ($/kW)',
    'POTENCIA PARA EL CARGO POR CAPACIDAD', 'POTENCIA PARA EL CARGO POR DISTRIBUCIÓN',
    '$ BASE', '$ INTERMEDIO', '$ PUNTA', '$ DISTRIBUCIÓN', '$ CAPACIDAD',
    '$ ENERGÍA', 'FP', '$ FP', 'SUBTOTAL', 'IVA 16%', 'DAP', 'TOTAL',
    'TOTAL A PAGAR (MXN)', 'CONTRAPRESTACIÓN',
    'HORAS BASE', 'HORAS INTERMEDIA', 'HORAS PUNTA ',
    'HORAS BASE SOLARES', 'HORAS INTERMEDIA SOLARES', 'TOTAL DE HORAS SOLARES',
    'CONSUMO BASE EN HORAS SOLARES', 'CONSUMO INTERMEDIO EN HORAS SOLARES',
    'CONSUMO EN HORAS SOLARES (kWh)',
    '%CONSUMO EN BASE RESPECTO AL TOTAL EN HORAS SOLARES',
    '%CONSUMO EN INTERMEDIO RESPECTO AL TOTAL EN HORAS SOLARES',
    '%CONSUMO EN HORAS SOLARES RESPECTO AL CONSUMO TOTAL',
    'HSP (NASA)', 'EFICIENCIA', 'POTENCIA DEL MÓDULO (kW)',
    'PSFV 100%', '#MÓDULOS 100%', 'PSFV PERIODO INTERMEDIO',
    '# MÓDULOS PERIODO INTERMEDIO', 'PSFV HORAS SOLARES', '#MÓDULOS HORAS SOLARES',
    'SISTEMA PROPUESTO (kW)', 'GENERACIÓN CON SFV PROPUESTO (kWh)',
    'GENERACIÓN CON SFV EN BASE', 'GENERACIÓN CON SFV EN INTERMEDIA',
    'EXCEDENTES/FALTANTES PROYECTADOS',
    'NUEVO CONSUMO ESPERADO EN BASE (kWh)', 'NUEVO CONSUMO ESPERADO EN INTERMEDIO (kWh)',
    'NUEVO CONSUMO ESPERADO (kWh)', 'PORCENTAJE DE ABATIMIENTO DE CONSUMO ENERGÉTICO',
    '=BK2', '=BL2', '=F2', 'kWh TOTAL', 'POTENCIA CALCULADA (kW)',
    '=N2', '=O2', '=P2', '=Q2', '=R2', '=S2', '=T2',
    '$ BASE', '$ INTERMEDIO', '$ PUNTA', '$ DISTRIBUCIÓN', '$ CAPACIDAD',
    '$ ENERGÍA', '=AC2', '=AD2', 'SUBTOTAL', 'IVA 16%', 'DAP',
    'TOTAL A PAGAR CON SISTEMA FOTOVOLTAICO', 'AHORRO', '', '',
]

FORMULA_COLUMNS = {
    3: "=B{row}-A{row}", 7: "=SUM(D{row}:F{row})",
    12: "=G{row}/(24*C{row}*0.57)", 21: "=MIN(J{row}:L{row})",
    22: "=MIN(K{row}:L{row})", 23: "=D{row}*P{row}",
    24: "=E{row}*Q{row}", 25: "=F{row}*R{row}",
    26: "=S{row}*V{row}", 27: "=T{row}*U{row}",
    28: "=SUMA(W{row}:AA{row})",
    29: "=SI(N{row}<95,REDONDEAR((3/5)*((95/N{row})-1),3), REDONDEAR((1/4)*(1-(90/N{row})),3))",
    30: "=(AB{row}+O{row})*AC{row}", 31: "=SUMA(O{row},AB{row},AD{row})",
    32: "=AE{row}*0.16", 33: "=AE{row}*0.08", 34: "=AE{row}+AF{row}+AG{row}",
    42: "=SUMA(AN{row}:AO{row})", 43: "=D{row}*AN{row}/AK{row}",
    44: "=E{row}*AO{row}/AL{row}", 45: "=SUMA(AQ{row},AR{row})",
    46: "=AQ{row}/AS{row}", 47: "=AR{row}/AS{row}", 48: "=AS{row}/G{row}",
    52: "=G{row}/(C{row}*AW{row}*AX{row})", 53: "=REDONDEAR.MAS(AZ{row}/AY{row},0)",
    54: "=E{row}/(C{row}*AW{row}*AX{row})", 55: "=REDONDEAR.MAS(BB{row}/AY{row},0)",
    56: "=AS{row}/(C{row}*AW{row}*AX{row})", 57: "=REDONDEAR.MAS(BD{row}/AY{row},0)",
    58: "=4560*AY{row}", 59: "=BF{row}*C{row}*AW{row}*AX{row}",
    60: "=BG{row}*AT{row}", 61: "=BG{row}*AU{row}",
    62: "=REDONDEAR.MAS(BG{row}-AS{row},0)", 63: "=D{row}-BH{row}",
    64: "=E{row}-BI{row}",
    65: "=SI(BJ{row}<0,G{row}-BG{row},G{row}-BG{row}+BJ{row})",
    66: "=SI(BJ{row}<0,BG{row}/G{row},(BG{row}-BJ{row})/G{row})",
    67: "=BK{row}", 68: "=BL{row}", 69: "=F{row}",
    70: "=SUMA(BO{row}:BQ{row})", 71: "=BR{row}/(24*C{row}*0.57)",
    72: "=N{row}", 73: "=O{row}", 74: "=P{row}", 75: "=Q{row}",
    76: "=R{row}", 77: "=S{row}", 78: "=T{row}",
    79: "=BO{row}*BV{row}", 80: "=BP{row}*BW{row}", 81: "=BQ{row}*BX{row}",
    82: "=BS{row}*BY{row}", 83: "=BS{row}*BZ{row}",
    84: "=SUMA(CA{row}:CC{row})", 85: "=AC{row}", 86: "=AD{row}",
    87: "=SUMA(BU{row},CF{row},CH{row})", 88: "=CI{row}*0.16",
    89: "=CI{row}*0.08", 90: "=SUMA(CI{row},CJ{row},CK{row})",
    91: "=AI{row}-CL{row}",
}


# ─── PDF field extraction ────────────────────────────────────

def _extract_pdf_fields(pdf_path: str) -> dict:
    with pdfplumber.open(pdf_path) as pdf:
        lines = []
        for page in pdf.pages:
            text = page.extract_text() or ""
            lines.extend(text.splitlines())
        full_text = " ".join(lines)
        # Get page1 text while still open
        text_page1 = pdf.pages[0].extract_text() or ""

    periodo_start, periodo_end = _extract_period_fechas(text_page1)
    dias = None
    if periodo_start and periodo_end:
        try:
            dias = (periodo_end - periodo_start).days
        except Exception:
            pass

    # TOTAL A PAGAR
    total_pagar = None
    for idx, line in enumerate(lines):
        if "TOTAL A PAGAR" in line.upper():
            for offset in range(1, 6):
                if idx + offset < len(lines):
                    next_line = lines[idx + offset].strip()
                    match = re.search(r"\$?\s*([\d,\.]+)$", next_line)
                    if match:
                        val = match.group(1).replace(",", "").replace(" ", "")
                        try:
                            total_pagar = float(val) if '.' in val else int(val)
                        except Exception:
                            continue
                        break
            if total_pagar is not None:
                break

    def get_number(label):
        match = re.search(label + r"[ :]*([0-9\.,]+)", full_text, re.IGNORECASE)
        val = match.group(1).replace(",", "") if match else None
        try:
            return float(val) if '.' in val else int(val)
        except Exception:
            return None

    kwh_base = get_number("kWh base")
    kwh_inter = get_number("kWh intermedia")
    kwh_punta = get_number("kWh punta")
    kwh_total = None
    if all(x is not None for x in [kwh_base, kwh_inter, kwh_punta]):
        kwh_total = kwh_base + kwh_inter + kwh_punta

    kw_base = get_number("kW base")
    kw_inter = get_number("kW intermedia")
    kw_punta = get_number("kW punta")
    kw_max = get_number("KWMax")
    kvarh = get_number("kVArh")
    fp = get_number("Factor de potencia %")

    contraprestacion = (
        "NET MET"
        if ("NETMET" in full_text.upper() or "NET MET" in full_text.upper())
        else ""
    )

    row = [""] * len(EXCEL_COLUMNS)
    row[0] = periodo_start
    row[1] = periodo_end
    row[2] = dias
    row[3] = kwh_base
    row[4] = kwh_inter
    row[5] = kwh_punta
    row[6] = kwh_total
    row[7] = kw_base
    row[8] = kw_inter
    row[9] = kw_punta
    row[10] = kw_max
    row[12] = kvarh
    row[13] = fp
    row[34] = total_pagar
    row[35] = contraprestacion

    address = _extract_address_from_lines(lines)

    return {
        "row": row,
        "period_start": periodo_start,
        "period_end": periodo_end,
        "address": address,
        "pdf_path": pdf_path,
    }


# ─── Excel output ────────────────────────────────────────────

def _fill_sheet1_explicit_formulas(wb, rows):
    ws = wb.worksheets[0]
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)
    for row_idx, data_row in enumerate(rows, start=3):
        for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
            if col_idx in FORMULA_COLUMNS and FORMULA_COLUMNS[col_idx]:
                ws.cell(row=row_idx, column=col_idx).value = FORMULA_COLUMNS[col_idx].format(row=row_idx)
            else:
                ws.cell(row=row_idx, column=col_idx).value = (
                    data_row[col_idx - 1] if (col_idx - 1) < len(data_row) else ""
                )


def _patch_range_formula(formula, src_start, src_end, dest_start, dest_end):
    pattern = rf"(?P<col>[A-Z]{{1,3}}){src_start}:(?P=col){src_end}"
    return re.sub(
        pattern,
        lambda m: f"{m.group('col')}{dest_start}:{m.group('col')}{dest_end}",
        formula,
    )


def _shift_formula_row_refs(formula, src_base_row, dest_base_row, src_rows):
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula
    delta = dest_base_row - src_base_row
    if delta == 0:
        return formula

    def repl(m):
        col = m.group(1)
        row = int(m.group(2))
        if src_base_row <= row <= (src_base_row + src_rows - 1):
            return f"{col}{row + delta}"
        return m.group(0)

    return re.sub(r"(\$?[A-Z]{1,3}\$?)(\d+)", repl, formula)


def _append_block_ranges(wb, pdf_count, template_ws):
    ws = wb.worksheets[0]
    data_start_row = 3
    data_end_row = data_start_row + pdf_count - 1
    if pdf_count <= 0:
        data_end_row = data_start_row

    dest_start = data_end_row + 1

    # Block 1: AY24:BN27 (4 rows)
    for src_offset, src_row in enumerate(range(24, 28)):
        dest_row = dest_start + src_offset
        for col_idx in range(
            openpyxl.utils.column_index_from_string('AY'),
            openpyxl.utils.column_index_from_string('BN') + 1,
        ):
            src_cell = template_ws.cell(row=src_row, column=col_idx)
            dest_cell = ws.cell(row=dest_row, column=col_idx)
            if isinstance(src_cell.value, str) and src_cell.value.startswith("="):
                formula = _patch_range_formula(src_cell.value, 3, 23, 3, data_end_row)
                dest_cell.value = formula
            else:
                dest_cell.value = src_cell.value
            if src_cell.has_style:
                dest_cell._style = src_cell._style

    # Block 2: CL24:CN26 (3 rows)
    for src_offset, src_row in enumerate(range(24, 27)):
        dest_row = dest_start + src_offset
        for col_idx in range(
            openpyxl.utils.column_index_from_string('CL'),
            openpyxl.utils.column_index_from_string('CN') + 1,
        ):
            src_cell = template_ws.cell(row=src_row, column=col_idx)
            dest_cell = ws.cell(row=dest_row, column=col_idx)
            val = src_cell.value
            if isinstance(val, str) and val.startswith("="):
                val = _patch_range_formula(val, 3, 23, 3, data_end_row)
                val = _shift_formula_row_refs(
                    val, src_base_row=24, dest_base_row=dest_start, src_rows=3
                )
                dest_cell.value = val
            else:
                dest_cell.value = val
            if src_cell.has_style:
                dest_cell._style = src_cell._style


def _normalize_rows(rows_or_meta):
    """
    Accepts:
      - list[list|tuple] of rows, or
      - list[(row, address)] where row is list/tuple, or
      - list[dict] with key 'row'
    Returns list[list] rows.
    """
    if not rows_or_meta:
        return []

    rows = []
    for idx, item in enumerate(rows_or_meta):
        if isinstance(item, dict) and "row" in item and isinstance(item["row"], (list, tuple)):
            rows.append(list(item["row"]))
            continue
        if isinstance(item, (list, tuple)) and len(item) == 2 and isinstance(item[0], (list, tuple)):
            rows.append(list(item[0]))
            continue
        if isinstance(item, (list, tuple)):
            rows.append(list(item))
            continue
        if isinstance(item, dict):
            found = None
            for v in item.values():
                if isinstance(v, (list, tuple)):
                    found = v
                    break
            if found is not None:
                rows.append(list(found))
                continue
        print(f"[cfe] Warning: skipping unusable row at index {idx}: {type(item)}")

    eff_idx = openpyxl.utils.column_index_from_string("AX") - 1
    out = []
    for r in rows:
        lr = list(r)
        if len(lr) < len(EXCEL_COLUMNS):
            lr.extend([""] * (len(EXCEL_COLUMNS) - len(lr)))
        # Default EFICIENCIA (AX) to 0.85 if empty
        if eff_idx < len(lr):
            v = lr[eff_idx]
            if v is None or (isinstance(v, str) and v.strip() == ""):
                lr[eff_idx] = 0.85
        out.append(lr)
    return out


def _create_output_excel(rows_data, output_path: str, template_path: str):
    wb = openpyxl.load_workbook(template_path)
    template_wb = openpyxl.load_workbook(template_path, data_only=False)
    template_ws = template_wb.worksheets[0]

    rows = _normalize_rows(rows_data)
    _fill_sheet1_explicit_formulas(wb, rows)
    _append_block_ranges(wb, len(rows), template_ws)

    wb.active = 0
    wb.properties.calcMode = "auto"
    wb.save(output_path)


# ─── Default schedule (GDMTH fallback) ──────────────────────

def _default_schedule_for_local_date(local_date):
    return [
        {"start": "00:00", "end": "06:00", "category": "base"},
        {"start": "06:00", "end": "18:00", "category": "intermedia"},
        {"start": "18:00", "end": "22:00", "category": "punta"},
        {"start": "22:00", "end": "24:00", "category": "base"},
    ]


# ─── Helpers ─────────────────────────────────────────────────

def _norm(s):
    """Normalize string for key matching (strip accents, lower, collapse spaces)."""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s).strip().lower()


def _parse_float_like(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(" ", "").replace(",", "")
    s = re.sub(r"[^\d\.\-]", "", s)
    if not s or s in (".", "-"):
        return None
    try:
        return float(s)
    except Exception:
        return None


# ─── Enrichment (CFE tariffs + NASA) ────────────────────────

def _enrich_row(meta: dict, col: dict, tariff_cache) -> list:
    """Enrich a single row with CFE tariffs and NASA solar hours."""
    row = meta["row"]
    ps = meta.get("period_start")
    pe = meta.get("period_end")
    address = meta.get("address") or ""

    # ── CFE tariffs ──
    if get_tariffs_for_period_start is not None and ps is not None:
        try:
            tariff_date = pe or ps
            tariffs = get_tariffs_for_period_start(
                tariff_date, cache=tariff_cache, location_text=address
            )

            if isinstance(tariffs, dict):
                tariffs_norm = {_norm(k): v for k, v in tariffs.items()}

                # Candidate key groups (possible key names returned by cfe_tariffs)
                candidates = {
                    "cargo_fijo": [
                        "cargo_fijo", "fijo", "cargo fijo",
                        "cargo_fijo($/mes)", "fijo($/mes)",
                    ],
                    "kwh_base": [
                        "kwh_base", "base", "energia_base",
                        "kwh base", "base ($/kwh)", "base($/kwh)",
                    ],
                    "kwh_intermedia": [
                        "kwh_intermedia", "intermedia", "intermedio",
                        "energia_intermedia", "kwh intermedia",
                        "intermedia ($/kwh)", "intermedio ($/kwh)",
                    ],
                    "kwh_punta": [
                        "kwh_punta", "punta", "energia_punta",
                        "kwh punta", "punta ($/kwh)",
                    ],
                    "distribucion": [
                        "distribucion", "distribucion ($/kw)",
                        "distribucion ($/kW)", "distribucion($/kw)",
                    ],
                    "capacidad": [
                        "capacidad", "capacidad ($/kw)",
                        "capacidad ($/kW)", "capacidad($/kw)",
                    ],
                }

                def pick(cand_list):
                    # Try exact normalized key match
                    for c in cand_list:
                        if _norm(c) in tariffs_norm and tariffs_norm[_norm(c)] not in (None, ""):
                            return _parse_float_like(tariffs_norm[_norm(c)])
                    # Try substring match
                    for k, v in tariffs_norm.items():
                        for c in cand_list:
                            if c in k and v not in (None, ""):
                                return _parse_float_like(v)
                    # Try original keys
                    for orig_k, orig_v in tariffs.items():
                        if _norm(orig_k) in [_norm(c) for c in cand_list] and orig_v not in (None, ""):
                            return _parse_float_like(orig_v)
                    return None

                def setv(header, value):
                    if value is not None and header in col:
                        row[col[header]] = value

                setv("CARGO FIJO ($/MES)", pick(candidates["cargo_fijo"]))

                # kWh rate headers may have small variations
                base_rate = pick(candidates["kwh_base"])
                inter_rate = pick(candidates["kwh_intermedia"])
                punta_rate = pick(candidates["kwh_punta"])
                distrib = pick(candidates["distribucion"])
                capacidad_val = pick(candidates["capacidad"])

                for h in ["kWh BASE ($/kWh)", "kWh BASE ($/kwh)", "kWh BASE"]:
                    setv(h, base_rate)
                for h in [
                    "kWh INTERMEDIA ($/kWh)", "kWh INTERMEDIA ($/kwh)",
                    "kWh INTERMEDIO ($/kWh)", "kWh INTERMEDIO ($/kwh)",
                ]:
                    setv(h, inter_rate)
                for h in ["kWh PUNTA ($/kWh)", "kWh PUNTA ($/kwh)", "kWh PUNTA"]:
                    setv(h, punta_rate)

                setv("DISTRIBUCIÓN ($/kW)", distrib)
                setv("CAPACIDAD ($/kW)", capacidad_val)

        except Exception as e:
            print(f"[cfe] Warning: no se pudieron obtener tarifas CFE para {ps}: {e}")

    # ── NASA POWER hours + solar hours + HSP ──
    if (
        compute_period_hours_and_solar_hours is not None
        and geocode_address is not None
        and ps is not None
        and pe is not None
    ):
        query = _clean_address_for_geocode(address)
        result = geocode_address(query)
        lat, lon = result if result else (None, None)

        if lat is not None and lon is not None:
            try:
                start_dt = datetime(ps.year, ps.month, ps.day, 0, 0, 0)
                end_dt = datetime(pe.year, pe.month, pe.day, 0, 0, 0) + timedelta(days=1)

                res = compute_period_hours_and_solar_hours(
                    start_dt,
                    end_dt,
                    float(lat),
                    float(lon),
                    schedule_for_local_date=_default_schedule_for_local_date,
                    irradiance_threshold_wm2=20.0,
                )

                if "HORAS BASE" in col:
                    row[col["HORAS BASE"]] = res.total_base_hours
                if "HORAS INTERMEDIA" in col:
                    row[col["HORAS INTERMEDIA"]] = res.total_intermedia_hours
                if "HORAS PUNTA " in col:
                    row[col["HORAS PUNTA "]] = res.total_punta_hours
                if "HORAS BASE SOLARES" in col:
                    row[col["HORAS BASE SOLARES"]] = res.solar_base_hours
                if "HORAS INTERMEDIA SOLARES" in col:
                    row[col["HORAS INTERMEDIA SOLARES"]] = res.solar_intermedia_hours
                if "TOTAL DE HORAS SOLARES" in col:
                    row[col["TOTAL DE HORAS SOLARES"]] = (
                        (res.solar_base_hours or 0) + (res.solar_intermedia_hours or 0)
                    )
                if "HSP (NASA)" in col:
                    row[col["HSP (NASA)"]] = res.hsp_nasa
            except Exception as e:
                print(f"[cfe] Warning: no se pudieron calcular horas NASA para '{address}': {e}")
        else:
            print(f"[cfe] Warning: geocoding failed for '{address}' (query: '{query}')")

    return row


# ─── Worker processor entry point ────────────────────────────

def process(ctx: JobContext) -> str:
    """Process CFE data extraction job. Returns relative output path."""

    ctx.report_progress(5, "Identificando archivos...")

    pdfs = ctx.input_files(".pdf")
    if not pdfs:
        raise ValueError("No se encontraron archivos PDF de recibos CFE")

    # Template: the CONSUMO.xlsx template
    template_path = None

    # 1) Check if provided via job config
    if ctx.template_abs and ctx.template_abs.exists():
        template_path = ctx.template_abs
    else:
        # 2) Check for xlsx in inputs
        xlsx_inputs = ctx.input_files(".xlsx")
        if xlsx_inputs:
            template_path = xlsx_inputs[0]

    # 3) Fall back to bundled template
    if not template_path and BUNDLED_TEMPLATE.exists():
        template_path = BUNDLED_TEMPLATE

    if not template_path:
        raise ValueError(
            "Se requiere la plantilla CONSUMO.xlsx. "
            "Súbela como template al crear el job."
        )

    ctx.report_progress(10, f"Extrayendo datos de {len(pdfs)} recibo(s) CFE...")

    # Build column index
    col = {str(c).strip(): i for i, c in enumerate(EXCEL_COLUMNS)}

    # Initialize tariff cache
    tariff_cache = TariffCache() if TariffCache is not None else None

    enriched_rows = []
    for i, pdf in enumerate(pdfs):
        pct = 10 + int(70 * (i + 1) / len(pdfs))
        ctx.report_progress(pct, f"Procesando {pdf.name} ({i + 1}/{len(pdfs)})...")

        try:
            meta = _extract_pdf_fields(str(pdf))
            row = _enrich_row(meta, col, tariff_cache)
            enriched_rows.append(row)
        except Exception as e:
            print(f"[cfe] Error processing {pdf.name}: {e}")
            traceback.print_exc()

    if not enriched_rows:
        raise ValueError("No se pudo extraer datos de ningún PDF")

    ctx.report_progress(85, "Generando Excel de salida...")

    output_file = ctx.output_path("consumo_cfe.xlsx")
    _create_output_excel(enriched_rows, str(output_file), str(template_path))

    ctx.report_progress(95, f"Procesados {len(enriched_rows)} recibo(s) exitosamente")
    return ctx.output_rel("consumo_cfe.xlsx")
