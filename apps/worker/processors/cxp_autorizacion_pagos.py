from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from math import fsum
from pathlib import Path
from typing import Iterable

import openpyxl
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import TABLOID, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .base import JobContext

APP_KEY = "cuentas_por_pagar_autorizacion_pagos"
PAYMENT_ORDER = {
    "CHEQUE NF": 0,
    "COMPROBACION": 1,
    "EFECTIVO": 2,
    "TARJETA": 3,
    "TRANSFERENCIA": 4,
}
WEEKDAY_LABELS = {
    0: "LUNES",
    1: "MARTES",
    2: "MIÉRCOLES",
    3: "JUEVES",
    4: "VIERNES",
    5: "SÁBADO",
    6: "DOMINGO",
}
SIGNATURES = [
    ("ELABORÓ", "Juan Carlos Rodriguez Silva", "ERA Cuentas por Pagar"),
    ("REVISÓ", "C.P. Jesús Salvador Chavez Martinez", "Coordinación Gerencia Administrativa"),
    ("AUTORIZÓ", "C.P. Ana Laura Jiménez Lara", "Dirección Grupo Loros"),
]


@dataclass
class PaymentEntry:
    razon: str
    unidad: str
    familia: str
    proyecto: str
    proveedor: str
    concepto: str
    detalle: str
    forma_pago: str
    importe: float
    fecha_pago: date | None


@dataclass
class WorkbookSnapshot:
    week_number: int | None
    start_date: date | None
    end_date: date | None
    target_weekday: int
    target_label: str
    entries: list[PaymentEntry]
    unit_summary: list[tuple[str, float]]
    concept_summary: list[tuple[str, float]]
    summary_total: float | None = None

    @property
    def total(self) -> float:
        if self.summary_total is not None:
            return round(self.summary_total, 2)
        return round(fsum(entry.importe for entry in self.entries), 2)


def _as_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _as_float(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        text = str(value).strip().replace("$", "").replace(",", "")
        return float(text) if text else 0.0
    except Exception:
        return 0.0


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def _format_money(value: float) -> str:
    return f"${value:,.2f}"


def _find_sheet(wb: openpyxl.Workbook, prefix: str) -> openpyxl.worksheet.worksheet.Worksheet:
    for name in wb.sheetnames:
        if name.strip().upper().startswith(prefix.strip().upper()):
            return wb[name]
    raise ValueError(f"No se encontró la hoja requerida: {prefix}")


def _extract_entries(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[PaymentEntry]:
    entries: list[PaymentEntry] = []
    for row_idx in range(12, ws.max_row + 1):
        unidad = _clean_text(ws.cell(row_idx, 2).value)
        forma_pago = _clean_text(ws.cell(row_idx, 8).value)
        importe = _as_float(ws.cell(row_idx, 9).value)
        if not unidad or not forma_pago or importe <= 0:
            continue
        entries.append(
            PaymentEntry(
                razon=_clean_text(ws.cell(row_idx, 1).value),
                unidad=unidad,
                familia=_clean_text(ws.cell(row_idx, 3).value),
                proyecto=_clean_text(ws.cell(row_idx, 4).value),
                proveedor=_clean_text(ws.cell(row_idx, 5).value),
                concepto=_clean_text(ws.cell(row_idx, 6).value),
                detalle=_clean_text(ws.cell(row_idx, 7).value),
                forma_pago=forma_pago,
                importe=round(importe, 2),
                fecha_pago=_as_date(ws.cell(row_idx, 17).value),
            )
        )
    return entries


def _resolve_target_weekday(
    entries: Iterable[PaymentEntry],
    end_date: date | None,
) -> int:
    if end_date is not None:
        return end_date.weekday()

    weekday_totals: dict[int, float] = defaultdict(float)
    for entry in entries:
        if entry.fecha_pago is None:
            continue
        weekday_totals[entry.fecha_pago.weekday()] += entry.importe

    if not weekday_totals:
        return 4

    return max(weekday_totals.items(), key=lambda item: (item[1], -item[0]))[0]


def _filter_entries_for_weekday(entries: Iterable[PaymentEntry], weekday: int) -> list[PaymentEntry]:
    return [entry for entry in entries if entry.fecha_pago is not None and entry.fecha_pago.weekday() == weekday]


def _extract_unit_summary(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[tuple[str, float]]:
    rows: list[tuple[str, float]] = []
    for row_idx in range(10, ws.max_row + 1):
        label = _clean_text(ws.cell(row_idx, 1).value)
        value = _as_float(ws.cell(row_idx, 2).value)
        if not label:
            if rows:
                break
            continue
        if value <= 0 and label != "TOTAL":
            continue
        rows.append((label, round(value, 2)))
        if label == "TOTAL":
            break
    return rows


def _find_day_column(ws: openpyxl.worksheet.worksheet.Worksheet, weekday_label: str) -> int:
    for col_idx in range(1, ws.max_column + 1):
        header = _clean_text(ws.cell(4, col_idx).value).upper()
        if weekday_label in header:
            return col_idx
    return 8


def _extract_concept_summary(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    weekday_label: str,
) -> list[tuple[str, float]]:
    day_col = _find_day_column(ws, weekday_label)
    rows: list[tuple[str, float]] = []
    for row_idx in range(5, ws.max_row + 1):
        label = _clean_text(ws.cell(row_idx, 1).value)
        value = _as_float(ws.cell(row_idx, day_col).value)
        if not label:
            if rows:
                break
            continue
        if value <= 0:
            continue
        rows.append((label, round(value, 2)))
    return rows


def _load_snapshot(workbook_path: Path) -> WorkbookSnapshot:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    detail_ws = _find_sheet(wb, "BD.CXP")
    summary_unit_ws = _find_sheet(wb, "RESUMEN POR U.N")
    summary_concept_ws = _find_sheet(wb, "DESLGOSE POR TIPO DE GAST")

    week_number_raw = detail_ws["B4"].value
    week_number = int(week_number_raw) if week_number_raw not in (None, "") else None
    start_date = _as_date(detail_ws["B5"].value)
    end_date = _as_date(detail_ws["B6"].value)

    entries = _extract_entries(detail_ws)
    target_weekday = _resolve_target_weekday(entries, end_date)
    target_label = WEEKDAY_LABELS.get(target_weekday, "VIERNES")
    filtered_entries = _filter_entries_for_weekday(entries, target_weekday)
    if not filtered_entries:
        raise ValueError("No se encontraron partidas para el día seleccionado en FECHA REAL DE PAGO.")

    unit_summary = _extract_unit_summary(summary_unit_ws)
    concept_summary = _extract_concept_summary(summary_concept_ws, target_label)
    summary_total = None
    for label, value in unit_summary:
        if label.strip().upper() == "TOTAL":
            summary_total = value
            break

    return WorkbookSnapshot(
        week_number=week_number,
        start_date=start_date,
        end_date=end_date,
        target_weekday=target_weekday,
        target_label=target_label,
        entries=filtered_entries,
        unit_summary=unit_summary,
        concept_summary=concept_summary,
        summary_total=summary_total,
    )


def _build_styles() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "cxp-title",
            parent=base["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=14,
            leading=17,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#0f172a"),
            spaceAfter=4,
        ),
        "meta": ParagraphStyle(
            "cxp-meta",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#334155"),
        ),
        "section": ParagraphStyle(
            "cxp-section",
            parent=base["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=12,
            textColor=colors.HexColor("#0f172a"),
            spaceBefore=6,
            spaceAfter=4,
        ),
        "unit": ParagraphStyle(
            "cxp-unit",
            parent=base["Heading3"],
            fontName="Helvetica-Bold",
            fontSize=9,
            leading=11,
            textColor=colors.HexColor("#1d4ed8"),
            spaceBefore=8,
            spaceAfter=4,
        ),
        "payment": ParagraphStyle(
            "cxp-payment",
            parent=base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#0f172a"),
            spaceBefore=4,
            spaceAfter=2,
        ),
        "cell": ParagraphStyle(
            "cxp-cell",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=7.5,
            leading=9,
            textColor=colors.HexColor("#0f172a"),
            alignment=TA_LEFT,
        ),
        "cell-right": ParagraphStyle(
            "cxp-cell-right",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=7.5,
            leading=9,
            textColor=colors.HexColor("#0f172a"),
            alignment=TA_RIGHT,
        ),
        "signature": ParagraphStyle(
            "cxp-signature",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#0f172a"),
        ),
    }


def _paragraph(text: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(text.replace("&", "&amp;"), style)


def _summary_table(
    title: str,
    rows: list[tuple[str, float]],
    styles: dict[str, ParagraphStyle],
) -> Table:
    data = [[_paragraph(f"<b>{title}</b>", styles["cell"]), ""]]
    for label, value in rows:
        data.append([
            _paragraph(label, styles["cell"]),
            _paragraph(_format_money(value), styles["cell-right"]),
        ])

    table = Table(data, colWidths=[74 * mm, 30 * mm])
    table.setStyle(
        TableStyle(
            [
                ("SPAN", (0, 0), (1, 0)),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("INNERGRID", (0, 1), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return table


def _detail_table(
    entries: list[PaymentEntry],
    styles: dict[str, ParagraphStyle],
) -> Table:
    data = [[
        _paragraph("<b>Razón</b>", styles["cell"]),
        _paragraph("<b>Familia</b>", styles["cell"]),
        _paragraph("<b>Proyecto</b>", styles["cell"]),
        _paragraph("<b>Proveedor</b>", styles["cell"]),
        _paragraph("<b>Concepto</b>", styles["cell"]),
        _paragraph("<b>Detalle</b>", styles["cell"]),
        _paragraph("<b>Importe</b>", styles["cell-right"]),
    ]]

    for entry in entries:
        data.append(
            [
                _paragraph(entry.razon or "-", styles["cell"]),
                _paragraph(entry.familia or "-", styles["cell"]),
                _paragraph(entry.proyecto or "-", styles["cell"]),
                _paragraph(entry.proveedor or "-", styles["cell"]),
                _paragraph(entry.concepto or "-", styles["cell"]),
                _paragraph(entry.detalle or "-", styles["cell"]),
                _paragraph(_format_money(entry.importe), styles["cell-right"]),
            ]
        )

    total = round(fsum(entry.importe for entry in entries), 2)
    data.append(
        [
            _paragraph("<b>Total</b>", styles["cell"]),
            "",
            "",
            "",
            "",
            "",
            _paragraph(f"<b>{_format_money(total)}</b>", styles["cell-right"]),
        ]
    )

    table = Table(
        data,
        repeatRows=1,
        colWidths=[22 * mm, 36 * mm, 42 * mm, 62 * mm, 28 * mm, 84 * mm, 24 * mm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f8fafc")),
                ("LINEABOVE", (0, -1), (-1, -1), 0.5, colors.HexColor("#94a3b8")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("INNERGRID", (0, 0), (-1, -2), 0.25, colors.HexColor("#e2e8f0")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return table


def _group_entries(entries: Iterable[PaymentEntry]) -> dict[str, dict[str, list[PaymentEntry]]]:
    grouped: dict[str, dict[str, list[PaymentEntry]]] = defaultdict(lambda: defaultdict(list))
    for entry in entries:
        grouped[entry.unidad][entry.forma_pago].append(entry)

    for unit_groups in grouped.values():
        for payment_entries in unit_groups.values():
            payment_entries.sort(key=lambda item: (-item.importe, item.proveedor, item.proyecto))

    return grouped


def _draw_page(canvas, doc) -> None:
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#64748b"))
    canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, 10 * mm, f"Página {doc.page}")
    canvas.restoreState()


def _build_pdf(snapshot: WorkbookSnapshot, output_path: Path) -> None:
    styles = _build_styles()
    story = []

    date_code = snapshot.end_date.strftime("%d/%m/%Y") if snapshot.end_date else "sin fecha"
    week_label = f"Semana {snapshot.week_number:02d}/{(snapshot.end_date or date.today()).year}" if snapshot.week_number else "Semana sin definir"
    range_label = "Sin rango"
    if snapshot.start_date and snapshot.end_date:
        range_label = f"Del {snapshot.start_date.strftime('%d/%m/%Y')} al {snapshot.end_date.strftime('%d/%m/%Y')}"

    story.append(_paragraph("AUTORIZACIÓN DE PAGOS", styles["title"]))
    story.append(_paragraph(f"{date_code} · {week_label} · {snapshot.target_label}", styles["meta"]))
    story.append(_paragraph(range_label, styles["meta"]))
    story.append(Spacer(1, 5 * mm))

    left_rows = snapshot.unit_summary or sorted(
        (
            (unit, round(fsum(entry.importe for entry in snapshot.entries if entry.unidad == unit), 2))
            for unit in sorted({entry.unidad for entry in snapshot.entries})
        ),
        key=lambda item: item[0],
    )
    right_rows = snapshot.concept_summary or sorted(
        (
            (concept, round(fsum(entry.importe for entry in snapshot.entries if entry.concepto == concept), 2))
            for concept in sorted({entry.concepto for entry in snapshot.entries})
        ),
        key=lambda item: item[0],
    )

    cover = Table(
        [[
            _summary_table(f"Unidad de negocio · {snapshot.target_label}", left_rows, styles),
            _summary_table(f"Concepto · {snapshot.target_label}", right_rows, styles),
        ]],
        colWidths=[110 * mm, 110 * mm],
    )
    cover.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story.append(cover)
    story.append(Spacer(1, 5 * mm))
    story.append(_paragraph("Detalle de partidas autorizadas", styles["section"]))

    grouped = _group_entries(snapshot.entries)
    for unidad in sorted(grouped):
        story.append(_paragraph(unidad, styles["unit"]))
        payment_groups = grouped[unidad]
        for forma_pago in sorted(payment_groups, key=lambda item: (PAYMENT_ORDER.get(item, 99), item)):
            story.append(_paragraph(forma_pago, styles["payment"]))
            story.append(_detail_table(payment_groups[forma_pago], styles))
            story.append(Spacer(1, 2 * mm))

    story.append(Spacer(1, 4 * mm))
    story.append(_paragraph(f"<b>TOTAL GLOBAL: {_format_money(snapshot.total)}</b>", styles["section"]))
    story.append(Spacer(1, 8 * mm))

    signatures = []
    top = []
    body = []
    role = []
    for title, name, position in SIGNATURES:
        top.append(_paragraph(f"<b>{title}</b>", styles["signature"]))
        body.append(_paragraph(name, styles["signature"]))
        role.append(_paragraph(position, styles["signature"]))
    signatures.extend([top, ["", "", ""], body, role])

    sign_table = Table(signatures, colWidths=[58 * mm, 58 * mm, 58 * mm])
    sign_table.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("LINEABOVE", (0, 2), (-1, 2), 0.5, colors.HexColor("#64748b")),
                ("TOPPADDING", (0, 2), (-1, 2), 8),
            ]
        )
    )
    story.append(sign_table)

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(TABLOID),
        leftMargin=9 * mm,
        rightMargin=9 * mm,
        topMargin=10 * mm,
        bottomMargin=14 * mm,
    )
    doc.build(story, onFirstPage=_draw_page, onLaterPages=_draw_page)


def process(ctx: JobContext) -> str:
    xlsx_inputs = [path for path in ctx.input_files() if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}]
    if not xlsx_inputs:
        raise ValueError("Se requiere un archivo Excel de provisión para generar la autorización de pagos.")

    workbook_path = xlsx_inputs[0]
    ctx.report_progress(10, "Leyendo archivo de provisión semanal...")
    snapshot = _load_snapshot(workbook_path)
    ctx.params["target_weekday"] = snapshot.target_label
    ctx.params["detail_rows"] = len(snapshot.entries)
    ctx.params["total_amount"] = snapshot.total

    date_token = (snapshot.end_date or date.today()).strftime("%d%m%y")
    filename = f"AUTORIZACION_DE_PAGOS_{date_token}.pdf"
    output_path = ctx.output_path(filename)

    ctx.report_progress(55, "Generando PDF de autorización de pagos...")
    _build_pdf(snapshot, output_path)
    ctx.report_progress(90, "PDF generado correctamente.")
    return ctx.output_rel(filename)

